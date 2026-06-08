from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from yield_report.agent.spec_model import RunContext, SkillError, SkillResult
from yield_report.skills.data_analysis import tool
from yield_report.skills.data_analysis.daily_report_analysis import (
    _read_worksheet_rows,
    select_product_rows,
)
from yield_report.skills.data_analysis.models import DataAnalysisRequest


def _save_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _force_stale_dimension(path: Path, dimension: str = "A1:A1") -> None:
    temp_path = path.with_suffix(".tmp.xlsx")
    with ZipFile(path, "r") as source, ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(
                    rb'<dimension ref="[^"]+"',
                    f'<dimension ref="{dimension}"'.encode(),
                    data,
                    count=1,
                )
            target.writestr(item, data)
    temp_path.replace(path)


def _daily_yield(path: Path) -> Path:
    return _save_workbook(
        path,
        {
            "CT": [
                ["", "CT良率及不良率By月周天汇总报表V1.0"],
                [],
                ["", "ProductCode\n产品型号", "Operation\n站点", "Factory\n归属工厂", "DefectGroup\n不良分组", "", "M06", "6/1", "6/2", "6/3"],
                ["", "M678", "CT", "CT产出数", "", "", 100, 100, 100, 100],
                ["", "", "", "CT良率", "", "", 0.94, 0.95, 0.94, 0.93],
                ["", "", "", "CT产出数_MVI产出占比", "", "", 0.1, 0.2, 0.3, 0.4],
                ["", "", "", "Array_AD", "", "", 0.04, 0.04, 0.05, 0.06],
                ["", "", "", "Array_Line", "", "", 0.03, 0.03, 0.04, 0.05],
                ["", "", "", "", "Array_AD", "暗点", 0.04, 0.04, 0.05, 0.07],
            ]
        },
    )


def _target(path: Path) -> Path:
    return _save_workbook(
        path,
        {"各产品Group目标": [["项目代号", "Group", "6月"], ["M678", "Array_AD", 0.01], ["M678", "Array_Line", 0.02]]},
    )


def _ct_exception(path: Path) -> Path:
    reply = (
        "【异常原因】清洗能力不足\n"
        "【Inline监控】AT可监控\n"
        "【是否止血】是\n"
        "【影响范围】16LOT\n"
        "【改善措施】机台保养"
    )
    return _save_workbook(
        path,
        {
            "CT异常波动调查": [
                ["通报日期", "产品", "不良", "日良损", "当月", "工艺整合&工艺 回复的改善及挽救进展"],
                ["2026-05-20 10:00:00", "M678", "暗点", 0.03, 0.01, reply],
                ["2026-06-01 10:00:00", "M678", "群亮点", 0.02, 0.01, reply],
            ]
        },
    )


def _product() -> dict[str, object]:
    return {
        "product_type": "M678",
        "product": "M678\nLTPO",
        "report_date": "2026-06-01",
        "target_yield": 0.95,
        "actual_yield": 0.88,
        "is_qualified": False,
        "daily_gap": -0.07,
    }


def test_read_worksheet_rows_resets_stale_read_only_dimension(tmp_path: Path) -> None:
    daily_yield = _save_workbook(
        tmp_path / "daily_yield.xlsx",
        {
            "CT": [
                ["", "ProductCode", "Operation", "DefectGroup"],
                ["", "M678", "CT", "CT output"],
                ["", "", "", "CT yield"],
            ]
        },
    )
    _force_stale_dimension(daily_yield)

    workbook = load_workbook(daily_yield, read_only=True, data_only=True)
    try:
        rows = _read_worksheet_rows(workbook["CT"])
    finally:
        workbook.close()

    assert len(rows[0]) == 4
    assert len(select_product_rows(rows, "M678")) == 3


def test_data_analysis_daily_report_returns_four_sections(tmp_path: Path) -> None:
    result = tool.run(
        DataAnalysisRequest(
            analysis_kind="daily_report",
            report_date="2026-06-01",
            daily_report_products=[_product()],
            source_files={
                "daily_yield": _daily_yield(tmp_path / "daily_yield.xlsx"),
                "target_decomposition": _target(tmp_path / "target.xlsx"),
                "ct_exception": _ct_exception(tmp_path / "ct_exception.xlsx"),
            },
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=tmp_path / "output"),
    )

    assert result.success is True
    facts = result.data["daily_report_facts"]
    sections = facts["products"][0]["sections"]
    assert set(sections) == {"gap", "trend", "known_exception", "new_exception"}
    assert sections["gap"]["status"] == "emitted"
    assert sections["trend"]["status"] == "emitted"
    assert sections["known_exception"]["status"] == "emitted"
    assert sections["new_exception"]["status"] == "emitted"


def test_data_analysis_daily_report_uses_product_models_without_product_rows(tmp_path: Path) -> None:
    result = tool.run(
        DataAnalysisRequest(
            analysis_kind="daily_report",
            report_date="2026-06-01",
            product_models=["M678"],
            source_files={
                "daily_yield": _daily_yield(tmp_path / "daily_yield.xlsx"),
                "target_decomposition": _target(tmp_path / "target.xlsx"),
                "ct_exception": _ct_exception(tmp_path / "ct_exception.xlsx"),
            },
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=tmp_path / "output"),
    )

    assert result.success is True
    products = result.data["daily_report_facts"]["products"]
    assert products[0]["product"]["product_type"] == "M678"
    assert products[0]["sections"]["gap"]["status"] == "emitted"


def test_data_analysis_daily_report_calls_report_download_when_sources_missing(
    monkeypatch,
    tmp_path: Path,
) -> None:
    calls = []

    def fake_download(request, context):
        calls.append(request.report_type)
        return SkillResult(
            skill_name="report_download",
            success=False,
            summary="download failed",
            error=SkillError(
                code="report_download.execution.failed",
                message="download failed",
                recoverable=True,
            ),
        )

    monkeypatch.setattr(
        "yield_report.skills.data_analysis.daily_report_analysis.report_download_tool.run",
        fake_download,
    )

    result = tool.run(
        DataAnalysisRequest(
            analysis_kind="daily_report",
            report_date="2026-06-01",
            daily_report_products=[_product()],
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=tmp_path / "output"),
    )

    assert result.success is False
    assert result.error is not None
    assert result.error.code == "data_analysis.daily_report.blocked"
    assert calls
    assert result.data["daily_report_facts"]["downstream_results"]


def test_daily_report_analysis_passes_effective_daily_yield_end_date(
    monkeypatch,
    tmp_path: Path,
) -> None:
    calls = []

    monkeypatch.setattr(
        "yield_report.skills.data_analysis.daily_report_analysis.effective_daily_yield_end_date",
        lambda report_date: date(2026, 6, 1),
    )

    def fake_download(request, context):
        calls.append(request)
        return SkillResult(
            skill_name="report_download",
            success=False,
            summary="download failed",
            error=SkillError(
                code="report_download.execution.failed",
                message="download failed",
                recoverable=True,
            ),
        )

    monkeypatch.setattr(
        "yield_report.skills.data_analysis.daily_report_analysis.report_download_tool.run",
        fake_download,
    )

    tool.run(
        DataAnalysisRequest(
            analysis_kind="daily_report",
            report_date="2026-06-02",
            daily_report_products=[_product()],
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=tmp_path / "output"),
    )

    daily_yield_call = next(call for call in calls if call.report_type == "daily_yield")
    assert daily_yield_call.end_date == "2026-06-01"
