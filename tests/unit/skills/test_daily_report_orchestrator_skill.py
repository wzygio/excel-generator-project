from __future__ import annotations

import subprocess
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook

from yield_report.agent.registry import build_default_runtime
from yield_report.agent.spec_model import RunContext, SkillCall, TaskSpec
from yield_report.skills.daily_report import task0_task2_orchestrator as base_orchestrator_module
from yield_report.skills.daily_report import task0_task4_orchestrator as orchestrator_module
from yield_report.skills.daily_report import tool
from yield_report.skills.daily_report.models import DailyReportRequest


def _write_minimal_full_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data Packet"
    sheet.append(
        [
            "产品类型",
            "产品",
            "日期",
            "Top3 Group Gap",
            "Top3 Group Batch Info",
            "1.1 过货影响",
            "1.2 批次分析",
            "1.3 当日异常",
            "1.4 已知异常",
            "月度分析",
        ]
    )
    sheet.append(
        [
            "C522",
            "C522",
            "2026-06-15",
            "{Group：Array_Pixel}-{当日良率：89.56%}-{良率目标：92.00%}-{Gap：-2.44%}",
            "{Group：Array_Pixel}-{最新批次：26/06/15}-{批次良率：89.00%}-{批次情况：该Group最新批次有异常}",
            "1.1 Array_Pixel 过货影响",
            "1.2 批次分析",
            "1.3 当日异常",
            "【异常】已知异常",
            "月度良率说明",
        ]
    )
    upload = workbook.create_sheet("Sheet1")
    upload.append(["RN", "产品类型", "产品", "日期", "当日异常", "当日异常_HTML", "月度良率说明"])
    upload.append(
        [
            None,
            "C522",
            "C522",
            "2026-06-15",
            "1.1 Array_Pixel 过货影响\n1.2 批次分析\n1.3 当日异常\n【异常】已知异常",
            '<strong><font color="#FF0000">【异常】</font></strong>已知异常',
            "月度良率说明",
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_child_scripts(scripts_dir: Path) -> None:
    scripts_dir.mkdir(parents=True)
    for name in (
        "task0_report_download.py",
        "task1_gap_analysis.py",
        "task2_extract_anomalies.py",
        "task3_batch_month_analysis.py",
        "task4_daily_report_generation.py",
    ):
        (scripts_dir / name).write_text("# test script\n", encoding="utf-8")


def _write_resource_files(duty_workspace: Path, suffix: str) -> dict[str, Path]:
    resource_dir = duty_workspace / "resources" / suffix
    resource_dir.mkdir(parents=True)
    paths = {
        "daily_yield": resource_dir / "V3良率及不良率By月周天汇总报表.xlsx",
        "batch_yield": resource_dir / "V3良率及不良率By批次汇总报表.xlsx",
        "target": duty_workspace / "resources" / "decrypted_files" / "良率目标表.xlsx",
    }
    for path in paths.values():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"PK\x03\x04placeholder")
    (duty_workspace / "resources" / "良率目标表.xlsx").write_bytes(b"\x00\x00encrypted")
    return paths


def test_daily_report_skill_runs_task0_task4_orchestrator(monkeypatch, tmp_path: Path) -> None:
    duty_workspace = tmp_path / "duty"
    _write_child_scripts(duty_workspace / "scripts")
    resource_paths = _write_resource_files(duty_workspace, "20260615-16：00")

    output_dir = tmp_path / "run" / "outputs"
    output_path = output_dir / "generated.xlsx"
    calls: list[list[str]] = []

    def fake_run(command, **kwargs):
        command_text = [str(part) for part in command]
        calls.append(command_text)
        if "task0_report_download.py" in command_text[1]:
            _write_minimal_full_workbook(output_path)
        return subprocess.CompletedProcess(
            args=command,
            returncode=0,
            stdout='{"ok": true}',
            stderr="",
        )

    monkeypatch.setattr(subprocess, "run", fake_run)
    monkeypatch.setattr(orchestrator_module, "_cleanup_hidden_excel_processes", lambda: None)
    monkeypatch.setattr(
        orchestrator_module, "_wait_for_file_unlock", lambda path, timeout_seconds: None
    )

    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-15",
            output_dir=output_dir,
            output_name="generated.xlsx",
            orchestrator_workspace=duty_workspace,
            orchestrator_now="2026-06-15 16:00",
            run_inspection=False,
            task2_max_anomaly_row=2365,
            source_files={
                "target_decomposition": str(duty_workspace / "resources" / "良率目标表.xlsx")
            },
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=output_dir),
    )

    assert result.success is True
    assert result.skill_name == "daily_report"
    assert [Path(call[1]).name for call in calls] == [
        "task0_report_download.py",
        "task1_gap_analysis.py",
        "task1_gap_analysis.py",
        "task2_extract_anomalies.py",
        "task3_batch_month_analysis.py",
        "task4_daily_report_generation.py",
    ]
    assert [call[0] for call in calls] == [
        "python",
        "python",
        "python",
        "python",
        "python",
        "python",
    ]
    assert calls[1][2:] == ["--self-test"]
    assert calls[2][2:] == [str(output_path), "--write", "--now", "2026-06-15 16:00"]
    assert calls[3][2:5] == ["--source", str(output_path), "--write"]
    assert calls[3][5:7] == ["--daily-yield", str(resource_paths["daily_yield"])]
    assert calls[3][-2:] == ["--max-anomaly-row", "2365"]
    assert calls[4][2:5] == ["--source", str(output_path), "--write"]
    assert calls[4][5:9] == [
        "--batch-report",
        str(resource_paths["batch_yield"]),
        "--period-report",
        str(resource_paths["daily_yield"]),
    ]
    assert calls[4][-2:] == ["--target", str(resource_paths["target"])]
    assert calls[5][2:] == ["--source", str(output_path), "--write"]
    assert result.artifacts[0].kind == "excel"
    assert Path(result.artifacts[0].path) == output_path
    assert result.data["workflow"] == [
        "basic-preparation",
        "task1-gap-analysis",
        "task2-extract-anomalies",
        "task3-batch-month-analysis",
        "task4-daily-report-generation",
    ]
    verification = result.data["verification"]
    assert verification["data_packet"]["row_count"] == 1
    assert verification["data_packet"]["nonblank_counts"]["1.2 批次分析"] == 1
    assert verification["data_packet"]["nonblank_counts"]["月度分析"] == 1
    assert verification["sheet1"]["nonblank_counts"]["当日异常"] == 1
    assert verification["sheet1"]["nonblank_counts"]["当日异常_HTML"] == 1
    assert verification["sheet1"]["nonblank_counts"]["月度良率说明"] == 1


def test_agent_runtime_runs_daily_report_orchestrator(monkeypatch, tmp_path: Path) -> None:
    duty_workspace = tmp_path / "duty"
    _write_child_scripts(duty_workspace / "scripts")

    output_dir = tmp_path / "specs" / "runs" / "run-daily" / "outputs"
    output_path = output_dir / "runtime-generated.xlsx"
    calls: list[list[str]] = []

    def fake_run(command, **kwargs):
        command_text = [str(part) for part in command]
        calls.append(command_text)
        if "task0_report_download.py" in command_text[1]:
            _write_minimal_full_workbook(output_path)
        return subprocess.CompletedProcess(command, 0, stdout='{"ok": true}', stderr="")

    monkeypatch.setattr(subprocess, "run", fake_run)
    monkeypatch.setattr(orchestrator_module, "_cleanup_hidden_excel_processes", lambda: None)
    monkeypatch.setattr(
        orchestrator_module, "_wait_for_file_unlock", lambda path, timeout_seconds: None
    )

    spec = TaskSpec(
        run_id="run-daily",
        user_goal="生成 2026-06-15 16:00 良率日报",
        workflow=[
            SkillCall(
                id="generate_daily_report",
                skill="daily_report",
                input={
                    "report_date": "2026-06-15",
                    "output_dir": str(output_dir),
                    "output_name": "runtime-generated.xlsx",
                    "orchestrator_workspace": str(duty_workspace),
                    "orchestrator_now": "2026-06-15 16:00",
                    "run_inspection": False,
                },
            )
        ],
    )
    context = RunContext(run_id="run-daily", workspace=tmp_path, output_dir=output_dir)

    results = build_default_runtime().run_spec(spec, context)

    assert results[0].success is True
    assert Path(results[0].data["output_file"]) == output_path
    assert [Path(call[1]).name for call in calls] == [
        "task0_report_download.py",
        "task1_gap_analysis.py",
        "task1_gap_analysis.py",
        "task2_extract_anomalies.py",
        "task3_batch_month_analysis.py",
        "task4_daily_report_generation.py",
    ]
    assert (tmp_path / "specs" / "runs" / "run-daily" / "run_summary.json").exists()


def test_daily_report_continues_when_task0_times_out_after_workbook_ready(
    monkeypatch,
    tmp_path: Path,
) -> None:
    duty_workspace = tmp_path / "duty"
    _write_child_scripts(duty_workspace / "scripts")

    output_dir = tmp_path / "run" / "outputs"
    output_path = output_dir / "timeout-generated.xlsx"
    calls: list[list[str]] = []

    def fake_run(command, **kwargs):
        command_text = [str(part) for part in command]
        calls.append(command_text)
        if "task0_report_download.py" in command_text[1]:
            _write_minimal_full_workbook(output_path)
            raise subprocess.TimeoutExpired(
                cmd=command,
                timeout=kwargs.get("timeout") or 1,
                output="task0 partial stdout",
                stderr="task0 hung during cleanup",
            )
        return subprocess.CompletedProcess(command, 0, stdout='{"ok": true}', stderr="")

    monkeypatch.setattr(subprocess, "run", fake_run)
    monkeypatch.setattr(orchestrator_module, "_cleanup_hidden_excel_processes", lambda: None)
    monkeypatch.setattr(
        orchestrator_module, "_wait_for_file_unlock", lambda path, timeout_seconds: None
    )

    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-15",
            output_dir=output_dir,
            output_name="timeout-generated.xlsx",
            orchestrator_workspace=duty_workspace,
            orchestrator_now="2026-06-15 16:00",
            run_inspection=False,
            task0_timeout_seconds=1,
        ),
        RunContext(run_id="run-timeout", workspace=tmp_path, output_dir=output_dir),
    )

    assert result.success is True
    assert [Path(call[1]).name for call in calls] == [
        "task0_report_download.py",
        "task1_gap_analysis.py",
        "task1_gap_analysis.py",
        "task2_extract_anomalies.py",
        "task3_batch_month_analysis.py",
        "task4_daily_report_generation.py",
    ]
    assert result.data["steps"][0]["returncode"] == orchestrator_module.TASK0_TIMEOUT_RETURNCODE
    assert "timed out after 1 seconds" in result.warnings[0]


def test_daily_report_skill_compares_generated_workbook_to_reference(
    monkeypatch,
    tmp_path: Path,
) -> None:
    duty_workspace = tmp_path / "duty"
    _write_child_scripts(duty_workspace / "scripts")

    output_dir = tmp_path / "output"
    output_path = output_dir / "generated.xlsx"
    reference_path = tmp_path / "reference.xlsx"
    _write_minimal_full_workbook(reference_path)

    def fake_run(command, **kwargs):
        command_text = [str(part) for part in command]
        if "task0_report_download.py" in command_text[1]:
            _write_minimal_full_workbook(output_path)
        return subprocess.CompletedProcess(command, 0, stdout='{"ok": true}', stderr="")

    monkeypatch.setattr(subprocess, "run", fake_run)

    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-15",
            output_dir=output_dir,
            output_name="generated.xlsx",
            orchestrator_workspace=duty_workspace,
            orchestrator_now="2026-06-15 16:00",
            run_inspection=False,
            reference_workbook=reference_path,
        ),
        RunContext(run_id="run-compare", workspace=tmp_path, output_dir=output_dir),
    )

    assert result.success is True
    assert result.data["comparison"]["match"] is True
    assert result.data["comparison"]["differences"] == []


def test_daily_report_verification_falls_back_to_excel_com_for_encrypted_workbooks(
    monkeypatch,
    tmp_path: Path,
) -> None:
    encrypted_path = tmp_path / "encrypted.xlsx"
    encrypted_path.write_text("not a zip workbook", encoding="utf-8")

    def raise_bad_zip(*args, **kwargs):
        raise BadZipFile("File is not a zip file")

    monkeypatch.setattr(base_orchestrator_module, "load_workbook", raise_bad_zip)
    monkeypatch.setattr(
        base_orchestrator_module,
        "_read_sheet_values_with_excel_com",
        lambda path, sheet_name: _fake_excel_com_values(sheet_name),
    )

    verification = orchestrator_module.verify_daily_report_workbook(encrypted_path)

    assert verification["data_packet"]["row_count"] == 1
    assert verification["data_packet"]["nonblank_counts"] == {
        "1.1 过货影响": 1,
        "1.2 批次分析": 1,
        "1.3 当日异常": 1,
        "1.4 已知异常": 1,
        "月度分析": 1,
    }
    assert verification["sheet1"]["nonblank_counts"] == {
        "当日异常": 1,
        "当日异常_HTML": 1,
        "月度良率说明": 1,
    }


def _fake_excel_com_values(sheet_name: str) -> list[list[object]]:
    if sheet_name == "Data Packet":
        return [
            [
                "产品类型",
                "1.1 过货影响",
                "1.2 批次分析",
                "1.3 当日异常",
                "1.4 已知异常",
                "月度分析",
            ],
            ["C522", "impact", "batch", "today", "known", "month"],
        ]
    if sheet_name == "Sheet1":
        return [
            ["产品类型", "产品", "日期", "当日异常", "当日异常_HTML", "月度良率说明"],
            ["C522", "C522", "2026-06-15", "plain", "<strong>html</strong>", "month"],
        ]
    raise ValueError(sheet_name)
