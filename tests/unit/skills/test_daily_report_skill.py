from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from yield_report.agent.spec_model import RunContext, SkillError, SkillResult
from yield_report.skills.daily_report import tool
from yield_report.skills.daily_report.implementation import (
    extract_exception_records,
    extract_shipped_products,
    extract_target_rates,
)
from yield_report.skills.daily_report.models import DailyReportRequest


def _save_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _template(path: Path) -> Path:
    return _save_workbook(
        path,
        {
            "sheet1": [
                ["", "title"],
                [],
                ["", "", "产品类型", "产品", "日期", "当日异常_HTML", "当日异常"],
                ["", "", "OLD", "OLD", "OLD", "OLD", "OLD"],
                ["", "", "STALE", "STALE", "STALE", "STALE", "STALE"],
            ]
        },
    )


def _spotfire(path: Path) -> Path:
    return _save_workbook(
        path,
        {
            "Sheet1": [
                ["产品类型", "产品", "日期", "目标", "良率", "是否达标", "当日GAP", "CT"],
                [
                    "产品类型",
                    "产品",
                    "日期",
                    "目标",
                    "良率",
                    "是否达标",
                    "当日GAP",
                    "投入数",
                    "ratio",
                    "当日异常",
                    "已HL异常",
                    "TOP3-GROUP",
                ],
                ["M678", "M678\nLTPO", 20260601.0, 0.95, 0.88, "不达标", -0.07, 100, 0.88, "", "", ""],
                ["M999", "M999\nLTPO", 20260601.0, 0.95, 0.96, "达标", 0.01, 100, 0.96, "", "", ""],
            ]
        },
    )


def _daily_yield(path: Path) -> Path:
    return _save_workbook(
        path,
        {
            "CT": [
                ["", "CT良率及不良率By月周天汇总报表V1.0"],
                [],
                [
                    "",
                    "ProductCode\n产品型号",
                    "Operation\n站点",
                    "Factory\n归属工厂",
                    "DefectGroup\n不良分组",
                    "",
                    "M06",
                    "6/1",
                    "6/2",
                    "6/3",
                ],
                ["", "M678", "CT", "CT良率", "", "", 0.94, 0.95, 0.94, 0.93],
                ["", "", "", "CT产出数_MVI产出占比", "", "", 0.1, 0.2, 0.3, 0.4],
                ["", "", "", "Array_AD", "", "", 0.04, 0.04, 0.05, 0.06],
                ["", "", "", "Array_Line", "", "", 0.03, 0.03, 0.04, 0.05],
                ["", "", "", "OLED_Mura", "", "", 0.01, 0.01, 0.02, 0.03],
                ["", "", "", "", "Array_AD", "暗点", 0.04, 0.04, 0.05, 0.07],
                ["", "", "", "", "Array_Line", "线不良", 0.03, 0.03, 0.04, 0.02],
            ]
        },
    )


def _target(path: Path) -> Path:
    return _save_workbook(
        path,
        {
            "各产品Group目标": [
                ["项目代号", "Group", "6月"],
                ["M678", "Array_AD", 0.01],
                ["M678", "Array_Line", 0.02],
                ["M678", "OLED_Mura", 0.01],
            ]
        },
    )


def _ct_exception(path: Path) -> Path:
    reply = (
        "【异常原因】清洗能力不足\n"
        "【Inline监控】AT可监控\n"
        "【是否止血】是\n"
        "【影响范围】16LOT\n"
        "【改善措施】机台保养"
    )
    return _save_workbook(
        path,
        {
            "CT异常波动调查": [
                ["no", "周别", "通报日期", "产品", "不良", "不良站点", "日良损", "当月", "当周", "本批次", "异常通报", "工艺整合&工艺 回复的改善及挽救进展"],
                [1, "W22", "2026-05-20 10:00:00+00:00", "M678", "暗点", "CT", 0.03, 0.01, 0.02, 0.04, "", reply],
                [2, "W23", "2026-06-01 10:00:00+00:00", "M678", "群亮点", "CT", 0.02, 0.01, 0.02, 0.03, "", reply],
                [3, "W23", "2026-06-01 10:00:00+00:00", "M999", "暗点", "CT", 0.02, 0.01, 0.02, 0.03, "", reply],
            ]
        },
    )


def test_daily_report_request_accepts_extended_fields(tmp_path: Path) -> None:
    request = DailyReportRequest(
        report_date="2026-06-01",
        product_models=["M678"],
        source_files={"spotfire": tmp_path / "spotfire.xlsx"},
        output_dir=tmp_path / "output",
        sections=[" gap ", "", "trend"],
    )

    assert request.product_models == ["M678"]
    assert request.sections == ["gap", "trend"]
    assert request.source_files["spotfire"] == tmp_path / "spotfire.xlsx"


def test_extract_shipped_products_filters_spotfire_shape(tmp_path: Path) -> None:
    path = _spotfire(tmp_path / "spotfire.xlsx")
    workbook = load_workbook(path, data_only=True)
    try:
        rows = [list(row) for row in workbook["Sheet1"].iter_rows(values_only=True)]
    finally:
        workbook.close()

    products = extract_shipped_products(rows)

    assert [item.product_type for item in products] == ["M678", "M999"]
    assert products[0].report_date == "2026-06-01"
    assert products[0].is_qualified is False


def test_target_rates_use_requested_month(tmp_path: Path) -> None:
    path = _target(tmp_path / "target.xlsx")
    workbook = load_workbook(path, data_only=True)
    try:
        rows = [list(row) for row in workbook["各产品Group目标"].iter_rows(values_only=True)]
    finally:
        workbook.close()

    rates = extract_target_rates(rows, "6月")

    assert rates[("M678", "Array_AD")] == 0.01


def test_exception_records_match_known_and_new_windows(tmp_path: Path) -> None:
    path = _ct_exception(tmp_path / "ct_exception.xlsx")
    workbook = load_workbook(path, data_only=True)
    try:
        rows = [list(row) for row in workbook["CT异常波动调查"].iter_rows(values_only=True)]
    finally:
        workbook.close()

    known = extract_exception_records(
        rows=rows,
        product_model="M678",
        report_date=__import__("datetime").date(2026, 6, 1),
        defect_codes=["暗点"],
        mode="known",
    )
    new = extract_exception_records(
        rows=rows,
        product_model="M678",
        report_date=__import__("datetime").date(2026, 6, 1),
        defect_codes=None,
        mode="new",
    )

    assert [item.defect_code for item in known] == ["暗点"]
    assert [item.defect_code for item in new] == ["群亮点"]
    assert known[0].exception_reason == "清洗能力不足"


def test_daily_report_skill_generates_excel_json_and_markdown(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-01",
            product_models=["M678"],
            template_ref=_template(tmp_path / "template.xlsx"),
            output_dir=output_dir,
            source_files={
                "spotfire": _spotfire(tmp_path / "spotfire.xlsx"),
                "daily_yield": _daily_yield(tmp_path / "daily_yield.xlsx"),
                "target_decomposition": _target(tmp_path / "target.xlsx"),
                "ct_exception": _ct_exception(tmp_path / "ct_exception.xlsx"),
            },
            output_name="daily_report.xlsx",
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=output_dir),
    )

    assert result.success is True
    assert {artifact.kind for artifact in result.artifacts} == {"excel", "json", "markdown"}
    assert result.data["products"][0]["gap_top_items"][0]["defect_group"] == "Array_AD"
    assert result.data["products"][0]["sections"]["gap"]["status"] == "emitted"
    assert result.data["products"][0]["sections"]["trend"]["status"] == "emitted"
    assert result.data["products"][0]["sections"]["known_exception"]["status"] == "emitted"
    assert result.data["products"][0]["sections"]["new_exception"]["status"] == "emitted"
    assert result.data["products"][0]["trend"]["is_declining"] is True
    assert result.data["products"][0]["trend"]["mvi_share_increasing"] is True
    assert result.data["downstream_results"] == []

    output_path = next(artifact.path for artifact in result.artifacts if artifact.kind == "excel")
    workbook = load_workbook(output_path, data_only=True)
    try:
        sheet = workbook["sheet1"]
        assert sheet["C4"].value == "M678"
        assert sheet["D4"].value == "M678\nLTPO"
        assert sheet["E4"].value == "2026-06-01"
        assert "2.1 当日Gap影响：Array_AD" in sheet["G4"].value
        assert "2.2 连续三天下降原因：6/1~6/3 CT良率持续下降" in sheet["G4"].value
        assert sheet["C5"].value is None
    finally:
        workbook.close()


def test_daily_report_skill_returns_structured_missing_file_error(tmp_path: Path) -> None:
    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-01",
            source_files={
                "spotfire": tmp_path / "missing.xlsx",
            },
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=tmp_path / "output"),
    )

    assert result.success is False
    assert result.error is not None
    assert result.error.code == "daily_report.file.missing_required"


def test_daily_report_skill_blocks_when_analysis_sources_missing(
    monkeypatch,
    tmp_path: Path,
) -> None:
    def fake_download(request, context):
        return SkillResult(
            skill_name="report_download",
            success=False,
            summary=f"missing {request.report_type}",
            error=SkillError(
                code="report_download.execution.failed",
                message=f"missing {request.report_type}",
                recoverable=True,
            ),
        )

    monkeypatch.setattr(
        "yield_report.skills.data_analysis.daily_report_analysis.report_download_tool.run",
        fake_download,
    )
    output_dir = tmp_path / "output"
    result = tool.run(
        DailyReportRequest(
            report_date="2026-06-01",
            product_models=["M678"],
            template_ref=_template(tmp_path / "template.xlsx"),
            output_dir=output_dir,
            source_files={"spotfire": _spotfire(tmp_path / "spotfire.xlsx")},
            output_name="daily_report.xlsx",
        ),
        RunContext(run_id="run-1", workspace=tmp_path, output_dir=output_dir),
    )

    assert result.success is False
    assert result.error is not None
    assert result.error.code == "daily_report.analysis.blocked"
    assert result.data["blocked_sections"]
    assert result.data["downstream_results"]
    assert not (output_dir / "daily_report.xlsx").exists()

    all_log = output_dir / "logs" / "all.log"
    daily_report_error_log = output_dir / "logs" / "skills_daily_report" / "error.log"
    assert "Daily report generation blocked" in all_log.read_text(encoding="utf-8")
    error_log_content = daily_report_error_log.read_text(encoding="utf-8")
    assert "Daily report generation blocked" in error_log_content
    assert "error_code=daily_report.analysis.blocked" in error_log_content
