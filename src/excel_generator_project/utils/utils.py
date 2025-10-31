# tests/test_data_processor.py
import sys
from pathlib import Path
# 动态添加项目根目录到 sys.path
sys.path.append(str(Path(__file__).parent.parent))

import logging, re
import pprint, shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from typing import Optional

from excel_generator_project.config import BASE_DIR, DATA_DIR, LOG_DIR

class Utils:
    @staticmethod
    def setup_logging(log_filename: str = "app.log"):
        """初始化日志系统，输出到文件(覆盖模式)和控制台。"""
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        log_filepath = LOG_DIR / log_filename

        log_format = '%(asctime)s - %(levelname)s - [%(module)s] - %(message)s'
        log_date_format = '%Y-%m-%d %H:%M:%S'

        root_logger = logging.getLogger()
        root_logger.setLevel(logging.INFO) # 或 DEBUG

        # 清除旧处理器 (保持不变)
        if root_logger.hasHandlers():
            # 移除处理器并关闭它们，确保文件句柄被释放
            for handler in root_logger.handlers[:]:
                handler.close()
                root_logger.removeHandler(handler)

        # --- [核心修改] 文件处理器，使用 mode='w' ---
        try:
            file_handler = logging.FileHandler(log_filepath, mode='w', encoding='utf-8') # <-- 添加 mode='w'
            file_handler.setFormatter(logging.Formatter(log_format, datefmt=log_date_format))
            root_logger.addHandler(file_handler)
        except Exception as e:
             # 如果文件无法以写入模式打开（例如权限问题），至少还能输出到控制台
             logging.error(f"无法以写入模式打开日志文件 '{log_filepath}': {e}")


        # --- 控制台处理器 (保持不变) ---
        # 避免重复添加控制台处理器（如果之前被清除了）
        if not any(isinstance(h, logging.StreamHandler) for h in root_logger.handlers):
             console_handler = logging.StreamHandler(sys.stdout)
             console_handler.setFormatter(logging.Formatter(log_format, datefmt=log_date_format))
             root_logger.addHandler(console_handler)

        logging.info(f"日志系统已初始化，将同时输出到文件 '{log_filepath}' (覆盖模式) 和控制台。")


    @staticmethod
    def setup_test_environment(log_filename: str, log_level=logging.INFO):
        """
        (已重构) 配置日志系统，实现文件和控制台的双路输出。
        """
        # 1. 获取根日志记录器
        root_logger = logging.getLogger()
        root_logger.setLevel(log_level)

        # 2. 清除任何已存在的处理器，防止日志重复输出
        if root_logger.hasHandlers():
            root_logger.handlers.clear()

        # 3. 创建一个统一的日志格式
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - [%(module)s] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        # 4. 创建处理器 1: 写入到文件
        log_dir = Path("./logs") # 建议将所有日志文件统一存放在logs目录
        log_dir.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_dir / log_filename, mode='w', encoding='utf-8')
        file_handler.setLevel(log_level)
        file_handler.setFormatter(formatter)
        root_logger.addHandler(file_handler)

        # 5. 创建处理器 2: 输出到标准错误流 (stderr)
        #    这是让 subprocess 能够捕获到日志的关键
        stream_handler = logging.StreamHandler(sys.stderr)
        stream_handler.setLevel(log_level)
        stream_handler.setFormatter(formatter)
        root_logger.addHandler(stream_handler)

        logging.info(f"日志系统已初始化，将同时输出到文件 '{log_filename}' 和控制台。")

    @staticmethod
    def extract_product_models(model_config: dict, custom_path: Optional[Path] = None) -> list:
        """
        (新的私有辅助方法) 执行提取tila_target的第一步：提取产品型号列表。
        """
        # --- 1. 解析并验证配置 ---
        if custom_path:
            source_file = custom_path
        else:
            source_file = model_config.get("source_file")
            if not source_file:
                logging.error("任务 'daily_yield_change' 配置中缺少 'source_file' 参数。")
                return []
        source_path = Path(source_file)
        
        sheet_name = model_config.get("sheet_name")
        pattern = model_config.get("pattern")
        start_cell = model_config.get("start_cell")
        step = model_config.get("step")
        end_row = model_config.get("end_row")

        if not all([source_path.is_file(), sheet_name, pattern, start_cell, isinstance(step, int), isinstance(end_row, int)]):
            logging.error(" 步骤1 'step1_product_models' 配置不完整、类型错误或源文件不存在。")
            return [] # 返回一个空列表

        # --- 2. 加载源数据文件 ---
        source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        if sheet_name not in source_wb.sheetnames:
            logging.error(f" 工作表 '{sheet_name}' 在文件 '{source_path.name}' 中未找到。")
            return []
        source_ws = source_wb[sheet_name]

        # --- 3. 迭代提取数据 ---
        col_letter = ''.join(filter(str.isalpha, start_cell or "A"))
        current_row = int(''.join(filter(str.isdigit, start_cell or "2")))
        
        models_list = []
        logging.info(f" 第一步：在 {source_path.name}[{sheet_name}] 中提取产品型号...")
        while current_row <= (end_row or float('inf')):
            cell_address = f"{col_letter}{current_row}"
            cell_value = source_ws[cell_address].value
            
            # --- 新增：边界条件判断逻辑 ---
            # 将 None 和空的字符串都视为空
            if cell_value is None or str(cell_value).strip() == "":
                break
            # --- 边界条件判断结束 ---

            if isinstance(cell_value, str) and cell_value.strip(): # 确保单元格是字符串且不为空
                match = re.search(str(pattern), cell_value)
            if match:
                model_name = match.group(1).strip()
                models_list.append(model_name)
                logging.info(f" 在单元格 {cell_address} 提取到型号: {model_name}")
            
            current_row += step or 4
        
        return models_list

    @staticmethod
    def get_safe_source_path(job_config: dict, 
                             path_key: str = "source_path",
                             file_key: str = "source_file") -> Optional[Path]:
        """
        (已重构) 增加了对Excel锁定文件的自动处理。
        """
        source_directory = job_config.get(path_key)
        file_name = job_config.get(file_key)
        if not source_directory or not file_name:
            logging.error(f"   任务配置中缺少必要的目录路径键 '{path_key}'。")
            return None

        initial_target_path = Path(source_directory) / file_name

        # 在复制前，自动解析并净化路径
        final_path_to_copy = Utils.resolve_lock_file(initial_target_path)
        if not final_path_to_copy.is_file():
            logging.error(f"   最终确定的源文件不存在或不是一个文件: '{final_path_to_copy}'")
            return None
            
        # 复制净化后的、最终确定的文件
        local_cache_dir = BASE_DIR / DATA_DIR
        local_cache_dir.mkdir(parents=True, exist_ok=True)
        
        return Utils.get_local_copy(final_path_to_copy, local_cache_dir)

    @staticmethod
    def get_local_copy(source_path, local_cache_dir) -> Path | None:
        """
        (新方法) 将源文件安全地复制到本地缓存目录，并返回本地文件的路径。

        Args:
            source_path_str (str): 原始文件路径（可以是网络路径）。
            local_cache_dir_str (str): 用于存放副本的本地目录路径。

        Returns:
            Path | None: 如果复制成功，返回本地副本的Path对象；否则返回None。
        """
        try:

            # 1. 检查源文件是否存在
            if not source_path.is_file():
                logging.error(f"    [Utils] 源文件不存在: {source_path}")
                return None

            # 2. 确保本地缓存目录存在
            local_cache_dir.mkdir(parents=True, exist_ok=True)
            
            # 3. 构建本地副本的目标路径
            destination_path = local_cache_dir / source_path.name
            if destination_path == source_path:
                logging.info(f"    [Utils] 本地副本路径与源文件路径相同: {destination_path}")
                return source_path

            # 4. 执行复制操作
            logging.info(f"    [Utils] 正在将 '{source_path.name}' 复制到本地...")
            # 使用 shutil.copy2 可以顺便复制文件的元数据（如修改时间）
            shutil.copy2(source_path, destination_path)
            logging.info(f"      -> 本地副本已创建于: {destination_path}")

            return destination_path

        except Exception as e:
            logging.error(f"    [Utils] 复制文件时发生错误: {e}", exc_info=True)
            return None
        
    @staticmethod
    def parse_fragments_flat(text_content: str, definitions: list[dict], default_style_name: str) -> list[tuple[str, str]]:
        """
        [策略一：扁平解析 - 最终版] 使用 finditer 稳定地处理单层标题。
        此版本逻辑与分层解析器同源，确保健壮性。
        """
        if not text_content or not str(text_content).strip():
            return []
        if not definitions:
            return [(text_content, default_style_name)]

        # 将所有扁平规则的 pattern 用 | (或) 连接起来
        all_patterns = [defn['pattern'] for defn in definitions]
        combined_pattern = re.compile("|".join(all_patterns), flags=re.MULTILINE)
        
        fragments = []
        last_end = 0
        
        # 使用 finditer 遍历所有匹配到的标题
        for match in combined_pattern.finditer(text_content):
            # 1. 添加上一个标题结尾到当前标题开头之间的“内容”
            content_before = text_content[last_end:match.start()]
            if content_before:
                fragments.append((content_before, default_style_name))
            
            # 2. 添加匹配到的“标题”本身
            title_text = match.group(0)
            # 简化处理：所有匹配到的标题都使用第一个定义的样式
            fragments.append((title_text, definitions[0]['style']))
            
            # 3. 更新扫描位置
            last_end = match.end()

        # 4. 添加最后一个标题之后的所有剩余“内容”
        remaining_content = text_content[last_end:]
        if remaining_content:
            fragments.append((remaining_content, default_style_name))
            
        # 5. 如果全文都找不到任何标题，则视作一整段内容
        if not fragments and text_content:
            fragments.append((text_content, default_style_name))
        
        return fragments

    @staticmethod
    def parse_fragments_hierarchical(text_content: str, definitions: list[dict], default_style_name: str) -> list[tuple[str, str]]:
        """
        [策略二：分层解析] 完整保留并调用你的多级 finditer 解析引擎。
        """
        if not text_content or not str(text_content).strip(): return []
        if not definitions: return [(text_content, default_style_name)]

        # --- 这是你提供的、功能完备的多级解析器的入口 ---
        final_fragments = []
        l1_def = definitions[0]
        l1_pattern = re.compile(l1_def['pattern'], flags=re.MULTILINE)
        
        last_end = 0
        # Pass 1: 扫描一级标题，将文本分为大的章节
        for l1_match in l1_pattern.finditer(text_content):
            content_before = text_content[last_end:l1_match.start()]
            if content_before: final_fragments.append((content_before, default_style_name))

            l1_title_text = l1_match.group(0)
            final_fragments.append((l1_title_text, l1_def['style']))
            
            chapter_start = l1_match.end()
            next_l1_match = next(l1_pattern.finditer(text_content, pos=chapter_start), None)
            chapter_end = next_l1_match.start() if next_l1_match else len(text_content)
            chapter_content = text_content[chapter_start:chapter_end]
            
            if l1_def.get('sub_definitions'):
                final_fragments.extend(
                Utils._parse_chapter(chapter_content, l1_def['sub_definitions'], l1_title_text, default_style_name)
                )
            elif chapter_content:
                final_fragments.append((chapter_content, default_style_name))

            last_end = chapter_end

        final_content = text_content[last_end:]
        if final_content: final_fragments.append((final_content, default_style_name))

        if not final_fragments:
            return [(text_content, default_style_name)]
        
        return final_fragments

 
    
    # --- 需要在 Utils 类外部或内部定义这个新的辅助函数 ---
    @staticmethod
    def _parse_chapter(chapter_text: str, l2_definitions: list, parent_title_text: str, default_style: str):
        fragments = []
        l2_def = l2_definitions[0]
        l2_pattern = re.compile(l2_def['pattern'])
        
        last_end = 0
        # Pass 2: 扫描二级标题
        for l2_match in l2_pattern.finditer(chapter_text):
            content_before = chapter_text[last_end:l2_match.start()]
            
            l2_title_text = l2_match.group(0)
            
            # 确定二级标题的样式
            chosen_l2_style = l2_def['default_style']
            for cond in l2_def.get('conditional_styles', []):
                if cond['if_parent_contains'] in parent_title_text:
                    chosen_l2_style = cond['style']
                    break
            
            # Pass 3: 在二级标题前的文本中找三级标题
            if l2_def.get('sub_definitions'):
                fragments.extend(
                    Utils._parse_section(content_before, l2_def['sub_definitions'], default_style)
                )
            else:
                if content_before: fragments.append((content_before, default_style))

            fragments.append((l2_title_text, chosen_l2_style))
            last_end = l2_match.end()

        # 处理最后一个二级标题后的内容（这里面也可能包含三级标题）
        final_content = chapter_text[last_end:]
        if l2_def.get('sub_definitions'):
            fragments.extend(
                Utils._parse_section(final_content, l2_def['sub_definitions'], default_style)
            )
        else:
            if final_content: fragments.append((final_content, default_style))
        
        return fragments

    @staticmethod
    def _parse_section(section_text: str, l3_definitions: list, default_style: str):
        fragments = []
        l3_def = l3_definitions[0]
        l3_pattern = re.compile(l3_def['pattern'])
        
        last_end = 0
        # Pass 3: 扫描三级标题
        for l3_match in l3_pattern.finditer(section_text):
            content_before = section_text[last_end:l3_match.start()]
            if content_before: fragments.append((content_before, default_style))
            
            title_text = l3_match.group(0)
            fragments.append((title_text, l3_def['style']))
            last_end = l3_match.end()

        final_content = section_text[last_end:]
        if final_content: fragments.append((final_content, default_style))
        return fragments


    @staticmethod
    def manual_verification_prompt(output_path: str):
        """提示进行人工验证。"""
        print("\n" + "="*50)
        print(" ✅ 自动化测试执行完毕！")
        print(f"    最终文件已生成于{output_path}，请手动打开该文件，进行最终验证：")
        print("    请检查所有在 config.yaml 中定义的任务是否都已执行。")
        print("="*50)

    @staticmethod
    def process_text_with_title_styling(
        text_content: str,
        definitions: list[dict],
        default_style_name: str = 'content_style'
    ) -> list[tuple[str, str]]:
        """
        [最终版-递归扫描] 根据可嵌套的定义规则，通过递归的 finditer 扫描，精确解析多层标题。
        此版本基于之前成功的 finditer 机制，并扩展以支持层级和条件化样式。
        """
        
        # 这是一个内部递归函数，是整个新引擎的核心
        def recursive_parser(text: str, current_definitions: list[dict], parent_title_text: str = ""):
            fragments = []
            
            # 1. 为当前层级的所有标题规则，构建一个组合正则表达式
            patterns = [f"(?P<{defn['name']}>{defn['pattern']})" for defn in current_definitions]
            if not patterns: # 递归的“出口”：如果没有更多子规则，则剩余文本都是内容
                if text:
                    fragments.append((text, default_style_name))
                return fragments

            combined_pattern = re.compile("|".join(patterns))
            name_to_defn_map = {defn['name']: defn for defn in current_definitions}
            
            last_end = 0
            # 2. 使用 finditer 扫描当前文本块中的所有本级标题
            for match in combined_pattern.finditer(text):
                # a. 上一个标题的结尾到当前标题的开头，是“中间内容”
                intermediate_content = text[last_end:match.start()]
                
                # b. 处理当前匹配到的标题本身
                title_text = match.group(0)
                matched_name = match.lastgroup
                current_defn = name_to_defn_map[matched_name]

                # c. 递归地解析“中间内容”
                # 我们需要找到这个“中间内容”应该使用哪套子规则来解析
                # 在这个设计中，它属于上一个标题的子内容
                # (为了简化，我们假设只有标题后的内容才包含子标题，标题前的内容都是正文)
                if intermediate_content:
                    fragments.append((intermediate_content, default_style_name))

                # d. 确定当前标题的样式（包含条件化逻辑）
                chosen_style = current_defn.get('default_style') or current_defn.get('style')
                if parent_title_text: # 只有在有父标题时才检查条件
                    for cond in current_defn.get('conditional_styles', []):
                        if cond['if_parent_contains'] in parent_title_text:
                            chosen_style = cond['style']
                            break
                fragments.append((title_text, chosen_style))

                # e. 更新扫描位置
                last_end = match.end()

            # 3. 处理最后一个标题之后剩余的文本
            remaining_content = text[last_end:]
            if remaining_content:
                # 同样，我们假设这部分内容是正文
                fragments.append((remaining_content, default_style_name))
            
            return fragments

        # -------------------------------------------------------------
        #  为了精确处理三级标题，我们需要一个更精细的“先分章节，再深入”的逻辑
        #  上面的纯递归逻辑过于理想化，下面是为您量身定制的、更稳健的实现
        # -------------------------------------------------------------
        
        final_fragments = []
        l1_def = definitions[0]
        l1_pattern = re.compile(l1_def['pattern'])
        
        last_end = 0
        # Pass 1: 扫描一级标题，将文本分为大的章节
        for l1_match in l1_pattern.finditer(text_content):
            # 添加一级标题前的内容
            content_before = text_content[last_end:l1_match.start()]
            if content_before: final_fragments.append((content_before, default_style_name))

            # 添加一级标题
            l1_title_text = l1_match.group(0)
            final_fragments.append((l1_title_text, l1_def['style']))
            
            # 定义当前章节的范围
            chapter_start = l1_match.end()
            next_l1_match = next(l1_pattern.finditer(text_content, pos=chapter_start), None)
            chapter_end = next_l1_match.start() if next_l1_match else len(text_content)
            chapter_content = text_content[chapter_start:chapter_end]
            
            # Pass 2 & 3: 在当前章节内，处理二级和三级标题
            if l1_def.get('sub_definitions'):
                final_fragments.extend(
                    Utils._parse_chapter(chapter_content, l1_def['sub_definitions'], l1_title_text, default_style_name)
                )
            else:
                if chapter_content:
                    final_fragments.append((chapter_content, default_style_name))

            last_end = chapter_end

        # 处理最后一个一级标题之后的内容
        final_content = text_content[last_end:]
        if final_content: final_fragments.append((final_content, default_style_name))

        # (这里的 pprint 仅用于调试，您可以保留或移除)
        print("\n" + "="*20 + " [最终版解析]生成的 fragments 列表 " + "="*20)
        pprint.pprint(final_fragments)
        print("="*65 + "\n")

        return final_fragments
    
    @staticmethod
    def format_string(template: str, data: dict) -> str:
        """
        一个更安全的字符串格式化辅助方法。
        - 自动为良率数据补全百分号 '%'。
        - 如果模板中的占位符在数据字典中不存在，会使用 'N/A' 代替并记录警告，而不是抛出异常。
        """
        # 1. 创建数据的安全副本，以避免修改原始字典
        safe_data = data.copy()

        # 2. 对特定字段进行格式化预处理
        #    只有当值不是 'N/A' 时才添加 '%'
        if 'daily_loss_rate' in safe_data and safe_data.get('daily_loss_rate') != 'N/A':
            safe_data['daily_loss_rate'] = str(safe_data['daily_loss_rate']) + '%'
        if 'monthly_loss_rate' in safe_data and safe_data.get('monthly_loss_rate') != 'N/A':
            safe_data['monthly_loss_rate'] = str(safe_data['monthly_loss_rate']) + '%'

        # 3. 定义一个内部类，用于处理缺失的键
        class SafeDict(dict):
            def __missing__(self, key):
                logging.warning(
                    f"模板中的占位符 '{{{key}}}' 在数据中未找到，将使用 'N/A' 代替。"
                )
                return 'N/A'
        
        # 4. 使用 .format_map 和自定义的SafeDict进行安全的格式化
        return template.format_map(SafeDict(safe_data))

    @staticmethod
    def find_latest_valid_file(directory: str, pattern: str) -> Optional[Path]:
        """
        在一个目录中根据模式查找最新的、有效的文件。
        - 会自动忽略Excel的临时锁定文件 (以 '~$' 开头)。
        """
        # (此方法的代码与我之前提供给你的版本完全一致)
        try:
            dir_path = Path(directory)
            latest_file = None
            latest_mtime = 0

            logging.info(f"  [Utils] 正在目录 '{dir_path.name}' 中搜索 (模式: '{pattern}')...")
            
            for file_path in dir_path.glob(pattern):
                # 核心过滤逻辑：不是锁定文件，并且是一个真实的文件
                if not file_path.name.startswith('~$') and file_path.is_file():
                    current_mtime = file_path.stat().st_mtime
                    if current_mtime > latest_mtime:
                        latest_mtime = current_mtime
                        latest_file = file_path

            if latest_file:
                logging.info(f"  [Utils] 找到最新的有效文件: {latest_file.name}")
            else:
                logging.warning(f"  [Utils] 在目录 '{dir_path.name}' 中未找到匹配的有效文件。")

            return latest_file
        except Exception as e:
            logging.error(f"  [Utils] 查找最新文件时发生错误: {e}", exc_info=True)
            return None

    @staticmethod
    def resolve_lock_file(source_path: Path) -> Path:
        """
        (新增的私有辅助方法) 
        检查给定的路径是否为Excel锁定文件(~$开头)，如果是，则尝试返回真实文件路径。
        """
        if source_path and source_path.name.startswith('~$'):
            # 移除 "~$" 前缀来构建真实文件名
            real_filename = source_path.name[2:]
            real_path = source_path.with_name(real_filename)
            
            # 检查真实文件是否存在
            if real_path.is_file():
                logging.info(f"  [Utils] 检测到Excel锁定文件，自动切换到真实文件: '{real_path.name}'")
                return real_path
            else:
                 logging.warning(f"  [Utils] 检测到锁定文件，但对应的真实文件不存在: '{real_path.name}'")
        
        # 如果不是锁定文件，或者真实文件不存在，则返回原始路径
        return source_path
    
    @staticmethod
    def generate_cell_list(seq_config: dict) -> list[str]:
        """根据序列配置，生成单元格地址列表。"""
        cell_list = []
        start_cell = seq_config.get('start_cell', 'A1')
        step = seq_config.get('step', 1)
        # 如果 end_row 未配置，则默认只处理 start_cell 本身
        row_match = re.search(r"(\d+)", start_cell)
        end_row = seq_config.get('end_row', int(row_match.group(1))) if row_match else 1

        col_match = re.match(r"([A-Z]+)", start_cell)
        col = col_match.group(1) if col_match else 'A'
        
        row_match = re.search(r"(\d+)", start_cell)
        start_row = int(row_match.group(1)) if row_match else 1

        for row in range(start_row, end_row + 1, step):
            cell_list.append(f"{col}{row}")
        
        return cell_list
    
    @staticmethod
    def apply_alignment_styles(output_path: Path, jobs_config: list):
        """
        (新增) 后期处理函数：打开一个已生成的Excel文件，并应用单元格对齐样式。
        """
        if not output_path.exists():
            logging.error(f"[Utils] 样式处理失败：目标文件不存在 '{output_path}'")
            return

        logging.info(f"[Utils] 开始对文件 '{output_path.name}' 进行后期样式处理...")
        try:
            # 1. 加载工作簿
            wb = load_workbook(output_path)
            
            # 2. 定义要应用的样式
            alignment_style = Alignment(wrap_text=True, vertical='top', horizontal='left')

            # 3. 遍历所有任务配置，找出需要应用样式的单元格
            for job_config in jobs_config:
                sheet_name = job_config.get('sheet_name')
                seq_config = job_config.get('cell_sequence')
                
                if not sheet_name or not seq_config:
                    continue

                ws = wb[sheet_name]
                # 4. 调用通用的列表生成函数
                cell_list = Utils.generate_cell_list(seq_config)
                
                logging.info(f"  正在为工作表 '{sheet_name}' 的 {len(cell_list)} 个单元格应用样式...")
                # 5. 循环应用样式
                for cell_address in cell_list:
                    # openpyxl 中，对不存在的单元格赋值会自动创建，所以无需检查
                    ws[cell_address].alignment = alignment_style

            # 6. 保存修改
            wb.save(output_path)
            logging.info(f"[Utils] 后期样式处理完成并已保存。")

        except Exception as e:
            logging.error(f"[Utils] 应用样式时发生错误: {e}", exc_info=True)

    