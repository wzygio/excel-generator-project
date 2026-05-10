# src/processors/excel_processor.py
import re, os
import json
import logging
from pathlib import Path
import datetime
from datetime import datetime as dt

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from excel_generator_project.config import PROJECT_ROOT, RESOURCES_DIR, TEMP_DIR
from excel_generator_project.utils.utils import Utils

class DataProcessor:
    """
    配置驱动的Excel报告处理器。
    """
    def __init__(self, config: dict, external_data: dict):
        """
        (已更新) 初始化处理器。
        'output_path' 是指本次处理流程中使用的临时文件路径。
        """
        self.config = config
        self.excel_handler = None # 统一命名
        self.calculated_data = {}
        self.external_data = external_data or {}
        self.cache_dir = TEMP_DIR
        self.product_models = Utils.extract_product_models(config.get("model_definitions", "")) # 提取日报中产品型号
        
    def run(self):
        """
        主调度方法。
        它会先复制模板到输出路径，然后根据配置中的任务列表，依次执行所有任务。
        """
        logging.info(f"--- DataProcessor 开始运行 ---")

        # 调度并执行所有任务
        jobs = self.config.get('data_processing_jobs', []) # 注意，这里需要匹配您YAML中的键名
        logging.info(f"发现 {len(jobs)} 个待处理任务...")

        for i, job_config in enumerate(jobs, 1):
            job_type = job_config.get("job_type")
            logging.info(f"--> 开始执行任务 {i}/{len(jobs)}: 类型 = {job_type}")
    
            if job_type == "parse_text_from_cells": # 统一调用新方法
                self._execute_parse_text_from_cells_job(job_config)
            elif job_type == "extract_tila_target": 
                # 提拉良率目标：self.calculated_data['tila_target']
                self._execute_extract_tila_target_job(job_config) 
            elif job_type == "extract_monthly_yield_estimate": 
                # 月度良率预估：self.calculated_data['prod_target'] 和 self.calculated_data['estimated_yield']
                self._execute_extract_monthly_yield_estimate_job(job_config) 
            elif job_type == "extract_risk_items": 
                # 风险品：self.calculated_data['risk_items']
                self._execute_extract_risk_items_job(job_config) 
            elif job_type == "prepare_summary_data":
                self._execute_prepare_summary_data(job_config)
            else:
                logging.warning(f"未知的任务类型 '{job_type}'，已跳过。")

        # 输出最新任务的结果
        logging.info(f"{json.dumps(self.calculated_data, ensure_ascii=False, indent=2)}")

        logging.info(f"--- DataProcessor 运行结束 ---")
        return self.calculated_data['summary_report_data']
    
    def _execute_prepare_summary_data(self, job_config: dict):
        """
        仅负责准备用于填充模板的数据，并将其存入 self.calculated_data。
        """
        logging.info("    开始执行“准备摘要数据”任务...")
        
        # 假设 __prepare_data_for_templating 是一个已存在的私有方法
        all_reports_data = self.__prepare_data_for_templating()
        
        if not all_reports_data:
            logging.warning("    没有可用于准备摘要的数据。")
            return
        
        # 将准备好的数据列表存入 self.calculated_data，供后续 Generator 使用
        self.calculated_data['summary_report_data'] = all_reports_data
        logging.info(f"    已成功为 {len(all_reports_data)} 个产品型号准备好摘要数据。")

    # --- 核心修改 2：使用以下代码替换现有的 __prepare_data_for_templating 方法 ---
    def __prepare_data_for_templating(self) -> list[dict]:
        """
        (已更新) 增加了"与提拉目标Gap"的衍生计算逻辑。
        """
        logging.info("开始准备模板数据...")
        all_reports_data = []
        num_models = len(self.product_models)
        logging.info(f"需要处理 {num_models} 个产品型号")

        # (从 self.calculated_data 中获取数据的部分保持不变)
        dc = self.calculated_data
        daily_yield_list = dc.get('daily_yield_change', [])
        bp_target_list = dc.get('bp_target', [])
        tila_target_list = dc.get('tila_target', [])
        prod_yield_list = dc.get('production_yield_values', [])
        month_yield_list = dc.get('monthly_yield_values', []) # <--- estimated_yield 在这里
        best_batch_list = dc.get('best_batch', [])
        latest_batch_list = dc.get('latest_batch', [])
        risk_items_map = dc.get('risk_items', "/") # <- 'risk_items' 现在是 {model: "完整字符串"}

        for i, model in enumerate(self.product_models):
            logging.info(f"正在处理产品型号 {i+1}/{num_models}: {model}")
            # --- 新增的Gap计算逻辑 ---
            # 1. 安全地获取原始数据
            estimated_yield_str = month_yield_list[i] if i < len(month_yield_list) else "／"
            tila_target_str = tila_target_list[i] if i < len(tila_target_list) else "／"
            # 获取ExceptionProcessor处理后的formatted_titles
            formatted_titles = self.external_data.get(model, {}).get('formatted_titles', [])
            # 将titles用换行符连接
            new_exceptions = '、'.join(formatted_titles) if formatted_titles else "/"

            # 2. 调用辅助函数进行安全转换
            float_yield = Utils.safe_convert_percent_to_float(estimated_yield_str)
            float_tila = Utils.safe_convert_percent_to_float(tila_target_str)

            # 3. 执行计算
            estimated_tila_gap_str = "／" # 默认值
            if float_yield is not None and float_tila is not None:
                gap_value = float_yield - float_tila
                # 将结果格式化回带一位小数的百分比字符串
                estimated_tila_gap_str = f"{gap_value:.1%}"
                logging.info(f"型号 {model} 的Gap计算结果: {estimated_yield_str} - {tila_target_str} = {estimated_tila_gap_str}")

            # 构建数据字典
            report_data = {
                'model_name': model,
                'daily_yield_change': daily_yield_list[i] if i < len(daily_yield_list) else "／",
                'new_exceptions': new_exceptions,
                'known_exceptions': "/",
                'bp_target': bp_target_list[i] if i < len(bp_target_list) else "／",
                'tila_target': tila_target_str, # 使用我们获取的原始字符串
                'prod_target': prod_yield_list[i] if i < len(prod_yield_list) else "／",
                'estimated_yield': estimated_yield_str, # 使用我们获取的原始字符串
                'estimated_tila_gap': estimated_tila_gap_str, # <-- 4. 插入新计算出的Gap值
                'best_batch': best_batch_list[i] if i < len(bp_target_list) else "／",
                'latest_batch': latest_batch_list[i] if i < len(bp_target_list) else "／",
                'risk_items': risk_items_map.get(model, "/"),
                'unique_exceptions': "/",
                'array_opportunities': "/",
                'oled_opportunities': "/",
                'tp_opportunities': "/"
            }
            all_reports_data.append(report_data)

        logging.info(f"模板数据准备完成，共处理 {len(all_reports_data)} 个产品型号的数据")
        return all_reports_data

    def _execute_parse_text_from_cells_job(self, job_config: dict):
        """
        (新增的通用方法)
        按步长遍历一系列单元格，读取其文本内容，应用正则表达式提取数据，
        并将结果存入由 'output_key' 指定的 self.calculated_data 键中。
        """
        # --- 1. 解析并验证配置 ---
        output_key = job_config.get("output_key") # <-- 关键：获取输出键名
        if not output_key:
            logging.error(f"任务 '{job_config.get('description')}' 配置错误：缺少 'output_key'。")
            return

        source_path = Utils.get_safe_source_path(job_config)
        if not source_path or not source_path.is_file():
            logging.error(f"  未能获取有效的源文件路径，任务 '{output_key}' 中止。")
            return

        sheet_name = job_config.get("sheet_name")
        pattern = job_config.get("pattern")
        sequence = job_config.get("cell_sequence", {})
        start_cell = sequence.get("start_cell")
        step = sequence.get("step")
        end_row = sequence.get("end_row")

        if not all([source_path.is_file(), sheet_name, pattern, start_cell, isinstance(step, int), isinstance(end_row, int)]):
            logging.error(f"任务 '{output_key}' 配置不完整、类型错误或源文件不存在。")
            return

        try:
            # --- 2. 加载源数据文件 ---
            source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            source_ws = source_wb[str(sheet_name)]  # 强制转换为字符串

            # --- 3. 步长迭代逻辑 ---
            col_letter = ''.join(filter(str.isalpha, start_cell))
            current_row = int(''.join(filter(str.isdigit, start_cell)))

            extracted_list = []
            consecutive_none_count = 0
            logging.info(f"  任务 '{output_key}': 开始在 {source_path.name}[{sheet_name}] 中搜索...")
            
            while current_row <= end_row:
                cell_address = f"{col_letter}{current_row}"
                cell_value = source_ws[cell_address].value
                
                if cell_value is None:
                    consecutive_none_count += 1
                else:
                    consecutive_none_count = 0
                
                if consecutive_none_count >= 2:
                    logging.info(f"  检测到连续两次空值，提取在第 {current_row} 行提前终止。")
                    break

                if isinstance(cell_value, str):
                    match = re.search(str(pattern), cell_value)
                    if match:
                        extracted_value = match.group(1).strip()
                        logging.info(f"  在单元格 {cell_address} 找到匹配项: {extracted_value}")
                        extracted_list.append(extracted_value)

                current_row += step

            # --- 4. 存储提取结果 ---
            if extracted_list:
                # <-- 关键：使用动态键名存入结果 ---
                self.calculated_data[output_key] = extracted_list
                logging.info(f"  数据提取成功: '{output_key}' 已存为包含 {len(extracted_list)} 个值的列表。")
            else:
                logging.warning(f"  未能在指定的单元格序列中找到与模式 '{pattern}' 匹配的数据。")

        except Exception as e:
            logging.error(f"  处理任务 '{output_key}' 时发生意外错误: {e}", exc_info=True)
        finally:
            if 'source_wb' in locals(): # 确保工作簿被关闭
                source_wb.close()
    
    def _execute_extract_tila_target_job(self, job_config: dict):
        """
        (新方法) 作为“提取提拉良率目标”这个多步骤任务的总调度方法。
        """
        logging.info(" 开始执行“提取提拉良率目标”多步骤任务...")

        # 从总任务配置中获取第一步的专属配置
        
        try:
            # 调用私有辅助方法来执行第一步的具体逻辑
            product_models = self.product_models
            logging.info(f" 第一步完成：提取到 {len(product_models)} 个产品型号。")
            
             # --- 准备第二、三步所需的工作表 ---
            # 根据 step2 的配置加载工作簿和工作表
            full_path = Utils.get_safe_source_path_year(job_config)
            if not full_path or not full_path.is_file():
                logging.error(f"    未能获取有效的源文件路径，任务中止。")
                return

            sheet_name = job_config.get("sheet_name", "Sheet1")
            source_wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            source_ws = source_wb[sheet_name]

             # --- 第二步: 为每个产品型号查找其所在的行 ---
            model_to_row_map = self.__tila_step2_find_model_rows(job_config.get("row_locator", {}), product_models, source_ws)
            logging.info(f" 第二步完成：已为 {len(model_to_row_map)} 个型号定位到行。")

            # --- 第三步: 查找当前月份所在的列 (只执行一次) ---
            target_column = self.__tila_step3_find_month_column(job_config.get("column_locator", {}), source_ws)
            if not target_column:
                logging.error(" 第三步失败：未能在指定行找到当前月份列，任务中止。")
                return
            logging.info(f" 第三步完成：找到当前月份所在列为 '{target_column}'。")

            # --- 第四步: 提取每个产品型号的提拉良率目标 ---
            final_values = []
            for model in product_models:
                if model in model_to_row_map:
                    target_row = model_to_row_map[model]
                    cell_address = f"{target_column}{target_row}"
                    value = source_ws[cell_address].value
                    if isinstance(value, (int, float)):
                        final_values.append(f"{value:.1%}")
                    logging.info(f" -> 型号 '{model}' (行 {target_row}) 的目标值为: {value}")
                else:
                    final_values.append("/") # 如果某个型号没找到对应的行，则添加None占位
                    logging.warning(f" -> 型号 '{model}' 未在指定列中找到，结果记为 None。")
            
            # 将最终提取到的数值列表存入 calculated_data
            self.calculated_data['tila_target'] = final_values
            logging.info(f" 任务完成: 'tila_target' 已存为包含 {len(final_values)} 个值的列表。")
        except Exception as e:
            logging.error(f" 处理任务 'extract_tila_target' 时发生意外错误: {e}", exc_info=True)
    

    
    def __tila_step2_find_model_rows(self, step2_config: dict, product_models: list, worksheet) -> dict:
        """
        (新方法) 执行提取tila_target的第二步：为每个产品型号查找其所在的行号。
        """
        search_col = step2_config.get("search_column")
        model_to_row_map = {} # 创建一个字典来存储 {型号: 行号} 的映射
        
        # 为了效率，我们只遍历一次工作表，而不是为每个型号都遍历一遍
        # 假设产品型号不会超过1000行
        for row in range(1, 1001):
            cell_address = f"{search_col}{row}"
            cell_value = worksheet[cell_address].value
            
            if isinstance(cell_value, str):
                # 检查单元格内容是否与我们的任何一个产品型号匹配
                for model in product_models:
                    if model in cell_value and model not in model_to_row_map:
                        model_to_row_map[model] = row
                        # 如果所有型号都找到了，就提前退出循环
                        if len(model_to_row_map) == len(product_models):
                            return model_to_row_map
        return model_to_row_map

    def __tila_step3_find_month_column(self, step3_config: dict, worksheet) -> str | None:
        """
        (新方法) 执行提取tila_target的第三步：在指定行中查找当前月份所在的列。
        """
        search_row = step3_config.get("search_row")
        month_format = step3_config.get("month_format")
        
        # 1. 自动计算当前月份
        current_month = dt.now().month
        search_text = (month_format or "{month}月").format(month=current_month)
        logging.info(f" 第三步：正在第 {search_row} 行查找内容包含 '{search_text}' 的单元格...")

        # 2. 遍历指定行来查找月份
        # 假设月份信息不会超过26列（Z列）
        for col in range(1, 50):
            cell = worksheet.cell(row=search_row, column=col)
            if cell.value and isinstance(cell.value, str) and search_text in cell.value:
                return cell.column_letter # 返回找到的列字母，例如 'F'
        
        return None # 如果没有找到，返回None
    
    def _execute_extract_monthly_yield_estimate_job(self, job_config: dict):
        """
        (新方法) 执行“提取月度良率预估”这个复杂任务的总调度方法。
        """
        logging.info("    开始执行“提取月度良率预估”任务...")
        
        try:
            # --- 步骤1: 动态定位源文件 ---
            file_locator_config = job_config.get("file_locator", {})
            initial_target_path = self.__find_latest_file(file_locator_config)
            if not initial_target_path: 
                logging.error("    未能找到源文件，任务中止。")
                return
            
            source_path = Utils.resolve_lock_file(initial_target_path)
            local_cache_dir = PROJECT_ROOT /  RESOURCES_DIR
            source_file_path = Utils.get_local_copy(source_path, local_cache_dir)
            if not source_file_path:
                logging.error("    未能根据 file_locator 配置找到源文件，任务中止。")
                return

            # --- 步骤2: 加载工作簿和工作表 ---
            locator_config = job_config.get("cell_locator", {})
            sheet_name = locator_config.get("sheet_name")
            source_wb = openpyxl.load_workbook(source_file_path, read_only=True, data_only=True)
            source_ws = source_wb[sheet_name]
            header_row = locator_config.get("header_row", 1)

            # --- 步骤3: 定位所有需要的列 ---
            # 3.1 定位“产品型号”所在的列
            model_col_config = locator_config.get("row_locator", {})
            model_col = self.__find_column_by_header(source_ws, model_col_config, header_row)
            
            # 3.2 定位数据列
            col_locator_config = locator_config.get("column_locator", {})
            prod_yield_col_config = col_locator_config.get("production_yield_col", {})
            month_yield_col_config = col_locator_config.get("monthly_yield_col", {})
            
            prod_yield_col = self.__find_column_by_header(source_ws, prod_yield_col_config, header_row)
            month_yield_col = self.__find_column_by_format(source_ws, month_yield_col_config, header_row)

            if not all([model_col, prod_yield_col, month_yield_col]):
                logging.error("    未能定位到所有必需的列（型号、排产、当月），任务中止。")
                return
            logging.info(f"    所有关键列定位成功: 型号列={model_col}, 排产良率列={prod_yield_col}, 当月预估列={month_yield_col}")

            # --- 步骤4: 遍历产品型号，查找行并提取数据 ---
            product_models = self.product_models
            if not product_models:
                logging.warning("    产品型号列表为空，无法进行数据提取。")
                return

            # --- 核心修改：使用优化后的逻辑构建映射 ---
            if model_col is not None:
                model_to_row_map = self.__map_models_to_rows(source_ws, model_col, header_row, product_models)
            
            # 打印调试信息，确认 C516 是否被正确找到
            if 'C516' not in model_to_row_map:
                logging.warning("    [调试] 依然未找到 C516 的行映射，请检查Excel中该单元格的内容是否标准。")
            elif 'C51' in model_to_row_map and model_to_row_map['C51'] == model_to_row_map['C516']:
                logging.error("    [调试] 严重错误：C51 和 C516 映射到了同一行！")

            # 根据产品型号列表的顺序，整理并提取最终数据
            production_yield_list = []
            monthly_yield_list = []
            
            for model in product_models:
                if model in model_to_row_map:
                    target_row = model_to_row_map[model]
                    prod_value = source_ws[f"{prod_yield_col}{target_row}"].value
                    month_value = source_ws[f"{month_yield_col}{target_row}"].value
                    
                    prod_str = f"{prod_value:.1%}" if isinstance(prod_value, (int, float)) else "/"
                    month_str = f"{month_value:.1%}" if isinstance(month_value, (int, float)) else "/"
                    
                    production_yield_list.append(prod_str)
                    monthly_yield_list.append(month_str)
                else:
                    production_yield_list.append("/")
                    monthly_yield_list.append("/")
                    logging.warning(f"      在 {model_col} 列中未找到型号 '{model}' 对应的行。")
            
            # --- 步骤5: 存储最终结果 ---
            self.calculated_data['production_yield_values'] = production_yield_list
            self.calculated_data['monthly_yield_values'] = monthly_yield_list
            logging.info("    任务完成: 已成功提取'排产良率'和'当月预估'的数据列表。")

        except Exception as e:
            logging.error(f"    处理任务 'extract_monthly_yield_estimate' 时发生意外错误: {e}", exc_info=True)



        return model_to_row_map
    def __find_latest_file(self, locator_config: dict) -> Path | None:
        """
        (修复Bug版) 根据配置在目录中查找最新的匹配文件。
        自动识别当前年份，优先查找当前年份，找不到则查找上一年份。
        不再依赖当前时间计算周数，而是直接扫描目录下存在的最大周数文件夹，避免跨年ISO周问题。
        """
        # 自动获取当前年份
        current_year = datetime.date.today().year
        
        # 处理base_path中的年份占位符
        base_path_template = locator_config.get("base_path", "")
        name_contains = locator_config.get("name_contains", "")
        dir_rule = locator_config.get("directory_rule", {}) 

        # 尝试查找文件，支持年份回退
        try:
            for year_offset in [0, -1]:  # 优先当前年份，然后上一年份
            
                # 格式化路径中的年份
                base_path_str = base_path_template.format(year=current_year + year_offset)
                base_path = Path(base_path_str)
                
                if not base_path.is_dir():
                    logging.warning(f"    路径不存在或不是一个目录: {base_path}")
                    continue
                
                logging.info(f"    正在尝试查找年份 {current_year + year_offset} 的数据...")
                
                # --- 第一阶段：查找最新子目录 (扫描并排序模式) ---
                if dir_rule:
                    prefix = dir_rule.get("prefix")
                    if prefix:
                        # 1. 获取所有子目录
                        all_subdirs = [p for p in base_path.iterdir() if p.is_dir()]
                        
                        # 2. 筛选以 prefix (例如'W') 开头的目录，并尝试解析后面的数字
                        valid_week_dirs = []
                        for p in all_subdirs:
                            if p.name.startswith(prefix):
                                try:
                                    # 提取 W 后面的数字 (例如 W51 -> 51)
                                    week_num = int(p.name[len(prefix):])
                                    valid_week_dirs.append((week_num, p))
                                except ValueError:
                                    # 忽略那些符合前缀但后面不是纯数字的文件夹 (如 W_Backup)
                                    continue
                        
                        # 3. 按周数倒序排列 (从大到小: 51, 50, 49...)
                        valid_week_dirs.sort(key=lambda x: x[0], reverse=True)

                        if not valid_week_dirs:
                            logging.warning(f"    在 '{base_path.name}' 下未找到任何格式为 '{prefix}+数字' 的子目录。")
                            continue

                        # 4. 遍历最近的几个文件夹 (例如最近5个)，尝试查找文件
                        for week_num, dir_path in valid_week_dirs[:5]:
                            logging.info(f"    正在检查目录: {dir_path.name}")
                            found_file = self.__search_file_in_directory(dir_path, name_contains)
                            if found_file:
                                logging.info(f"    --> 锁定目标文件: {found_file.name}")
                                return found_file
                                
                        logging.warning(f"    在最近的 {len(valid_week_dirs[:5])} 个周文件夹中均未找到包含 '{name_contains}' 的文件。")
                        continue

                # --- 第二阶段：在基础路径直接查找（如果没有二级目录规则） ---
                found_file = self.__search_file_in_directory(base_path, name_contains)
                if found_file:
                    return found_file
                    
        except Exception as e:
            logging.error(f"    在当前年份 {current_year} 和上一年份 {current_year - 1} 中均未找到目标文件。")
            return None


    def __search_file_in_directory(self, directory: Path, name_contains: str) -> Path | None:
        """在指定目录中查找最新的匹配文件"""
        if not name_contains:
            return None
        
        current_month = datetime.date.today().month
        month_string_to_find = f"{current_month}月"

        latest_file = None
        latest_time = 0

        logging.info(f"    正在目录 '{directory.name}' 中查找包含 '{name_contains}' 和 '{month_string_to_find}' 的最新文件...")
        for filename in os.listdir(directory):
            if (name_contains in filename and 
                not filename.startswith('~$') and 
                month_string_to_find in filename):
                file_path = directory / filename
                if file_path.is_file():  # 确保是文件
                    mod_time = file_path.stat().st_mtime
                    if mod_time > latest_time:
                        latest_time = mod_time
                        latest_file = file_path
        
        if latest_file:
            logging.info(f"      -> 找到最新文件: {latest_file.name}")
        else:
            logging.warning(f"    未能找到包含 '{name_contains}' 的文件。")
        
        return latest_file

    
    def __find_column_by_header(self, worksheet, header_config: dict, header_row: int) -> str | None:
        """(新辅助方法) 根据表头名称和出现次数查找列。"""
        header_name = header_config.get("header_name")
        occurrence = header_config.get("occurrence", 1)
        
        count = 0
        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str) and (header_name or "") in cell.value:
                count += 1
                if count == occurrence:
                    return cell.column_letter
        return None

    def __find_column_by_format(self, worksheet, format_config: dict, header_row: int) -> str | None:
        """(新辅助方法) 根据动态格式化的表头名称查找列。"""
        header_format = format_config.get("header_format")
        current_month = dt.now().month
        search_text = (header_format or "{month}月").format(month=current_month)

        for cell in worksheet[header_row]:
            if cell.value and isinstance(cell.value, str) and search_text in cell.value:
                return cell.column_letter
        return None
    
    def __map_models_to_rows(self, worksheet, model_col: str, header_row: int, product_models: list) -> dict:
        """
        (修复版) 建立 {产品型号: 行号} 的映射。
        策略：
        1. 优先使用【严格精确匹配】(cell == model)，彻底解决 C51 匹配到 C451 的问题。
        2. 仅在确实需要时，使用【单词边界正则】作为兜底（确保匹配到的 C51 左右没有数字或字母）。
        """
        model_to_row_map = {}
        
        # 1. 准备数据结构
        # 使用 Set 进行 O(1) 的精确查找，同时去除配置中的首尾空格
        models_set = set(str(m).strip() for m in product_models)
        
        logging.info(f"    正在构建型号-行号映射 (策略: 严格精确匹配)...")

        # 2. 遍历 Excel 行
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            cell_value = worksheet[f"{model_col}{row_idx}"].value
            
            if isinstance(cell_value, str):
                cell_text = cell_value.strip() # 去除单元格内容的首尾空格
                if not cell_text:
                    continue

                # --- 策略 A: 严格精确匹配 (解决 C51 vs C451 问题的关键) ---
                # 只有当单元格内容完全等于 "C51" 时才算匹配
                if cell_text in models_set:
                    model_to_row_map[cell_text] = row_idx
                    continue 

                # --- 策略 B: 智能边界匹配 (可选兜底) ---
                # 如果您的表格 N 列非常干净（如图所示），上面的策略 A 已经足够且最安全。
                # 但为了防止像 "Product: C51" 这样的情况漏掉，我们可以加一个带边界检查的逻辑。
                # 它可以防止 "C451" (C左边有数字) 或 "C516" (1右边有数字) 被匹配。
                
                # 注意：如果您的N列确实只有型号，建议注释掉下面这段 策略B 的代码，只保留 策略A
                """
                for model in product_models:
                    model_str = str(model).strip()
                    # 使用正则检查单词边界：确保找到的 model 前后不是字母或数字
                    # \b 在处理中文时可能不准，所以用 lookbehind/lookahead 模拟
                    # 意思：前面不能是字母数字，后面也不能是字母数字
                    pattern = r'(?<![a-zA-Z0-9])' + re.escape(model_str) + r'(?![a-zA-Z0-9])'
                    if re.search(pattern, cell_text):
                        model_to_row_map[model_str] = row_idx
                        break
                """

        return model_to_row_map
    
    def _execute_extract_risk_items_job(self, job_config: dict):
        """
        (已重构 - 暂时移除基线逻辑) 
        1.提取模板中的风险品 2.获取更新释放计划 3.合并并生成最终报告字符串。
        """
        logging.info("    开始执行“提取风险品信息”任务...")
        
        # --- 1. 解析配置 ---
        reader_cfg = job_config.get("template_reader_config", {})
        parser_cfg = job_config.get("parser_config", {})
        plan_finder_cfg = job_config.get("plan_finder_config", {})
        
        model_text_map = {} # {model: "OLED: ... TP: ..."}
        model_data_map = {} # {model: [('Item1', 'Yield1'), ('Item2', 'Yield2')]}
        
        try:
            # --- 2. 步骤1: 从(新)日报模板中提取“风险品”文本块和(名称, 良率)元组 ---
            source_path = Utils.get_safe_source_path(reader_cfg)
            wb = openpyxl.load_workbook(str(source_path), read_only=True, data_only=True)
            ws = wb[reader_cfg.get("sheet_name")]
            cell_sequence = reader_cfg.get("cell_sequence", {})
            block_pattern = reader_cfg.get("block_pattern")
            
            col_letter = ''.join(filter(str.isalpha, cell_sequence.get('start_cell')))
            start_row = int(''.join(filter(str.isdigit, cell_sequence.get('start_cell'))))
            step = cell_sequence.get('step')

            patterns = [parser_cfg.get(k) for k in parser_cfg if k.startswith('item_pattern')]
            
            for i, model in enumerate(self.product_models):
                current_row = start_row + i * step
                cell_value = ws[f"{col_letter}{current_row}"].value
                
                if isinstance(cell_value, str):
                    block_match = re.search(block_pattern, cell_value, re.DOTALL)
                    if block_match:
                        text_block = block_match.group(1).strip()
                        model_text_map[model] = text_block # 存储原始文本块
                        
                        # 立即解析出 (名称, 良率) 元组
                        model_data_map[model] = []
                        for pat in patterns:
                            if pat:
                                matches = re.findall(pat, text_block)
                                model_data_map[model].extend(matches)
            wb.close()
            logging.info("      -> 步骤1完成：已从新模板中解析出风险品名称和良率。")

            # --- 3. 步骤2: (暂时去掉) 从“旧日报”中获取“基线”释放计划 ---
            # baseline_cfg = plan_finder_cfg.get("old_release_plan_finder", {})
            # baseline_plan_map = self.__risk_items_build_baseline_plan_lookup(baseline_cfg)
            baseline_plan_map = {} # 暂时置为空字典
            logging.info("      -> 步骤2跳过：当前已禁用从旧日报提取基线数据。")

            # --- 4. 步骤3: 从“风险品汇总表”中获取“更新”释放计划 ---
            update_cfg = plan_finder_cfg.get("new_release_plan_finder", {})
            update_plan_map = self.__risk_items_build_update_plan_lookup(update_cfg)
            logging.info(f"      -> 步骤3完成：已从汇总表获取 {len(update_plan_map)} 条“更新”释放计划。")

            # --- 5. 步骤4: 合并释放计划 (当前仅包含汇总表内容) ---
            final_plan_map = baseline_plan_map.copy()
            # 只有当汇总表里的计划不为空时，才执行覆盖
            for key, new_plan in update_plan_map.items():
                if new_plan and str(new_plan).strip():
                    final_plan_map[key] = new_plan
            logging.info(f"      -> 步骤4完成：已合并释放计划，共 {len(final_plan_map)} 条有效计划。")

            # --- 6. 步骤5: 组装最终的报告字符串 ---
            final_report_strings = {}
            for model, text_block in model_text_map.items():
                modified_text_block = text_block
                
                # 遍历此型号在模板中找到的 (名称, 良率) 元组
                for risk_name_raw, yield_str_raw in model_data_map.get(model, []):
                    risk_name = risk_name_raw.strip()
                    yield_str = yield_str_raw.strip()
                    
                    # 从合并后的最终计划中查找
                    plan_str = final_plan_map.get((model, risk_name)) # 默认为空
                    
                    if plan_str: # 只有当计划非空时才进行拼接
                        original_line_pattern = re.compile(re.escape(risk_name) + r'\s*' + re.escape(yield_str))
                        replacement_line = f"{risk_name} {yield_str}：{plan_str}"
                        modified_text_block = original_line_pattern.sub(replacement_line, modified_text_block)
                
                final_report_strings[model] = modified_text_block
            
            if final_report_strings:
                self.calculated_data['risk_items'] = final_report_strings
                logging.info("      -> 步骤5完成：已生成最终的风险品信息字符串。")

        except Exception as e:
            logging.error(f"    处理任务 'extract_risk_items' 时发生意外错误: {e}", exc_info=True)
        finally:
            if 'wb' in locals() and wb:
                wb.close()

    
    
    def __risk_items_build_baseline_plan_lookup(self, config: dict) -> dict:
        """
        (新增) 步骤1：从“最近的旧日报”中，按产品型号匹配，提取“基线”释放计划。
        """
        baseline_plan_map = {}
        if not config:
            logging.warning("    未配置 'old_release_plan_finder'，跳过基线计划提取。")
            return baseline_plan_map

        # 1. 查找旧日报文件
        dynamic_path_cfg = config.get("dynamic_path_config", {})
        source_path = Utils.find_previous_report_file(dynamic_path_cfg) # 复用 ExceptionProcessor 的查找方法
        if not source_path:
            logging.warning("    未能找到旧日报文件，无法提取基线计划。")
            return baseline_plan_map

        source_file_path = Utils.get_local_copy(source_path, self.cache_dir)
        if not source_file_path:
            logging.warning("    复制旧日报文件失败，无法提取基线计划。")
            return {}

        # 2. 解析配置
        sheet_name = config.get("sheet_name")
        model_col = config.get("product_model_column")
        model_start_row = config.get("model_start_row")
        model_step = config.get("model_step")
        risk_col = config.get("risk_content_column")
        risk_offset = config.get("risk_content_offset")
        block_pattern = config.get("block_pattern")
        parser_pattern = config.get("parser_pattern")

        if not all([sheet_name, model_col, model_start_row, model_step, risk_col, risk_offset is not None, block_pattern, parser_pattern]):
            logging.error("    'old_release_plan_finder' 配置不完整，任务中止。")
            return {}

        try:
            # 3. 加载旧日报
            wb = openpyxl.load_workbook(source_file_path, read_only=True, data_only=True)
            ws = wb[str(sheet_name)]

            # 4. 按产品型号精确匹配
            for model in self.product_models:
                found = False

                for row_idx in range(model_start_row, ws.max_row, model_step): # type: ignore

                    cell_value = ws[f"{model_col}{row_idx}"].value
                    if isinstance(cell_value, str) and model in cell_value:
                        # 找到了产品型号，现在去D列+偏移量处查找
                        risk_row = row_idx + risk_offset # type: ignore
                        risk_cell_value = ws[f"{risk_col}{risk_row}"].value
                        
                        if isinstance(risk_cell_value, str):
                            # 提取风险品文本块
                            block_match = re.search(block_pattern, risk_cell_value, re.DOTALL) # type: ignore
                            if block_match:
                                text_block = block_match.group(1).strip()
                                # 解析 (名称, 计划) 元组
                                matches = re.findall(parser_pattern, text_block) # type: ignore
                                logging.info(f"        [解析条目数量] {len(matches)} 个")
                                for risk_name_raw, plan_str_raw in matches:
                                    risk_name = risk_name_raw.strip()
                                    plan_str = plan_str_raw.strip()
                                    if risk_name and plan_str:
                                        baseline_plan_map[(model, risk_name)] = plan_str
                        found = True
                        logging.info(f" [基线匹配成功] {model} - {risk_name}: {plan_str}")
                        break # 找到了这个model，继续找下一个
                
                if not found:
                    logging.warning(f"    在旧日报的 {model_col} 列未找到型号 '{model}'。")
            
            wb.close()
            return baseline_plan_map
            
        except Exception as e:
            logging.error(f"    提取“基线”释放计划时发生错误: {e}", exc_info=True)
            if 'wb' in locals() and wb: wb.close()
            return {}
        
    def __risk_items_build_update_plan_lookup(self, config: dict) -> dict:
        """
        (新增) 步骤2：从“风险品汇总表”中，提取“更新”释放计划。
        (此方法逻辑基于旧的 __build_release_plan_lookup)
        """
        update_plan_map = {}
        if not config:
            logging.warning("    未配置 'new_release_plan_finder'，跳过更新计划提取。")
            return update_plan_map
            
        source_path = Utils.get_safe_source_path(config)
        sheet_name = config.get("sheet_name")
        if not source_path or not sheet_name:
            logging.error(f"    未能获取有效的“风险品汇总表”路径，任务中止。")
            return {}

        try:
            worksheet = openpyxl.load_workbook(source_path, data_only=True)[sheet_name]
            
            header_row = config.get("header_row", 1)
            model_col_header = config.get("product_model_column")
            risk_col_header = config.get("risk_item_column")
            plan_col_header = config.get("release_plan_column")

            model_col = self.__find_column_by_header(worksheet, {'header_name': model_col_header}, header_row)
            risk_col = self.__find_column_by_header(worksheet, {'header_name': risk_col_header}, header_row)
            plan_col = self.__find_column_by_header(worksheet, {'header_name': plan_col_header}, header_row)
            if not all([model_col, risk_col, plan_col]):
                logging.error("    未能在“风险品汇总表”中定位到所有必需的列。")
                return {}

            # 构建查询字典
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                model = worksheet[f"{model_col}{row_idx}"].value
                risk = worksheet[f"{risk_col}{row_idx}"].value
                plan = worksheet[f"{plan_col}{row_idx}"].value
                if model and risk:
                    update_plan_map[(str(model).strip(), str(risk).strip())] = str(plan).strip() if plan else ""
            
            return update_plan_map
        except Exception as e:
            logging.error(f"    加载“风险品汇总表”时出错: {e}", exc_info=True)
            return {}