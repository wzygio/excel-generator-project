import logging
import xlsxwriter
import zipfile
import shutil
import os, re

from openpyxl import load_workbook
from xml.etree import ElementTree as ET

from excel_generator_project.utils.utils import Utils # 假设您的文本处理函数在这里

# 预定义XML命名空间
NAMESPACES = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
}

class FontProcessor:
    """
    (已重构) 封装了Excel富文本处理的高层业务流程。
    """
    def __init__(self, template_path, output_path, config: dict):
        logging.info("FontProcessor 实例已创建，正在初始化...")
        self.template_path = template_path
        self.output_path = output_path
        self.style_configs = config.get('style_definitions', {})
        self.batch_jobs = config.get('font_processing_jobs', [])
        
        try:
            shutil.copy(self.template_path, self.output_path)
            with zipfile.ZipFile(self.output_path, 'r') as target_zip:
                zip_contents = {name: target_zip.read(name) for name in target_zip.namelist()}
            
            # --- 核心改动：创建并持有一个XML辅助类的实例 ---
            self.xml_helper = _RichTextXMLHelper(zip_contents, self.style_configs)
            
            logging.info(f"模板文件 '{template_path}' 已成功加载到内存。")
        except Exception as e:
            logging.error(f"初始化失败：无法加载模板文件。错误: {e}", exc_info=True)
            raise

    def run(self):
        """
        执行配置文件中定义的所有批量处理任务。
        """
        try:
            logging.info(f"准备执行 {len(self.batch_jobs)} 个批量处理任务。")
            for job_config in self.batch_jobs:
                self._batch_process_single_job(job_config)

            # --- 核心改动：通过辅助类来执行回写 ---
            self.xml_helper.repack_excel_file(self.output_path)
            
            logging.info(" 所有批量处理任务均已成功完成！")
            return True
        except Exception as e:
            logging.error(f"批量处理主流程发生严重错误: {e}", exc_info=True)
            return False

    def _batch_process_single_job(self, job_config: dict):
        """(已重构) 作为高层调度器，协调各个辅助方法完成任务。"""
        job_name = job_config.get('job_name', '未命名任务')
        sheet_name = job_config.get('sheet_name')
        default_style = 'part1_content_style' # 默认内容样式
        logging.info(f"--- 开始执行批量任务: '{job_name}' ---")

        # 步骤 1: 调用辅助方法，生成需要处理的单元格列表
        cell_list = Utils.generate_cell_list(job_config.get('cell_sequence', {}))
        if not cell_list:
            logging.warning(f"任务 '{job_name}': 未能生成任何单元格，任务中止。")
            return

        # 步骤 3: 准备只读的工作簿，用于安全地读取原始数据
        try:
            if not sheet_name:
                logging.error(f"任务 '{job_name}': 未提供工作表名")
                return
            
            wb_template_readonly = load_workbook(self.template_path, data_only=True)
            ws_template_readonly = wb_template_readonly[sheet_name]
        except Exception as e:
            logging.error(f"任务 '{job_name}': 无法加载只读模板 '{self.template_path.name}' 或工作表 '{sheet_name}'. 错误: {e}")
            return
        
        # 步骤 4: 循环处理每一个单元格
        logging.info(f"任务 '{job_name}': 准备处理 {len(cell_list)} 个单元格...")
        for cell_address in cell_list:
            cell_readonly = ws_template_readonly[cell_address]
            cell_value = cell_readonly[0].value if isinstance(cell_readonly, tuple) else cell_readonly.value
            text_content = str(cell_value) if cell_value is not None else ""

            if not text_content.strip():
                continue

            # (数据解析部分的逻辑保持不变)
            title_definitions = job_config.get('title_definitions', [])
            is_hierarchical = title_definitions and 'sub_definitions' in title_definitions[0]
            if is_hierarchical:
                fragments = Utils.parse_fragments_hierarchical(text_content, title_definitions, default_style)
            else:
                fragments = Utils.parse_fragments_flat(text_content, title_definitions, default_style)
            
            # 调用核心处理流水线，传入样式ID
            self.xml_helper.run_single_cell_pipeline(
                fragments, 
                sheet_name, 
                cell_address
            )
        logging.info(f"--- 批量任务 '{job_name}' 执行完毕 ---")

    def _generate_cell_list(self, seq_config: dict) -> list[str]:
        """根据序列配置，生成单元格地址列表。"""
        cell_list = []
        start_cell = seq_config.get('start_cell', 'A1')
        step = seq_config.get('step', 1)
        end_row = seq_config.get('end_row', 1)

        col_match = re.match(r"([A-Z]+)", start_cell)
        col = col_match.group(1) if col_match else 'A'
        
        row_match = re.search(r"(\d+)", start_cell)
        start_row = int(row_match.group(1)) if row_match else 1

        for row in range(start_row, end_row + 1, step):
            cell_list.append(f"{col}{row}")
        
        return cell_list

# ======================================================================
#  (新增) 底层XML操作辅助类
# ======================================================================
class _RichTextXMLHelper:
    """
    一个私有的辅助类，封装了所有与.xlsx文件内部XML结构的底层操作。
    """
    def __init__(self, zip_contents: dict, style_configs: dict):
        self.zip_contents = zip_contents
        self.style_configs = style_configs

    def run_single_cell_pipeline(self, fragments: list, sheet_name: str, cell_address: str):
        """(已回退) 在内存中对单个单元格执行富文本注入。"""
        donor_file = 'data/temp/donor.xlsx'
        try:
            if not self._create_rich_text_donor_file(donor_file, fragments):
                raise Exception("创建'供体'文件失败。")
            
            rich_text_element = self._extract_rich_text_from_donor(donor_file)
            if rich_text_element is None:
                raise Exception("从'供体'文件提取富文本失败。")

            # 调用已移除了 style_index 参数的下一层方法
            self._inject_rich_text(rich_text_element, sheet_name, cell_address)

        except Exception as e:
            logging.error(f"处理单元格 {cell_address} 的流程失败: {e}", exc_info=True)
            raise
        finally:
            if os.path.exists(donor_file):
                os.remove(donor_file)
    
    def repack_excel_file(self, output_path: str):
        """将内存中的 self.zip_contents 重新打包成 .xlsx 文件。"""
        logging.info(f"▶️ 开始将修改后的内容重新打包到 {output_path}。")
        try:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                for filename, content in self.zip_contents.items():
                    new_zip.writestr(filename, content)
            logging.info(f"✅ 文件已成功重新打包并保存。")
        except Exception as e:
            logging.error(f"❌ 重新打包文件时发生错误: {e}", exc_info=True)
            raise

        
    def _create_rich_text_donor_file(self, donor_filename: str, fragments: list) -> bool:
        """根据传入的样式配置字典，动态创建多种格式。"""
        workbook = None
        try:
            workbook = xlsxwriter.Workbook(donor_filename)
            worksheet = workbook.add_worksheet()
            format_objects = {name: workbook.add_format(config) for name, config in self.style_configs.items()}
            
            # args_for_writing 的构建逻辑保持不变
            args_for_writing = []
            for text, style_name in fragments:
                if not text: continue
                current_format = format_objects.get(style_name, workbook.add_format())
                args_for_writing.extend([current_format, text])
            
            # 智能判断：根据片段数量选择不同的写入方法
            if len(args_for_writing) > 2:
                # 如果有多个片段（真正的富文本），使用 write_rich_string
                worksheet.write_rich_string('A1', *args_for_writing)
            elif len(args_for_writing) == 2:
                # 如果只有一个片段，使用更基础的 write 方法，它没有“至少两个片段”的限制
                # args_for_writing 在这里是 [format_object, text_string]
                worksheet.write('A1', args_for_writing[1], args_for_writing[0])
            # 如果 args_for_writing 为空，则什么也不写入，创建一个空文件是符合预期的
            
            workbook.close()
            return True
        except Exception as e:
            logging.error(f"创建 '供体' 文件时发生错误: {e}", exc_info=True)
            if workbook: workbook.close()
            return False

    def _extract_rich_text_from_donor(self, donor_path: str) -> ET.Element | None:
        """从"供体"文件中提取富文本XML片段。"""
        SHARED_STRINGS_PATH = 'xl/sharedStrings.xml'
        try:
            with zipfile.ZipFile(donor_path, 'r') as donor_zip:
                with donor_zip.open(SHARED_STRINGS_PATH) as f:
                    xml_content = f.read()
            ET.register_namespace('', NAMESPACES['main'])
            root = ET.fromstring(xml_content)
            return root.find('main:si', NAMESPACES)
        except (KeyError, ET.ParseError, Exception) as e:
            logging.error(f"提取富文本时出错: {e}", exc_info=True)
            return None

    def _inject_rich_text(self, rich_text_element: ET.Element, sheet_name: str, cell_address: str):
        """(已回退) 不再处理样式索引。"""
        new_string_index, new_shared_strings_xml = self._inject_into_shared_strings(rich_text_element)
        self.zip_contents['xl/sharedStrings.xml'] = new_shared_strings_xml

        worksheet_path = self._get_worksheet_path(sheet_name)
        if not worksheet_path:
            raise Exception(f"在文件中找不到名为 '{sheet_name}' 的工作表。")

        original_worksheet_xml = self.zip_contents[worksheet_path]
        # 调用已移除了 style_index 参数的下一层方法
        new_worksheet_xml = self._update_worksheet_cell(original_worksheet_xml, cell_address, new_string_index)
        self.zip_contents[worksheet_path] = new_worksheet_xml

    def _get_worksheet_path(self, sheet_name: str) -> str | None:
        """
        (已更新) 直接从内存中的 self.zip_contents 读取rels和workbook文件。
        此版本能智能处理相对和绝对工作表路径。
        """
        try:
            # 1. 从 workbook.xml.rels 中构建 关系ID -> 目标路径 的映射
            rels_root = ET.fromstring(self.zip_contents['xl/_rels/workbook.xml.rels'])
            # 使用 '*' 通配符来忽略命名空间，使查找更健壮
            id_target_map = {rel.get('Id'): rel.get('Target') for rel in rels_root.findall('{*}Relationship')}
            
            # 2. 从 workbook.xml 中根据工作表名称查找其关系ID (rId)
            workbook_root = ET.fromstring(self.zip_contents['xl/workbook.xml'])
            for sheet in workbook_root.findall('main:sheets/main:sheet', NAMESPACES):
                if sheet.get('name') == sheet_name:
                    # 注意 rId 的属性名包含了完整的命名空间
                    rId = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    target = id_target_map.get(rId)
                    
                    if not target:
                        return None

                    # --- 核心修正逻辑 ---
                    # 3. 智能判断并构建最终路径
                    # 首先清理路径中可能存在的 '../' 前缀
                    clean_target = target.lstrip('../')
                    
                    # 检查清理后的路径是否已经以 'xl/' 开头
                    if clean_target.startswith('xl/'):
                        # 如果是，则它已经是我们需要的完整内部路径，直接返回
                        return clean_target
                    else:
                        # 如果不是，则它是一个相对路径，我们需要为它补上 'xl/' 前缀
                        return f'xl/{clean_target}'
            
            # 如果循环结束都没找到匹配的工作表名称
            return None
        except KeyError as e:
            logging.error(f"    获取工作表路径时，缺少必要的内部文件: {e}")
            return None
        except Exception as e:
            logging.error(f"    获取工作表路径时发生未知错误: {e}", exc_info=True)
            return None

    def _inject_into_shared_strings(self, rich_text_element: ET.Element) -> tuple[int, bytes]:
        """
        (已更新) 将富文本元素注入 self.zip_contents['xl/sharedStrings.xml']。
        如果 sharedStrings.xml 不存在，则会先在内存中创建它。
        """
        shared_strings_path = 'xl/sharedStrings.xml'

        # 检查 'xl/sharedStrings.xml' 是否已存在于内存的zip内容中
        if shared_strings_path in self.zip_contents:
            # 如果存在，则像以前一样正常解析
            root = ET.fromstring(self.zip_contents[shared_strings_path])
        else:
            # 如果不存在，创建一个全新的、最小化的XML结构
            logging.warning(f"模板文件中未找到 '{shared_strings_path}'，将自动在内存中创建。")
            
            # 这是一个符合 a minimal sharedStrings.xml 规范的XML模板字符串
            blank_sst_xml_template = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"></sst>'
            )
            root = ET.fromstring(blank_sst_xml_template)
        
        # 1. 将从 "供体" 文件中提取出的富文本元素(<si>)追加到根元素(<sst>)下
        root.append(rich_text_element)
        
        # 2. 重新计算字符串总数并更新 <sst> 标签的属性
        #    为确保准确，我们直接查找所有 <si> 节点来计数
        all_si_nodes = root.findall('main:si', NAMESPACES)
        count = len(all_si_nodes)
        root.set('count', str(count))
        root.set('uniqueCount', str(count)) # 在简单情况下，这两个值通常相等
        
        # 3. 返回新注入元素的索引（从0开始）和更新后的整个XML文件内容
        new_string_index = count - 1
        new_xml_bytes = ET.tostring(root, encoding='UTF-8', xml_declaration=True)
        
        return new_string_index, new_xml_bytes

    def _update_worksheet_cell(self, worksheet_xml: bytes, cell_address: str, new_index: int) -> bytes:
        """(已回退) 使用文本替换策略更新工作表XML，不再添加 s 属性。"""
        worksheet_str = worksheet_xml.decode('utf-8')
        # 创建一个不包含 s="..." 样式属性的新单元格XML
        new_cell_xml = f'<c r="{cell_address}" t="s"><v>{new_index}</v></c>'
        cell_pattern = re.compile(f'<c r="{cell_address}"[^>]*>.*?</c>', re.DOTALL)
        new_worksheet_str, num_replacements = cell_pattern.subn(new_cell_xml, worksheet_str)
        if num_replacements == 0:
            # (创建新行和新单元格的逻辑...)
            match = re.search(r'(\d+)$', cell_address)
            row_number_str = match.group(1) if match else '1'
            row_pattern = re.compile(f'<row r="{row_number_str}"[^>]*>(.*?)</row>', re.DOTALL)
            row_match = row_pattern.search(worksheet_str)
            if row_match:
                new_row_xml = row_match.group(0).replace('</row>', f'{new_cell_xml}</row>')
                new_worksheet_str = worksheet_str.replace(row_match.group(0), new_row_xml)
            else:
                new_row_xml = f'<row r="{row_number_str}">{new_cell_xml}</row>'
                new_worksheet_str = worksheet_str.replace('</sheetData>', f'{new_row_xml}</sheetData>')
        return new_worksheet_str.encode('utf-8')
    
    

    @staticmethod
    def _get_style_index_from_donor(style_properties: dict) -> str | None:
        """
        [新增] 创建一个临时的“样式供体”文件，专门用于提取应用了特定格式的单元格的样式索引(s attribute)。
        """
        style_donor_file =  'data/temp/style_donor.xlsx'
        workbook = None
        try:
            # 1. 创建一个包含特定样式的 donor 文件
            workbook = xlsxwriter.Workbook(style_donor_file)
            worksheet = workbook.add_worksheet()
            cell_format = workbook.add_format(style_properties)
            worksheet.write('A1', 'dummy', cell_format) # 写入带样式的虚拟内容
            workbook.close()

            # 2. 解压 donor 文件并从中提取 's' 属性
            with zipfile.ZipFile(style_donor_file, 'r') as zip_ref:
                sheet_xml_content = zip_ref.read('xl/worksheets/sheet1.xml')
                root = ET.fromstring(sheet_xml_content)
                # 使用 .// 查找 A1 单元格
                cell_a1 = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c[@r="A1"]')
                if cell_a1 is not None:
                    style_index = cell_a1.get('s')
                    logging.info(f"成功从样式供体文件中提取到样式索引: '{style_index}'")
                    return style_index
            return None
        except Exception as e:
            logging.error(f"提取样式索引时出错: {e}", exc_info=True)
            return None
        finally:
            if os.path.exists(style_donor_file):
                os.remove(style_donor_file)