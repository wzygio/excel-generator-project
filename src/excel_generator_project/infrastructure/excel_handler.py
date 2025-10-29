# src/excel_writer.py
import logging

from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color

class ExcelHandler:
    def __init__(self, template_path=None):
        self.template_path = template_path
        self.wb = None
        self.ws = None

    def read_cell(self, cell_address):
        """读取指定单元格的内容"""
        if not self.ws:
            raise ValueError("请先加载模板文件")
        cell = self.ws[cell_address]
        # 检查是否为合并单元格
        if isinstance(cell, tuple):
            # 如果是合并单元格，读取第一个单元格的值
            return cell[0].value
        return cell.value
    
    def write_cell(self, cell_address, value):
        """向指定单元格写入数据"""
        if not self.ws or not cell_address:
            raise ValueError("Invalid worksheet or cell address")
        cell = self.ws[cell_address]

        # 检查是否为合并单元格
        if isinstance(cell, tuple):
            # 如果是合并单元格，只对第一个单元格赋值
            cell = cell[0]
        cell.value = value
        # 设置样式
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

    def write_data(self, data, start_row, start_col):
        """将数据写入Excel"""
        for r_idx, row in enumerate(data.values, start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                if not self.ws:
                    raise ValueError("Invalid worksheet or cell address")
                cell = self.ws.cell(row=r_idx, column=c_idx, value=value)
                # 设置样式
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))


    def load_template(self):
        """加载模板文件"""
        if not self.template_path:
            raise ValueError("模板文件路径不能为空")
        try:
            self.wb = load_workbook(self.template_path)
            self.ws = self.wb.active
        except Exception as e:
            raise ValueError(f"加载模板文件时出错：{e}")
        
    def set_active_sheet(self, sheet_name: str):
        """
        根据名称设置当前活动的工作表。
        
        :param sheet_name: 目标工作表的名称
        """
        if not self.wb:
            raise ValueError("工作簿未加载，请先调用 load_template()")
        
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用的工作表: {self.wb.sheetnames}")
        
        self.ws = self.wb[sheet_name]
        logging.info(f"活动工作表已切换至: '{sheet_name}'")
        
    def create_new_file(self, output_path, cell_value, target_cell="B2"):
        """创建一个新的 Excel 文件，并写入指定内容"""
        new_wb = Workbook()
        new_ws = new_wb.active
        if new_ws is None:
            raise ValueError("无法创建工作表")
        new_ws[target_cell] = cell_value  # 写入内容到目标单元格
        new_wb.save(output_path)

    def save(self, file_path):
        """
        保存当前工作簿到指定路径。
        :param file_path: 保存的目标文件路径
        """
        if not self.wb:
            raise ValueError("没有可保存的工作簿，请先加载或创建一个工作簿")
        try:
            self.wb.save(file_path)
            logging.info(f"文件已保存至: {file_path}")
        except Exception as e:
            raise ValueError(f"保存文件时出错：{e}")