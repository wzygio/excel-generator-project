# src/generators/report_generator.py
import logging
from pathlib import Path
import shutil

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 假设 ExcelHandler 和 Utils 的路径如下
from excel_generator_project.infrastructure.excel_handler import ExcelHandler
from excel_generator_project.utils.utils import Utils

class ReportGenerator:
    """
    报告生成器。负责接收来自各个处理器的数据，并将其写入到最终的Excel报告中。
    """
    def __init__(self, template_path: Path, output_path: Path, config: dict):
        """
        初始化报告生成器。
        """
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.config = config
        self.excel_handler = None
        # 生成器也需要产品型号列表来确保写入顺序
        self.product_models = Utils.extract_product_models(config.get('model_definitions', []))
        logging.info("ReportGenerator 初始化成功。")

    def run(self):
        """
        主调度方法，负责完整的“文件准备 -> 任务执行 -> 文件保存”流程。
        """
        logging.info("--- ReportGenerator 开始运行 ---")

        # 步骤 1: 准备输出文件和Excel处理器
        try:
            logging.info(f"正在将模板复制到输出路径: {self.output_path}")
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy(self.template_path, self.output_path)
            
            self.excel_handler = ExcelHandler(str(self.output_path))
            self.excel_handler.load_template()
        except Exception as e:
            logging.error(f"准备输出文件时出错: {e}", exc_info=True)
            return

        # 步骤 2: 按顺序执行所有在配置中定义的写入任务
        jobs = self.config.get('report_generator_jobs', [])
        logging.info(f"发现 {len(jobs)} 个报告生成任务...")
        for i, job_config in enumerate(jobs, 1):
            job_type = job_config.get("job_type")
            logging.info(f"--> 开始执行任务 {i}/{len(jobs)}: 类型 = {job_type}")

            if job_type == "write_processed_exceptions":
                self._execute_write_processed_exceptions(job_config)
            elif job_type == "write_summary_reports":
                self._execute_write_summary_reports(job_config)
            else:
                logging.warning(f"未知的任务类型 '{job_type}'，已跳过。")

        # 步骤 3: 保存所有修改到输出文件
        try:
            logging.info(f"所有任务执行完毕，正在将结果保存到文件: {self.output_path.name}")
            self.excel_handler.save(self.output_path)
        except Exception as e:
            logging.error(f"保存文件时出错: {e}", exc_info=True)
            
        logging.info("--- ReportGenerator 运行结束 ---")

    def _execute_write_processed_exceptions(self, job_config: dict):
        """
        将 ExceptionProcessor 处理好的异常模块文本，按顺序写回到目标Excel中。
        """
        if not self.excel_handler:
            logging.error("  ExcelHandler 未初始化，无法执行写入任务。")
            return

        # 1. 从配置中获取数据源的键名
        data_source_key = job_config.get('data_source_key')
        # 从注入的外部数据中获取 ExceptionProcessor 的处理结果
        processed_data = self.config.get('EXTERNAL_DATA', {}).get(data_source_key, {})

        if not processed_data:
            logging.warning("  未在配置中找到有效的异常数据源，写入任务跳过。")
            return

        logging.info(f"  开始写入 '{data_source_key}' 的数据...")
        sheet_name = job_config['sheet_name']
        data_column_name = job_config['data_column']
        rules = job_config.get('sequential_extraction_rules', {})
        module_definitions = rules.get('module_definitions', [])

         # --- 核心修改 1: 在方法开头，获取默认的后备文本模板 ---
        default_text = self.config.get("report_text_templates", {}).get(
            "default_empty_exception_module", 
            "错误: 默认模板 'default_empty_exception_module' 未在配置中定义。"
        )

        try:
            self.excel_handler.set_active_sheet(sheet_name)
            if not self.excel_handler.ws:
                logging.error(f"  工作表 '{sheet_name}' 不存在。")
                return

            # 2. 从表头找到目标列的索引，进而得到列字母
            header_row = self.excel_handler.ws[1]
            col_idx = None
            for cell in header_row:
                if cell.value == data_column_name:
                    col_idx = cell.column
                    break
            
            if col_idx is None:
                logging.error(f"  在工作表 '{sheet_name}' 中未找到列标题 '{data_column_name}'。")
                return
            
            col_letter = get_column_letter(col_idx)

        except Exception as e:
            logging.error(f"  准备写入时出错（例如，工作表不存在）: {e}", exc_info=True)
            return

        # 3. 遍历产品列表，计算位置并写入数据
        for product_index, model in enumerate(self.product_models):
            model_data = processed_data.get(model, {})
            previous_exceptions = model_data.get('previous_exceptions', {})
            
            for module_def in module_definitions:
                key_name = module_def['key_name']
                text_to_write = previous_exceptions.get(key_name, default_text)

                start_row = module_def['start_row']
                step = module_def['step']
                target_row = start_row + (product_index * step)
                
                target_cell_address = f"{col_letter}{target_row}"
                logging.info(f"    正在将产品 '{model}' 的 '{key_name}' 写入到单元格 {target_cell_address}")
                self.excel_handler.write_cell(target_cell_address, text_to_write)
                
                # 设置单元格样式
                self.excel_handler.ws[target_cell_address].alignment = Alignment(
                    wrap_text=True, vertical='top', horizontal='left'
                )
                
    # --- 新增的任务执行方法 ---
    def _execute_write_summary_reports(self, job_config: dict):
        """
        将 DataProcessor 准备好的摘要数据，填充模板并写入到Excel中。
        """
        logging.info("  开始执行'写入摘要报告'任务...")

        # 1. 从配置中获取数据源的键名和写入参数
        data_source_key = job_config.get('data_source_key')
        # 从主流程注入的外部数据中，获取DataProcessor的处理结果
        summary_data = self.config.get('EXTERNAL_DATA', {}).get(data_source_key, [])

        if not summary_data:
            logging.warning("  未找到有效的摘要数据，写入任务跳过。")
            return

        template_name = job_config.get("template_name")
        target_cells_cfg = job_config.get("target_cells", {})
        start_cell = target_cells_cfg.get("start_cell")
        step = target_cells_cfg.get("step")

        # 2. 获取模板字符串
        template_string = self.config.get("report_text_templates", {}).get(template_name)
        if not all([start_cell, step, template_string]):
            logging.error("  'write_summary_reports' 任务配置不完整，任务中止。")
            return

         # 3. 循环填充模板并写入到对应的Excel单元格
        col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))

        logging.info(f"    将为 {len(summary_data)} 个产品型号写入摘要报告...")
        for i, data_dict in enumerate(summary_data):
            current_row = start_row + i * step
            cell_address = f"{col_letter}{current_row}"
            
            # 添加类型检查
            if not isinstance(data_dict, dict):
                logging.error(f"数据项 {i} 不是字典类型，而是 {type(data_dict)}")
                continue

            # 使用更安全的 _format_string 辅助方法填充模板
            final_text = Utils.format_string(template_string, data_dict)
            
            # 写入单元格
            if not self.excel_handler or not self.excel_handler.ws:
                logging.error("  ExcelHandler 未初始化或工作表不存在，无法写入单元格。")
                return
            
            self.excel_handler.write_cell(cell_address, final_text)
            # 设置单元格样式 (与你最初的实现完全一致)
            self.excel_handler.ws[cell_address].alignment = Alignment(
                wrap_text=True, vertical='top', horizontal='left'
            )
            logging.info(f"      -> 已将型号 '{data_dict.get('model_name', '')}' 的报告写入单元格 {cell_address}")

        logging.info("    所有摘要报告均已写入完毕。")