"""Shared Excel sheet reader for standard and enterprise-encrypted workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

XLSX_MAGIC = b"PK\x03\x04"


class ExcelReadError(Exception):
    """Raised when an Excel sheet cannot be read."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True)
class SheetData:
    """In-memory worksheet values."""

    path: Path
    sheet_name: str
    rows: list[list[Any]]


class ExcelSheetReader:
    """Read standard xlsx or local enterprise-encrypted Excel via COM."""

    def read_sheet(self, path: Path, sheet_name: str | None = None) -> SheetData:
        path = Path(path)
        if not path.exists():
            raise ExcelReadError(
                "excel.file.missing",
                f"Excel file does not exist: {path}",
                details={"path": str(path), "sheet_name": sheet_name},
            )
        if _is_standard_xlsx(path):
            return self._read_with_openpyxl(path, sheet_name)
        return self._read_with_com(path, sheet_name)

    def _read_with_openpyxl(self, path: Path, sheet_name: str | None) -> SheetData:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            actual_sheet = _resolve_sheet_name(workbook.sheetnames, sheet_name)
            worksheet = workbook[actual_sheet]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            return SheetData(path=path, sheet_name=actual_sheet, rows=rows)
        except ExcelReadError:
            raise
        except Exception as exc:
            raise ExcelReadError(
                "excel.file.read_failed",
                f"Failed to read Excel file: {path}",
                details={"path": str(path), "sheet_name": sheet_name, "error": str(exc)},
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()

    def _read_with_com(self, path: Path, sheet_name: str | None) -> SheetData:
        try:
            import pythoncom  # type: ignore[import-not-found]
            import win32com.client  # type: ignore[import-not-found]
        except ImportError as exc:
            raise ExcelReadError(
                "excel.file.com_unavailable",
                "Reading encrypted Excel requires pywin32 and local Excel COM.",
                details={"path": str(path), "sheet_name": sheet_name},
            ) from exc

        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(path.resolve()), 0, True)
            sheet_names = [
                str(workbook.Worksheets(index).Name)
                for index in range(1, workbook.Worksheets.Count + 1)
            ]
            actual_sheet = _resolve_sheet_name(sheet_names, sheet_name)
            worksheet = workbook.Worksheets(actual_sheet)
            rows = _com_values_to_rows(worksheet.UsedRange.Value)
            return SheetData(path=path, sheet_name=actual_sheet, rows=rows)
        except ExcelReadError:
            raise
        except Exception as exc:
            raise ExcelReadError(
                "excel.file.com_read_failed",
                f"Failed to read Excel file through COM: {path}",
                details={"path": str(path), "sheet_name": sheet_name, "error": str(exc)},
            ) from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()


def _resolve_sheet_name(sheet_names: list[str], requested: str | None) -> str:
    if not sheet_names:
        raise ExcelReadError("excel.file.no_sheets", "Excel workbook has no sheets.")
    if requested is None:
        return sheet_names[0]
    if requested in sheet_names:
        return requested
    requested_norm = requested.strip().lower()
    for sheet_name in sheet_names:
        if sheet_name.strip().lower() == requested_norm:
            return sheet_name
    raise ExcelReadError(
        "excel.file.missing_sheet",
        f"Excel workbook is missing sheet: {requested}",
        details={"requested": requested, "available_sheets": sheet_names},
    )


def _is_standard_xlsx(path: Path) -> bool:
    try:
        with Path(path).open("rb") as file:
            return file.read(4) == XLSX_MAGIC
    except OSError:
        return False


def _com_values_to_rows(raw_values: Any) -> list[list[Any]]:
    if raw_values is None:
        return []
    if not isinstance(raw_values, tuple):
        return [[raw_values]]
    if raw_values and not isinstance(raw_values[0], tuple):
        return [list(raw_values)]
    return [list(row) for row in raw_values]
