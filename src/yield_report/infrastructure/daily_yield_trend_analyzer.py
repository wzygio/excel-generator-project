"""Deterministic yield trend analyzer for monthly/weekly/daily reports."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

XLSX_MAGIC = b"PK\x03\x04"


class DailyYieldTrendAnalysisError(Exception):
    """Raised when yield trend analysis cannot be completed."""


@dataclass
class DailyYieldPoint:
    label: str
    yield_value: float | None

    @property
    def is_valid(self) -> bool:
        return self.yield_value is not None and self.yield_value > 0


@dataclass
class YieldDriver:
    metric_name: str
    first_label: str
    last_label: str
    first_value: float
    last_value: float
    delta: float
    reason: str


@dataclass
class DailyYieldTrendResult:
    product_model: str
    file_path: Path
    sheet_name: str
    metric_name: str
    points: list[DailyYieldPoint]
    result_text: str
    time_grain: str = "daily"
    requested_periods: int | None = None
    actual_period_count: int = 0
    warnings: list[str] = field(default_factory=list)
    drivers: list[YieldDriver] = field(default_factory=list)


class DailyYieldTrendAnalyzer:
    """Read a yield row and produce a concise trend conclusion."""

    preferred_sheets = ("屏体综合良率", "MVI", "CT")
    preferred_metrics = ("屏体综合良率", "MVI良率", "CT良率")

    def can_handle(
        self,
        *,
        user_query: str,
        target_metrics: list[str],
        analysis_logic: str,
        time_grain: str = "",
    ) -> bool:
        text = user_query.upper()
        metric_text = " ".join(target_metrics).upper()
        has_ct = "CT" in text or "CT" in metric_text
        has_yield = (
            "良率" in user_query
            or "良率" in " ".join(target_metrics)
            or "YIELD" in text
            or "YIELD" in metric_text
        )
        has_trend = any(keyword in user_query for keyword in ["趋势", "变化", "波动"]) or (
            "趋势" in analysis_logic or "TREND" in analysis_logic.upper()
        )
        return has_yield and has_trend and not has_ct

    def analyze(
        self,
        file_path: Path,
        product_model: str,
        days: int = 7,
        time_grain: str = "daily",
        requested_periods: int | None = None,
    ) -> DailyYieldTrendResult:
        normalized_grain = _normalize_time_grain(time_grain)
        periods = requested_periods or (days if normalized_grain == "daily" else 3)
        rows_by_sheet = self._read_candidate_sheets(file_path)
        for sheet_name in self.preferred_sheets:
            rows = rows_by_sheet.get(sheet_name)
            if not rows:
                continue
            try:
                return self._analyze_rows(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    rows=rows,
                    product_model=product_model,
                    time_grain=normalized_grain,
                    requested_periods=periods,
                )
            except DailyYieldTrendAnalysisError:
                logger.debug("Yield sheet not usable: %s", sheet_name, exc_info=True)
                continue

        raise DailyYieldTrendAnalysisError(
            f"Cannot find {normalized_grain} yield trend row for product={product_model}"
        )

    def _analyze_rows(
        self,
        *,
        file_path: Path,
        sheet_name: str,
        rows: list[list[Any]],
        product_model: str,
        time_grain: str,
        requested_periods: int,
    ) -> DailyYieldTrendResult:
        header_index = self._find_header_row(rows)
        header = rows[header_index]
        product_col = self._find_header_col(header, ["ProductCode", "产品型号"])
        time_cols = self._find_time_columns(header, time_grain)
        if product_col is None:
            raise DailyYieldTrendAnalysisError("Cannot identify product column")
        if not time_cols:
            raise DailyYieldTrendAnalysisError(f"Cannot identify {time_grain} columns")

        row, metric_name = self._find_yield_row(
            rows=rows,
            start_index=header_index + 1,
            product_col=product_col,
            product_model=product_model,
        )
        chosen_cols = time_cols[-requested_periods:]
        points = [
            DailyYieldPoint(
                label=str(header[col_index]).strip(),
                yield_value=_to_float(_cell(row, col_index)),
            )
            for col_index in chosen_cols
        ]
        drivers = self._find_potential_drivers(
            rows=rows,
            start_index=header_index + 1,
            product_col=product_col,
            product_model=product_model,
            chosen_cols=chosen_cols,
            header=header,
            main_metric_name=metric_name,
        )
        warnings: list[str] = []
        if len(chosen_cols) < requested_periods:
            warnings.append(
                f"请求 {requested_periods} 个{_period_unit(time_grain)}，源表仅提供 "
                f"{len(chosen_cols)} 个{_grain_label(time_grain)}周期。"
            )
        result_text = self._format_result(
            file_path=file_path,
            product_model=product_model,
            sheet_name=sheet_name,
            metric_name=metric_name,
            points=points,
            time_grain=time_grain,
            requested_periods=requested_periods,
            warnings=warnings,
            drivers=drivers,
        )
        return DailyYieldTrendResult(
            product_model=product_model,
            file_path=file_path,
            sheet_name=sheet_name,
            metric_name=metric_name,
            points=points,
            result_text=result_text,
            time_grain=time_grain,
            requested_periods=requested_periods,
            actual_period_count=len(chosen_cols),
            warnings=warnings,
            drivers=drivers,
        )

    def _read_candidate_sheets(self, file_path: Path) -> dict[str, list[list[Any]]]:
        if not _is_standard_xlsx(file_path):
            raise DailyYieldTrendAnalysisError("Yield analyzer requires standard xlsx")

        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            rows_by_sheet: dict[str, list[list[Any]]] = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                reset_dimensions = getattr(worksheet, "reset_dimensions", None)
                if callable(reset_dimensions):
                    try:
                        reset_dimensions()
                    except Exception:
                        logger.debug("Unable to reset worksheet dimensions", exc_info=True)
                rows_by_sheet[sheet_name] = [list(row) for row in worksheet.iter_rows(values_only=True)]
            return rows_by_sheet
        finally:
            workbook.close()

    @staticmethod
    def _find_header_row(rows: list[list[Any]]) -> int:
        for index, row in enumerate(rows[:12]):
            joined = " ".join(str(cell) for cell in row if cell is not None)
            if "ProductCode" in joined or "产品型号" in joined:
                return index
        raise DailyYieldTrendAnalysisError("Cannot identify header row")

    @staticmethod
    def _find_header_col(row: list[Any], keywords: list[str]) -> int | None:
        for index, cell in enumerate(row):
            text = str(cell or "")
            if any(keyword in text for keyword in keywords):
                return index
        return None

    @staticmethod
    def _find_time_columns(row: list[Any], time_grain: str) -> list[int]:
        pattern = {
            "monthly": r"^M\d{1,2}$",
            "weekly": r"^W\d{1,2}$",
            "daily": r"^\d{1,2}/\d{1,2}$",
        }[_normalize_time_grain(time_grain)]
        columns: list[int] = []
        for index, cell in enumerate(row):
            text = str(cell or "").strip().upper()
            if re.match(pattern, text):
                columns.append(index)
        return columns

    @staticmethod
    def _find_daily_columns(row: list[Any]) -> list[int]:
        return DailyYieldTrendAnalyzer._find_time_columns(row, "daily")

    def _find_yield_row(
        self,
        *,
        rows: list[list[Any]],
        start_index: int,
        product_col: int,
        product_model: str,
    ) -> tuple[list[Any], str]:
        current_product = ""
        expected_product = product_model.upper()
        for row in rows[start_index:]:
            product_value = str(_cell(row, product_col) or "").strip()
            if product_value:
                current_product = product_value.upper()
            if current_product != expected_product:
                continue
            for metric_name in self.preferred_metrics:
                if any(str(cell or "").strip() == metric_name for cell in row[:8]):
                    return row, metric_name
        raise DailyYieldTrendAnalysisError(f"Cannot find yield row for product={product_model}")

    def _find_potential_drivers(
        self,
        *,
        rows: list[list[Any]],
        start_index: int,
        product_col: int,
        product_model: str,
        chosen_cols: list[int],
        header: list[Any],
        main_metric_name: str,
    ) -> list[YieldDriver]:
        if len(chosen_cols) < 2:
            return []

        current_product = ""
        expected_product = product_model.upper()
        first_col = chosen_cols[0]
        last_col = chosen_cols[-1]
        first_label = str(_cell(header, first_col) or "").strip()
        last_label = str(_cell(header, last_col) or "").strip()
        drivers: list[YieldDriver] = []
        for row in rows[start_index:]:
            product_value = str(_cell(row, product_col) or "").strip()
            if product_value:
                current_product = product_value.upper()
            if current_product != expected_product:
                continue

            metric_name = _extract_metric_label(row, product_col, expected_product)
            if not metric_name or metric_name == main_metric_name:
                continue
            first_value = _to_float(_cell(row, first_col))
            last_value = _to_float(_cell(row, last_col))
            if first_value is None or last_value is None:
                continue
            delta = last_value - first_value
            reason = _driver_reason(metric_name, delta)
            if not reason:
                continue
            drivers.append(
                YieldDriver(
                    metric_name=metric_name,
                    first_label=first_label,
                    last_label=last_label,
                    first_value=first_value,
                    last_value=last_value,
                    delta=delta,
                    reason=reason,
                )
            )

        drivers.sort(key=lambda item: abs(item.delta), reverse=True)
        return drivers[:3]

    @staticmethod
    def _format_result(
        *,
        file_path: Path,
        product_model: str,
        sheet_name: str,
        metric_name: str,
        points: list[DailyYieldPoint],
        time_grain: str,
        requested_periods: int,
        warnings: list[str],
        drivers: list[YieldDriver],
    ) -> str:
        valid_points = [point for point in points if point.is_valid]
        grain_label = _grain_label(time_grain)
        period_name = _period_name(time_grain)
        title = _format_title(product_model, time_grain, requested_periods)
        lines = [
            title,
            "",
            f"- 数据源: `{file_path}`",
            f"- Sheet: `{sheet_name}`",
            f"- 指标行: `{metric_name}`",
            "",
            f"| {period_name} | {grain_label}良率 | 数据状态 |",
            "| --- | ---: | --- |",
        ]
        for point in points:
            yield_text = "-" if point.yield_value is None else _format_percent(point.yield_value)
            status = "有效" if point.is_valid else "无有效良率"
            lines.append(f"| {point.label} | {yield_text} | {status} |")

        lines.extend(["", "### 核心结论"])
        for warning in warnings:
            lines.append(f"- 数据限制: {warning}")

        if len(valid_points) >= 2:
            first = valid_points[0]
            last = valid_points[-1]
            delta = (last.yield_value or 0) - (first.yield_value or 0)
            average = sum(point.yield_value or 0 for point in valid_points) / len(valid_points)
            min_point = min(valid_points, key=lambda point: point.yield_value or 0)
            max_point = max(valid_points, key=lambda point: point.yield_value or 0)
            direction = "上升" if delta > 0 else "下降" if delta < 0 else "持平"
            lines.append(f"- 有效{period_name} {len(valid_points)}/{len(points)} 个。")
            lines.append(
                f"- 良率从 {first.label} 的 {_format_percent(first.yield_value)} "
                f"到 {last.label} 的 {_format_percent(last.yield_value)}，"
                f"{direction} {_format_pct_points(delta)}。"
            )
            lines.append(
                f"- 当前可用{grain_label}平均良率 {_format_percent(average)}；最低为 {min_point.label} "
                f"{_format_percent(min_point.yield_value)}，最高为 {max_point.label} "
                f"{_format_percent(max_point.yield_value)}。"
            )
            if delta < 0:
                lines.append("- 恶化判断: 末期良率低于首期，存在恶化。")
                lines.extend(_format_driver_lines(drivers))
            else:
                lines.append("- 恶化判断: 末期良率未低于首期，未观察到整体恶化。")
        elif len(valid_points) == 1:
            only = valid_points[0]
            lines.append(
                f"- 只有 1 个{period_name}存在有效良率: {only.label} "
                f"{_format_percent(only.yield_value)}，不足以判断连续趋势。"
            )
        else:
            lines.append(f"- 当前{grain_label}周期没有有效良率数据，无法判断趋势。")
        return "\n".join(lines)


def _is_standard_xlsx(path: Path) -> bool:
    try:
        with Path(path).open("rb") as file:
            return file.read(4) == XLSX_MAGIC
    except OSError:
        return False


def _cell(row: list[Any] | None, index: int | None) -> Any:
    if row is None or index is None or index >= len(row):
        return None
    return row[index]


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_percent(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.2f}%"


def _format_pct_points(value: float) -> str:
    sign = "+" if value > 0 else ""
    return f"{sign}{value * 100:.2f} pct"


def _normalize_time_grain(time_grain: str) -> str:
    normalized = str(time_grain or "").strip().lower()
    if normalized in {"monthly", "month", "m", "月", "月度"}:
        return "monthly"
    if normalized in {"weekly", "week", "w", "周", "周度"}:
        return "weekly"
    return "daily"


def _grain_label(time_grain: str) -> str:
    return {"monthly": "月度", "weekly": "周度", "daily": "日度"}[
        _normalize_time_grain(time_grain)
    ]


def _period_name(time_grain: str) -> str:
    return {"monthly": "月份", "weekly": "周别", "daily": "日期"}[
        _normalize_time_grain(time_grain)
    ]


def _period_unit(time_grain: str) -> str:
    return {"monthly": "月", "weekly": "周", "daily": "天"}[
        _normalize_time_grain(time_grain)
    ]


def _format_title(product_model: str, time_grain: str, requested_periods: int) -> str:
    normalized = _normalize_time_grain(time_grain)
    if normalized == "daily" and requested_periods == 7:
        return f"## {product_model} 最近一周日度良率变化趋势"
    return (
        f"## {product_model} 最近{requested_periods}个{_period_unit(normalized)}"
        f"{_grain_label(normalized)}良率变化趋势"
    )


def _extract_metric_label(row: list[Any], product_col: int, expected_product: str) -> str:
    ignored = {expected_product, "屏体", "MVI", "CT", "ARRAY", "MODULE", "CELL"}
    for index, cell in enumerate(row[:8]):
        if index == product_col:
            continue
        text = str(cell or "").strip()
        if not text or text.upper() in ignored:
            continue
        if any(keyword in text for keyword in ["良率", "不良", "占比", "NG", "Fail", "FAIL"]):
            return text
    return ""


def _driver_reason(metric_name: str, delta: float) -> str:
    upper = metric_name.upper()
    if "良率" in metric_name or "YIELD" in upper:
        return "细分良率下降" if delta < 0 else ""
    if any(keyword in metric_name for keyword in ["不良", "占比"]) or any(
        keyword in upper for keyword in ["NG", "FAIL"]
    ):
        return "不良或占比上升" if delta > 0 else ""
    return ""


def _format_driver_lines(drivers: list[YieldDriver]) -> list[str]:
    if not drivers:
        return [
            "- 恶化原因线索: 当前源表未提供足够的同产品细分变化行来直接归因；"
            "建议结合不良 Code、异常闭环表或制程站别明细继续定位。"
        ]

    lines = ["- 恶化原因线索: 同源表中与恶化方向一致的细分项如下，需结合现场/异常记录复核。"]
    for driver in drivers:
        lines.append(
            f"  - {driver.metric_name}: {driver.first_label} {_format_percent(driver.first_value)} "
            f"-> {driver.last_label} {_format_percent(driver.last_value)}，"
            f"{driver.reason} {_format_pct_points(driver.delta)}。"
        )
    return lines
