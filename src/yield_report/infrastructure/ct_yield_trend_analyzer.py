"""Deterministic CT yield trend analyzer for monthly/weekly/daily yield reports."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

XLSX_MAGIC = b"PK\x03\x04"


class CtYieldTrendAnalysisError(Exception):
    """Raised when CT yield trend analysis cannot be completed."""


@dataclass
class CtYieldPoint:
    label: str
    yield_value: float | None
    output_count: float | None

    @property
    def has_valid_output(self) -> bool:
        if self.output_count is None:
            return self.yield_value is not None and self.yield_value > 0
        return self.output_count > 0 and self.yield_value is not None


@dataclass
class CtYieldTrendResult:
    product_model: str
    file_path: Path
    sheet_name: str
    points: list[CtYieldPoint]
    result_text: str


class CtYieldTrendAnalyzer:
    """Read CT daily-yield rows and produce a concise trend conclusion."""

    def can_handle(
        self,
        *,
        user_query: str,
        target_metrics: list[str],
        analysis_logic: str,
    ) -> bool:
        text = user_query.upper()
        metric_text = " ".join(target_metrics).upper()
        logic_text = analysis_logic.upper()
        has_ct_yield = ("CT" in text or "CT" in metric_text) and (
            "良率" in user_query or "良率" in " ".join(target_metrics)
        )
        has_trend = any(keyword in user_query for keyword in ["趋势", "变化", "波动"]) or (
            "趋势" in analysis_logic or "TREND" in logic_text
        )
        return has_ct_yield and has_trend

    def analyze(
        self,
        file_path: Path,
        product_model: str,
        days: int = 7,
    ) -> CtYieldTrendResult:
        rows = self._read_sheet_rows(file_path, sheet_name="CT")
        if not rows:
            raise CtYieldTrendAnalysisError("CT sheet is empty")

        header_index = self._find_header_row(rows)
        header = rows[header_index]
        product_col = self._find_header_col(header, ["ProductCode", "产品型号"])
        operation_col = self._find_header_col(header, ["Operation", "站点"])
        metric_col = self._find_header_col(header, ["Factory", "归属工厂"])
        date_cols = self._find_daily_columns(header)

        if product_col is None or metric_col is None:
            raise CtYieldTrendAnalysisError("Cannot identify product or metric columns in CT sheet")
        if not date_cols:
            raise CtYieldTrendAnalysisError("Cannot identify daily columns in CT sheet")

        chosen_cols = date_cols[-days:]
        yield_row = self._find_metric_row(
            rows=rows,
            start_index=header_index + 1,
            product_col=product_col,
            operation_col=operation_col,
            metric_col=metric_col,
            product_model=product_model,
            metric_name="CT良率",
        )
        output_row = self._find_metric_row(
            rows=rows,
            start_index=header_index + 1,
            product_col=product_col,
            operation_col=operation_col,
            metric_col=metric_col,
            product_model=product_model,
            metric_name="CT产出数",
        )

        points: list[CtYieldPoint] = []
        for col_index in chosen_cols:
            points.append(
                CtYieldPoint(
                    label=str(header[col_index]).strip(),
                    yield_value=_to_float(_cell(yield_row, col_index)),
                    output_count=_to_float(_cell(output_row, col_index)) if output_row else None,
                )
            )

        result_text = self._format_result(
            file_path=file_path,
            product_model=product_model,
            sheet_name="CT",
            points=points,
        )
        return CtYieldTrendResult(
            product_model=product_model,
            file_path=file_path,
            sheet_name="CT",
            points=points,
            result_text=result_text,
        )

    def _read_sheet_rows(self, file_path: Path, sheet_name: str) -> list[list[Any]]:
        if _is_standard_xlsx(file_path):
            return self._read_sheet_rows_openpyxl(file_path, sheet_name)
        return self._read_sheet_rows_com(file_path, sheet_name)

    @staticmethod
    def _read_sheet_rows_openpyxl(file_path: Path, sheet_name: str) -> list[list[Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise CtYieldTrendAnalysisError(f"Sheet not found: {sheet_name}")
            worksheet = workbook[sheet_name]
            return [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    @staticmethod
    def _read_sheet_rows_com(file_path: Path, sheet_name: str) -> list[list[Any]]:
        try:
            import pythoncom  # type: ignore[import-untyped]
            import win32com.client  # type: ignore[import-untyped]
        except ImportError as exc:
            raise CtYieldTrendAnalysisError(
                "Enterprise-encrypted Excel requires pywin32/Excel COM"
            ) from exc

        excel = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(Path(file_path).resolve()))
            worksheet = workbook.Sheets(sheet_name)
            row_count = int(worksheet.UsedRange.Rows.Count)
            col_count = int(worksheet.UsedRange.Columns.Count)
            rows: list[list[Any]] = []
            for row_index in range(1, row_count + 1):
                rows.append(
                    [worksheet.Cells(row_index, col_index).Value for col_index in range(1, col_count + 1)]
                )
            return rows
        except Exception as exc:
            raise CtYieldTrendAnalysisError(f"Excel COM read failed: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    @staticmethod
    def _find_header_row(rows: list[list[Any]]) -> int:
        for index, row in enumerate(rows[:12]):
            joined = " ".join(str(cell) for cell in row if cell is not None)
            if ("ProductCode" in joined or "产品型号" in joined) and (
                "Factory" in joined or "归属工厂" in joined
            ):
                return index
        raise CtYieldTrendAnalysisError("Cannot identify CT sheet header row")

    @staticmethod
    def _find_header_col(row: list[Any], keywords: list[str]) -> int | None:
        for index, cell in enumerate(row):
            text = str(cell or "")
            if any(keyword in text for keyword in keywords):
                return index
        return None

    @staticmethod
    def _find_daily_columns(row: list[Any]) -> list[int]:
        columns: list[int] = []
        for index, cell in enumerate(row):
            text = str(cell or "").strip()
            if re.match(r"^\d{1,2}/\d{1,2}$", text):
                columns.append(index)
        return columns

    @staticmethod
    def _find_metric_row(
        *,
        rows: list[list[Any]],
        start_index: int,
        product_col: int,
        operation_col: int | None,
        metric_col: int,
        product_model: str,
        metric_name: str,
    ) -> list[Any]:
        current_product = ""
        current_operation = ""
        expected_product = product_model.upper()

        for row in rows[start_index:]:
            product_value = str(_cell(row, product_col) or "").strip()
            operation_value = str(_cell(row, operation_col) or "").strip() if operation_col is not None else ""
            metric_value = str(_cell(row, metric_col) or "").strip()

            if product_value:
                current_product = product_value.upper()
            if operation_value:
                current_operation = operation_value.upper()

            if (
                current_product == expected_product
                and (not current_operation or current_operation == "CT")
                and metric_value == metric_name
            ):
                return row

        raise CtYieldTrendAnalysisError(
            f"Cannot find metric row: product={product_model}, metric={metric_name}"
        )

    @staticmethod
    def _format_result(
        *,
        file_path: Path,
        product_model: str,
        sheet_name: str,
        points: list[CtYieldPoint],
    ) -> str:
        valid_points = [point for point in points if point.has_valid_output]
        total_output = sum(point.output_count or 0 for point in points)

        lines = [
            f"## {product_model} 近一周日度 CT 良率变化趋势",
            "",
            f"- 数据源: `{file_path}`",
            f"- Sheet: `{sheet_name}`",
            "- 口径: ProductCode=目标型号，Operation=CT，指标行=CT良率；若 CT产出数为 0，则标记为无有效产出日。",
            "",
            "| 日期 | CT产出数 | CT良率 | 数据状态 |",
            "| --- | ---: | ---: | --- |",
        ]
        for point in points:
            status = "有效" if point.has_valid_output else "无有效产出"
            output_text = "-" if point.output_count is None else f"{point.output_count:.0f}"
            yield_text = "-" if point.yield_value is None else _format_percent(point.yield_value)
            lines.append(f"| {point.label} | {output_text} | {yield_text} | {status} |")

        lines.extend(["", "### 核心结论"])
        if len(valid_points) >= 2:
            first = valid_points[0]
            last = valid_points[-1]
            delta = (last.yield_value or 0) - (first.yield_value or 0)
            average = sum(point.yield_value or 0 for point in valid_points) / len(valid_points)
            min_point = min(valid_points, key=lambda point: point.yield_value or 0)
            max_point = max(valid_points, key=lambda point: point.yield_value or 0)
            direction = "上升" if delta > 0 else "下降" if delta < 0 else "持平"
            lines.append(
                f"- 有效生产日 {len(valid_points)}/{len(points)} 天，总 CT 产出数 {total_output:.0f}。"
            )
            lines.append(
                f"- 有效日 CT良率从 {first.label} 的 {_format_percent(first.yield_value)} "
                f"到 {last.label} 的 {_format_percent(last.yield_value)}，"
                f"{direction} {_format_pct_points(delta)}。"
            )
            lines.append(
                f"- 有效日平均 CT良率 {_format_percent(average)}；最低为 {min_point.label} "
                f"{_format_percent(min_point.yield_value)}，最高为 {max_point.label} "
                f"{_format_percent(max_point.yield_value)}。"
            )
            if len(valid_points) < len(points):
                lines.append("- 部分日期 CT产出数为 0，趋势判断已排除这些无有效产出日。")
        elif len(valid_points) == 1:
            only = valid_points[0]
            lines.append(
                f"- 近一周只有 1 天存在有效 CT 产出：{only.label}，CT良率 "
                f"{_format_percent(only.yield_value)}，不足以判断连续趋势。"
            )
        else:
            lines.append("- 近一周没有有效 CT 产出数据，无法判断日度 CT 良率趋势。")

        return "\n".join(lines)


def _is_standard_xlsx(path: Path) -> bool:
    try:
        with Path(path).open("rb") as file:
            return file.read(4) == XLSX_MAGIC
    except OSError:
        return False


def _cell(row: list[Any] | None, index: int | None) -> Any:
    if row is None or index is None or index >= len(row):
        return None
    return row[index]


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_percent(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.2f}%"


def _format_pct_points(value: float) -> str:
    sign = "+" if value > 0 else ""
    return f"{sign}{value * 100:.2f} pct"
