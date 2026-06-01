"""Structured analysis for daily report generation."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from yield_report.agent.spec_model import ArtifactRef, RunContext, SkillError, SkillResult
from yield_report.core.query_parser import ReportType
from yield_report.infrastructure.file_decryption import decrypt_excel_file
from yield_report.skills.data_analysis.models import DataAnalysisRequest
from yield_report.skills.report_download import tool as report_download_tool
from yield_report.skills.report_download.models import ReportDownloadRequest

logger = logging.getLogger(__name__)

TOOL_NAME = "data_analysis"
XLSX_MAGIC = b"PK\x03\x04"
HEADER_SCAN_ROWS = 12

DEFAULT_SECTIONS = ["gap", "trend", "known_exception", "new_exception"]

SOURCE_PATTERNS: dict[str, list[str]] = {
    "daily_yield": ["V3良率及不良率By月周天*.xlsx"],
    "target_decomposition": ["*良率目标拆解*.xlsx"],
    "gap_template": ["*Gap分析模板*.xlsx", "日良率Gap分析模板*.xlsx"],
    "ct_exception": ["CT良率异常波动管理表*.xlsx"],
    "code_mapping": ["大数据值班当日新增不良HL模板*.xlsx"],
}

DOWNLOAD_TYPES: dict[str, ReportType] = {
    "daily_yield": ReportType.DAILY_YIELD,
    "target_decomposition": ReportType.TARGET_DECOMPOSITION,
    "gap_template": ReportType.GAP_TEMPLATE,
    "ct_exception": ReportType.CT_EXCEPTION,
}

DEFECT_GROUPS = [
    "Array_AD",
    "Array_Line",
    "Array_Mura",
    "Array_Pixel",
    "ARRAY_RS查杀",
    "ARRAY其他",
    "CELL其他",
    "OLED_Mura",
    "OLED_RS查杀",
    "OLED其他",
    "OLED_色偏",
    "TP_RS查杀",
    "TP_Short NG",
    "TP其他",
    "TP 容值NG",
    "外观不良",
]

CONCENTRATION_RULES = {
    "Array_AD": "受SCA R品集中过货影响",
    "Array_Line": "受ACA R品集中过货影响",
    "Array_Mura": "受MVI集中过货影响",
    "Array_Pixel": "受MVI集中过货影响",
    "ARRAY_RS查杀": "受CUT站点新批次刚刚投入影响",
    "ARRAY其他": "受MVI集中过货影响",
    "CELL其他": "受MVI集中过货影响",
    "OLED_Mura": "受MVI集中过货影响",
    "OLED_RS查杀": "受CUT站点新批次刚刚投入影响",
    "OLED其他": "受CUT站点新批次刚刚投入影响",
    "OLED_色偏": "受MVI集中过货影响",
    "TP_RS查杀": "受CUT站点新批次刚刚投入影响",
    "TP_Short NG": "受35001站点集中过货影响",
    "TP其他": "受MVI集中过货影响",
    "TP 容值NG": "受35006站点集中过货影响",
    "外观不良": "受APP1站点新批次刚刚投入影响",
}


@dataclass
class SheetData:
    path: Path
    sheet_name: str
    rows: list[list[Any]]


class DailyReportStructuredAnalyzer:
    """Build daily-report facts for all shipped products."""

    def __init__(self, request: DataAnalysisRequest, context: RunContext | None = None) -> None:
        self.request = request
        self.context = context or RunContext(run_id="manual-run", workspace=Path.cwd())
        self.workspace = self.context.workspace.resolve()
        self.output_dir = _resolve_output_dir(self.context)
        self.normalized_dir = self.output_dir / "_daily_report_sources" / self.context.run_id
        self.sections = request.sections or DEFAULT_SECTIONS
        self.products = list(request.daily_report_products)
        self.report_date = _normalize_date(request.report_date or _first_product_date(self.products))
        self.source_files: dict[str, Path] = {}
        self.downstream_results: list[dict[str, Any]] = []
        self.warnings: list[str] = []
        self._sheet_cache: dict[tuple[str, str], SheetData] = {}

    def run(self) -> SkillResult:
        product_models = [_string_cell(product.get("product_type")) for product in self.products]
        self._resolve_sources(product_models)

        target_rates = self._load_target_rates()
        code_mapping = self._load_code_mapping()

        product_results: list[dict[str, Any]] = []
        blocked_sections: list[dict[str, str]] = []
        for product in self.products:
            sections = self._analyze_product(product, target_rates, code_mapping)
            blocked_sections.extend(
                {
                    "product_type": _string_cell(product.get("product_type")),
                    "section": section_name,
                    "reason": section.get("reason", ""),
                }
                for section_name, section in sections.items()
                if section.get("status") == "blocked"
            )
            report_text = _compose_report_text(sections)
            product_results.append(
                {
                    "product": product,
                    "sections": sections,
                    "report_text": report_text,
                    "warnings": [
                        section.get("reason", "")
                        for section in sections.values()
                        if section.get("status") in {"blocked", "skipped"} and section.get("reason")
                    ],
                }
            )

        payload = {
            "analysis_kind": "daily_report",
            "report_date": self.report_date,
            "products": product_results,
            "source_files": {alias: str(path) for alias, path in self.source_files.items()},
            "downstream_results": self.downstream_results,
            "blocked_sections": blocked_sections,
            "warnings": self.warnings,
        }
        if blocked_sections:
            return SkillResult(
                skill_name=TOOL_NAME,
                success=False,
                summary=f"日报结构化分析阻断: {len(blocked_sections)} 个必需段落缺少数据",
                data={"daily_report_facts": payload},
                warnings=self.warnings,
                error=SkillError(
                    code="data_analysis.daily_report.blocked",
                    message="日报 2.1-2.4 必需数据不完整，已阻断成功日报生成。",
                    recoverable=True,
                    details={"blocked_sections": blocked_sections},
                ),
            )

        artifacts = [
            ArtifactRef(
                kind="excel",
                path=path,
                description=f"daily_report source: {alias}",
                metadata={"alias": alias, "skill": TOOL_NAME},
            )
            for alias, path in self.source_files.items()
            if path.exists()
        ]
        return SkillResult(
            skill_name=TOOL_NAME,
            success=True,
            summary=f"日报结构化分析完成: {len(product_results)} 个产品",
            artifacts=artifacts,
            data={"daily_report_facts": payload},
            warnings=self.warnings,
        )

    def _resolve_sources(self, product_models: list[str]) -> None:
        for alias in ["daily_yield", "target_decomposition", "gap_template", "ct_exception", "code_mapping"]:
            path = self._source_from_request(alias) or self._find_local_source(alias)
            if alias == "gap_template" and path is None and "target_decomposition" in self.source_files:
                continue
            if path is None and alias in DOWNLOAD_TYPES:
                path = self._download_source(alias, product_models)
            if path is None:
                continue
            normalized = self._normalize_source(alias, path)
            if normalized is not None:
                self.source_files[alias] = normalized

        if "daily_yield" in self.source_files and self.products:
            missing = self._products_missing_in_daily_yield(product_models)
            if missing:
                downloaded = self._download_source("daily_yield", missing)
                if downloaded is not None:
                    normalized = self._normalize_source("daily_yield", downloaded)
                    if normalized is not None:
                        self.source_files["daily_yield"] = normalized
                        self._sheet_cache = {
                            key: value
                            for key, value in self._sheet_cache.items()
                            if key[0] != "daily_yield"
                        }

    def _source_from_request(self, alias: str) -> Path | None:
        path = self.request.source_files.get(alias)
        if path is None:
            return None
        resolved = _resolve_path(Path(path), self.workspace)
        return resolved if resolved.exists() else None

    def _find_local_source(self, alias: str) -> Path | None:
        patterns = SOURCE_PATTERNS.get(alias, [])
        for directory in [self.workspace / "resources", self.workspace / "docs" / "project_files"]:
            if not directory.exists():
                continue
            for pattern in patterns:
                for path in sorted(directory.glob(pattern)):
                    if path.is_file() and not path.name.startswith("~$"):
                        return path
        return None

    def _download_source(self, alias: str, product_models: list[str]) -> Path | None:
        report_type = DOWNLOAD_TYPES.get(alias)
        if report_type is None:
            return None
        request = ReportDownloadRequest(report_type=report_type)
        if alias == "daily_yield":
            request.end_date = self.report_date
            request.product_models = product_models or None

        result = report_download_tool.run(request, self.context)
        self.downstream_results.append(
            {
                "skill": "report_download",
                "alias": alias,
                "success": result.success,
                "summary": result.summary,
                "warnings": result.warnings,
                "error": result.error.model_dump(mode="json") if result.error else None,
                "artifacts": [str(artifact.path) for artifact in result.artifacts],
            }
        )
        if not result.success:
            return None
        for artifact in result.artifacts:
            if artifact.path and artifact.path.exists() and artifact.path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
                return artifact.path
        for file_item in result.data.get("files", []):
            if isinstance(file_item, dict) and file_item.get("success") and file_item.get("file_path"):
                path = Path(file_item["file_path"])
                if path.exists():
                    return path
        return None

    def _normalize_source(self, alias: str, path: Path) -> Path | None:
        path = Path(path)
        if _can_open_with_openpyxl(path):
            return path

        output_dir = self.normalized_dir / alias
        output_dir.mkdir(parents=True, exist_ok=True)
        try:
            normalized = decrypt_excel_file(path, output_dir)
        except Exception as exc:
            self.warnings.append(f"源文件 {alias} 规范化失败: {exc}")
            return None
        if not _can_open_with_openpyxl(normalized):
            self.warnings.append(f"源文件 {alias} 规范化后仍无法由 openpyxl 读取: {normalized}")
            return None
        return normalized

    def _products_missing_in_daily_yield(self, product_models: list[str]) -> list[str]:
        if not product_models:
            return []
        sheet = self._read_sheet("daily_yield", "CT")
        if sheet is None:
            return product_models
        missing = [
            model
            for model in product_models
            if model and len(select_product_rows(sheet.rows, model)) <= 1
        ]
        if missing:
            self.warnings.append(f"daily_yield 缺少产品 CT 数据，将尝试下载: {', '.join(missing)}")
        return missing

    def _read_sheet(self, alias: str, sheet_name: str) -> SheetData | None:
        key = (alias, sheet_name)
        if key in self._sheet_cache:
            return self._sheet_cache[key]
        path = self.source_files.get(alias)
        if path is None:
            return None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            self.warnings.append(f"读取源文件 {alias} 失败: {exc}")
            return None
        try:
            actual_sheet = _resolve_sheet_name(workbook.sheetnames, sheet_name)
            if actual_sheet is None:
                self.warnings.append(f"源文件 {alias} 缺少 Sheet: {sheet_name}")
                return None
            worksheet = workbook[actual_sheet]
            rows = _read_worksheet_rows(worksheet)
            data = SheetData(path=path, sheet_name=actual_sheet, rows=rows)
            self._sheet_cache[key] = data
            return data
        finally:
            workbook.close()

    def _load_target_rates(self) -> dict[tuple[str, str], float]:
        for alias, sheet_name in [
            ("target_decomposition", "各产品Group目标"),
            ("gap_template", "良率目标"),
        ]:
            sheet = self._read_sheet(alias, sheet_name)
            if sheet is None:
                continue
            rates = extract_target_rates(sheet.rows, _month_label(self.report_date))
            if rates:
                return rates
        return {}

    def _load_code_mapping(self) -> dict[str, dict[str, str]]:
        sheet = self._read_sheet("code_mapping", "Code归属表")
        return extract_code_mapping(sheet.rows) if sheet else {}

    def _analyze_product(
        self,
        product: dict[str, Any],
        target_rates: dict[tuple[str, str], float],
        code_mapping: dict[str, dict[str, str]],
    ) -> dict[str, dict[str, Any]]:
        gap = self._analyze_gap(product, target_rates, code_mapping)
        trend = self._analyze_trend(product)
        known = self._analyze_known_exception(product, gap)
        new = self._analyze_new_exception(product)
        return {
            "gap": gap,
            "trend": trend,
            "known_exception": known,
            "new_exception": new,
        }

    def _analyze_gap(
        self,
        product: dict[str, Any],
        target_rates: dict[tuple[str, str], float],
        code_mapping: dict[str, dict[str, str]],
    ) -> dict[str, Any]:
        product_model = _string_cell(product.get("product_type"))
        if _is_product_qualified(product):
            return _section("skipped", "产品达标，跳过 2.1。")
        if "daily_yield" not in self.source_files:
            return _section("blocked", "缺少 daily_yield，无法计算 2.1 当日 Gap 影响。")
        if not target_rates:
            return _section("blocked", "缺少良率目标数据，无法计算 Defect Group 正 Gap。")

        sheet = self._read_sheet("daily_yield", "CT")
        rows = select_product_rows(sheet.rows, product_model) if sheet else []
        if len(rows) <= 1:
            return _section("blocked", f"daily_yield 中未找到产品 {product_model} 的 CT 数据。")
        last_col = latest_daily_column(rows)
        if last_col is None:
            return _section("blocked", "daily_yield 中未找到日度数据列。")

        headers = [_norm_header(value) for value in rows[0]]
        group_col = _find_col(headers, ["defectgroup", "不良分组"])
        code_col = (group_col + 1) if group_col is not None else None
        gap_items: list[dict[str, Any]] = []
        for row in rows[1:]:
            group = _string_cell(_get_cell(row, group_col)) or _row_label(row)
            code = _string_cell(_get_cell(row, code_col))
            if group not in DEFECT_GROUPS:
                continue
            if code and code not in DEFECT_GROUPS:
                continue
            actual = _to_float(_get_cell(row, last_col))
            target = target_rates.get((product_model, group))
            if actual is None or target is None:
                continue
            gap = actual - target
            if gap > 0:
                gap_items.append(
                    {
                        "defect_group": group,
                        "actual_loss_rate": actual,
                        "target_loss_rate": target,
                        "gap": gap,
                        "concentration_reason": CONCENTRATION_RULES.get(group, ""),
                    }
                )
        gap_items.sort(key=lambda item: item["gap"], reverse=True)
        gap_items = gap_items[:3]
        if not gap_items:
            return _section("no_match", "未发现正 Gap 的 Defect Group。", facts={"gap_top_items": []})

        group_names = [item["defect_group"] for item in gap_items]
        defect_codes = top_defect_codes(rows, group_names, last_col)
        if not defect_codes and code_mapping:
            group_set = set(group_names)
            defect_codes = [
                code
                for code, metadata in code_mapping.items()
                if metadata.get("group") in group_set
            ][:3]
        facts = {"gap_top_items": gap_items, "top_defect_codes": defect_codes}
        text = _format_gap_text(gap_items)
        return _section("emitted", "", facts=facts, text=text)

    def _analyze_trend(self, product: dict[str, Any]) -> dict[str, Any]:
        product_model = _string_cell(product.get("product_type"))
        if "daily_yield" not in self.source_files:
            return _section("blocked", "缺少 daily_yield，无法分析 2.2 连续三天下降原因。")
        sheet = self._read_sheet("daily_yield", "CT")
        rows = select_product_rows(sheet.rows, product_model) if sheet else []
        if len(rows) <= 1:
            return _section("blocked", f"daily_yield 中未找到产品 {product_model} 的 CT 数据。")

        date_cols = latest_daily_columns(rows, count=3)
        if len(date_cols) < 3:
            return _section("blocked", "daily_yield 日度数据列不足三天。")
        ct_row = find_metric_row(rows, "CT良率")
        if ct_row is None:
            return _section("blocked", "daily_yield 未找到 CT良率 行。")
        mvi_row = find_metric_row(rows, "CT产出数_MVI产出占比")
        ct_values = [_to_float(_get_cell(ct_row, col)) for col in date_cols]
        mvi_values = [
            _to_float(_get_cell(mvi_row, col)) if mvi_row is not None else None
            for col in date_cols
        ]
        if any(value is None for value in ct_values):
            return _section("blocked", "CT良率 最新三天数据不完整。")
        ct_numbers = [float(value) for value in ct_values if value is not None]
        mvi_numbers = [float(value) for value in mvi_values if value is not None]
        facts = {
            "trend": {
                "checked": True,
                "is_declining": _strictly_descending(ct_numbers),
                "mvi_share_increasing": len(mvi_numbers) == 3 and _strictly_ascending(mvi_numbers),
                "date_labels": [str(_get_cell(rows[0], col)) for col in date_cols],
                "ct_yield_values": ct_numbers,
                "mvi_share_values": mvi_numbers,
                "skipped_reason": "",
            }
        }
        if not facts["trend"]["is_declining"]:
            return _section("skipped", "最新三天 CT良率未严格下降。", facts=facts)
        return _section("emitted", "", facts=facts, text=_format_trend_text(facts["trend"]))

    def _analyze_known_exception(self, product: dict[str, Any], gap: dict[str, Any]) -> dict[str, Any]:
        if gap.get("status") == "blocked":
            return _section("blocked", "2.1 被阻断，无法确定 2.3 的 Gap Top3 Defect Code。")
        defect_codes = gap.get("facts", {}).get("top_defect_codes") or []
        if not defect_codes:
            return _section("no_match", "未识别到 Gap Top3 内的 Defect Code。", facts={"records": []})
        if "ct_exception" not in self.source_files:
            return _section("blocked", "缺少 ct_exception，无法匹配近 30 天已知异常。")
        sheet = self._read_sheet("ct_exception", "CT异常波动调查")
        records = extract_exception_records(
            rows=sheet.rows if sheet else [],
            product_model=_string_cell(product.get("product_type")),
            report_date=_parse_date(self.report_date),
            defect_codes=defect_codes,
            mode="known",
        )
        facts = {"records": records}
        if not records:
            return _section("no_match", "近 30 天未匹配到已知异常。", facts=facts)
        return _section("emitted", "", facts=facts, text="\n\n".join(_format_exception_record("2.3", item) for item in records))

    def _analyze_new_exception(self, product: dict[str, Any]) -> dict[str, Any]:
        if "ct_exception" not in self.source_files:
            return _section("blocked", "缺少 ct_exception，无法匹配当日新增异常。")
        sheet = self._read_sheet("ct_exception", "CT异常波动调查")
        records = extract_exception_records(
            rows=sheet.rows if sheet else [],
            product_model=_string_cell(product.get("product_type")),
            report_date=_parse_date(self.report_date),
            defect_codes=None,
            mode="new",
        )
        facts = {"records": records}
        if not records:
            return _section("no_match", "当天未匹配到新增异常。", facts=facts)
        return _section("emitted", "", facts=facts, text="\n\n".join(_format_exception_record("【异常】", item) for item in records))


def run_daily_report_analysis(
    request: DataAnalysisRequest,
    context: RunContext | None = None,
) -> SkillResult:
    return DailyReportStructuredAnalyzer(request, context=context).run()


def extract_target_rates(rows: list[list[Any]], month_label: str) -> dict[tuple[str, str], float]:
    header_index = _find_header_row(rows, ["项目代号", "Group"])
    if header_index is None:
        return {}
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["项目代号"])
    group_col = _find_col(headers, ["group"])
    month_col = _find_col(headers, [_norm_header(month_label)])
    if product_col is None or group_col is None or month_col is None:
        return {}
    rates: dict[tuple[str, str], float] = {}
    for row in rows[header_index + 1 :]:
        product = _string_cell(_get_cell(row, product_col))
        group = _string_cell(_get_cell(row, group_col))
        value = _to_float(_get_cell(row, month_col))
        if product and group and value is not None:
            rates[(product, group)] = value
    return rates


def extract_code_mapping(rows: list[list[Any]]) -> dict[str, dict[str, str]]:
    header_index = _find_header_row(rows, ["DefectDesc", "GroupType"])
    if header_index is None:
        return {}
    headers = [_norm_header(value) for value in rows[header_index]]
    desc_col = _find_col(headers, ["defectdesc", "代码描述"])
    group_col = _find_col(headers, ["grouptype", "组别"])
    factory_col = _find_col(headers, ["factory", "归属工厂"])
    mapping: dict[str, dict[str, str]] = {}
    for row in rows[header_index + 1 :]:
        desc = _string_cell(_get_cell(row, desc_col))
        if desc:
            mapping[desc] = {
                "group": _string_cell(_get_cell(row, group_col)),
                "factory": _string_cell(_get_cell(row, factory_col)),
            }
    return mapping


def select_product_rows(rows: list[list[Any]], product_model: str) -> list[list[Any]]:
    header_index = _find_header_row(rows, ["ProductCode", "Operation"])
    if header_index is None:
        return []
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["productcode", "产品型号"])
    if product_col is None:
        return []
    selected: list[list[Any]] = [rows[header_index]]
    current_product = ""
    for row in rows[header_index + 1 :]:
        product_value = _string_cell(_get_cell(row, product_col))
        if product_value:
            current_product = product_value
        if current_product == product_model:
            selected.append(row)
    return selected


def latest_daily_column(rows: list[list[Any]]) -> int | None:
    columns = latest_daily_columns(rows, count=1)
    return columns[-1] if columns else None


def latest_daily_columns(rows: list[list[Any]], count: int) -> list[int]:
    if not rows:
        return []
    return [
        index
        for index, value in enumerate(rows[0])
        if _is_daily_label(_string_cell(value))
    ][-count:]


def find_metric_row(rows: list[list[Any]], metric_name: str) -> list[Any] | None:
    wanted = _norm_header(metric_name)
    for row in rows[1:]:
        if _norm_header(_row_label(row)) == wanted:
            return row
    return None


def top_defect_codes(
    rows: list[list[Any]],
    defect_groups: list[str],
    last_col: int,
    limit: int = 3,
) -> list[str]:
    if not rows:
        return []
    headers = [_norm_header(value) for value in rows[0]]
    group_col = _find_col(headers, ["defectgroup", "不良分组"])
    code_col = (group_col + 1) if group_col is not None else None
    group_set = set(defect_groups)
    scored: list[tuple[float, str]] = []
    current_group = ""
    for row in rows[1:]:
        explicit_group = _string_cell(_get_cell(row, group_col))
        row_label = _row_label(row)
        if explicit_group:
            current_group = explicit_group
        elif row_label in group_set:
            current_group = row_label
        if current_group not in group_set:
            continue
        code = _string_cell(_get_cell(row, code_col))
        if not code or code in DEFECT_GROUPS or code == current_group:
            continue
        value = _to_float(_get_cell(row, last_col))
        if value is not None:
            scored.append((value, code))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [code for _, code in scored[:limit]]


def extract_exception_records(
    *,
    rows: list[list[Any]],
    product_model: str,
    report_date: date,
    defect_codes: list[str] | None,
    mode: str,
) -> list[dict[str, str]]:
    if mode == "known" and defect_codes is not None and not defect_codes:
        return []
    header_index = _find_header_row(rows, ["产品", "不良", "通报日期"])
    if header_index is None:
        return []
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["产品"])
    defect_col = _find_col(headers, ["不良"])
    date_col = _find_col(headers, ["通报日期"])
    reply_col = _find_col(headers, ["工艺整合&工艺回复的改善及挽救进展", "工艺整合"])
    daily_col = _find_col(headers, ["日良损"])
    month_col = _find_col(headers, ["当月"])
    week_col = _find_col(headers, ["当周"])
    batch_col = _find_col(headers, ["本批次"])
    if product_col is None or defect_col is None or date_col is None:
        return []

    defect_set = set(defect_codes or [])
    start_date = report_date - timedelta(days=30)
    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        product = _string_cell(_get_cell(row, product_col))
        if product != product_model:
            continue
        defect = _string_cell(_get_cell(row, defect_col))
        if defect_set and defect not in defect_set:
            continue
        record_date = _parse_date_value(_get_cell(row, date_col))
        if record_date is None:
            continue
        if mode == "known" and not (start_date <= record_date <= report_date):
            continue
        if mode == "new" and record_date != report_date:
            continue
        reply = _string_cell(_get_cell(row, reply_col))
        records.append(
            {
                "product_model": product,
                "defect_code": defect,
                "report_datetime": _string_cell(_get_cell(row, date_col)),
                "daily_loss": _format_loss(_get_cell(row, daily_col)),
                "monthly_loss": _format_loss(_get_cell(row, month_col)),
                "weekly_loss": _format_loss(_get_cell(row, week_col)),
                "batch_loss": _format_loss(_get_cell(row, batch_col)),
                "raw_reply": reply,
                **_extract_reply_fields(reply),
            }
        )
    return records


def _section(
    status: str,
    reason: str,
    *,
    facts: dict[str, Any] | None = None,
    text: str = "",
) -> dict[str, Any]:
    return {
        "status": status,
        "reason": reason,
        "facts": facts or {},
        "text": text,
    }


def _compose_report_text(sections: dict[str, dict[str, Any]]) -> str:
    lines = [
        section.get("text", "").strip()
        for name in ["gap", "trend", "known_exception", "new_exception"]
        for section in [sections.get(name, {})]
        if section.get("status") == "emitted" and section.get("text", "").strip()
    ]
    return "\n\n".join(lines) if lines else "无"


def _format_gap_text(gap_items: list[dict[str, Any]]) -> str:
    total_actual = sum(float(item["actual_loss_rate"]) for item in gap_items)
    total_gap = sum(float(item["gap"]) for item in gap_items)
    groups = "、".join(str(item["defect_group"]) for item in gap_items)
    reasons = "、".join(
        sorted({str(item.get("concentration_reason", "")) for item in gap_items if item.get("concentration_reason")})
    )
    line = f"2.1 当日Gap影响：{groups} total不良率{_percent(total_actual)}，GAP {_percent(total_gap)}"
    if reasons:
        line += f"；当日Gap主要{reasons}"
    return line


def _format_trend_text(trend: dict[str, Any]) -> str:
    labels = trend.get("date_labels") or []
    label_range = f"{labels[0]}~{labels[-1]}" if len(labels) >= 2 else "近三天"
    if trend.get("mvi_share_increasing"):
        return (
            f"2.2 连续三天下降原因：{label_range} CT良率持续下降，"
            f"受MVI产出占比提升影响({_percent_chain(trend.get('mvi_share_values') or [])})"
        )
    return f"2.2 连续三天下降原因：{label_range} CT良率持续下降，未识别到MVI产出占比连续升高证据"


def _format_exception_record(prefix: str, record: dict[str, str]) -> str:
    title = f"{prefix} {record.get('defect_code', '')}"
    if record.get("daily_loss"):
        title += f" {record['daily_loss']}"
    if record.get("monthly_loss"):
        title += f"(当月 {record['monthly_loss']})"
    return "\n".join(
        [
            title,
            f"【异常原因】{record.get('exception_reason') or '/'}",
            f"【Inline监控】{record.get('inline_monitoring') or '/'}",
            f"【是否止血】{record.get('is_stopped') or '/'}",
            f"【影响范围】{record.get('impact_scope') or '/'}",
            f"【改善措施】{record.get('improvement_measures') or '/'}",
        ]
    )


def _is_product_qualified(product: dict[str, Any]) -> bool:
    is_qualified = product.get("is_qualified")
    daily_gap = _to_float(product.get("daily_gap"))
    return bool(is_qualified is True or (daily_gap is not None and daily_gap >= 0))


def _first_product_date(products: list[dict[str, Any]]) -> str:
    for product in products:
        value = _string_cell(product.get("report_date"))
        if value:
            return value
    return date.today().isoformat()


def _resolve_output_dir(context: RunContext) -> Path:
    output_dir = Path(context.output_dir)
    return output_dir if output_dir.is_absolute() else context.workspace / output_dir


def _resolve_path(path: Path, workspace: Path) -> Path:
    return path if path.is_absolute() else workspace / path


def _find_header_row(rows: list[list[Any]], required_headers: list[str]) -> int | None:
    required = [_norm_header(item) for item in required_headers]
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        normalized = [_norm_header(value) for value in row]
        if all(any(req in cell for cell in normalized) for req in required):
            return index
    return None


def _find_col(headers: list[str], candidates: list[str]) -> int | None:
    normalized_candidates = [_norm_header(item) for item in candidates]
    for candidate in normalized_candidates:
        for index, header in enumerate(headers):
            if candidate and candidate == header:
                return index
    for candidate in normalized_candidates:
        for index, header in enumerate(headers):
            if candidate and candidate in header:
                return index
    return None


def _resolve_sheet_name(sheet_names: list[str], requested: str) -> str | None:
    if requested in sheet_names:
        return requested
    requested_norm = requested.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() == requested_norm:
            return sheet_name
    return None


def _read_worksheet_rows(worksheet: Any) -> list[list[Any]]:
    # FineReport/Excel-COM exports can keep a stale A1:A1 sheet dimension.
    # In openpyxl read-only mode that hides every real data column unless reset first.
    reset_dimensions = getattr(worksheet, "reset_dimensions", None)
    if callable(reset_dimensions):
        try:
            reset_dimensions()
        except Exception:
            logger.debug("Unable to reset worksheet dimensions", exc_info=True)
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _get_cell(row: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _row_label(row: list[Any]) -> str:
    for index in [3, 4]:
        value = _string_cell(_get_cell(row, index))
        if value:
            return value
    return ""


def _string_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if "%" in str(value) else number


def _normalize_date(value: str) -> str:
    parsed = _parse_date_value(value)
    return parsed.isoformat() if parsed else str(value or date.today().isoformat())


def _parse_date(value: str) -> date:
    parsed = _parse_date_value(value)
    if parsed is None:
        raise ValueError(f"Unable to parse date: {value}")
    return parsed


def _parse_date_value(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    compact = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(?:\.0)?", text)
    if compact:
        return date(int(compact.group(1)), int(compact.group(2)), int(compact.group(3)))
    text = text.replace("/", "-")
    if text.endswith("+00:00"):
        text = text[:-6]
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _month_label(report_date: str) -> str:
    return f"{_parse_date(report_date).month}月"


def _is_daily_label(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}", value.strip()))


def _strictly_descending(values: list[float]) -> bool:
    return len(values) >= 2 and all(left > right for left, right in zip(values, values[1:]))


def _strictly_ascending(values: list[float]) -> bool:
    return len(values) >= 2 and all(left < right for left, right in zip(values, values[1:]))


def _extract_reply_fields(text: str) -> dict[str, str]:
    return {
        "exception_reason": _extract_tag(text, "异常原因"),
        "inline_monitoring": _extract_tag(text, "Inline监控"),
        "is_stopped": _extract_tag(text, "是否止血"),
        "impact_scope": _extract_tag(text, "影响范围"),
        "improvement_measures": _extract_tag(text, "改善措施"),
    }


def _extract_tag(text: str, tag: str) -> str:
    pattern = rf"【{re.escape(tag)}】([\s\S]*?)(?=【[^】]+】|$)"
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _format_loss(value: Any) -> str:
    number = _to_float(value)
    if number is None:
        return _string_cell(value)
    return _percent(number)


def _percent(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value * 100:.2f}%"


def _percent_chain(values: list[float]) -> str:
    return "→".join(_percent(value) for value in values)


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", _string_cell(value).lower())


def _is_standard_xlsx(path: Path) -> bool:
    try:
        with path.open("rb") as file:
            return file.read(4) == XLSX_MAGIC
    except OSError:
        return False


def _can_open_with_openpyxl(path: Path) -> bool:
    if not _is_standard_xlsx(path):
        return False
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        return bool(workbook.sheetnames)
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.close()
