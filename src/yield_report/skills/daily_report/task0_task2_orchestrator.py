"""Task0-Task2 orchestrator adapter for the daily_report skill."""

from __future__ import annotations

import json
import os
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from yield_report.agent.spec_model import ArtifactRef, RunContext, SkillError, SkillResult
from yield_report.skills.daily_report.models import DailyReportRequest

TOOL_NAME = "daily_report"
DEFAULT_DUTY_WORKSPACE = Path(r"D:\wzy\工作-值班工作\相关文件")
DEFAULT_OUTPUT_PREFIX = "V3良率日报每日异常填报表"
DEFAULT_CHILD_PYTHON = "python"
TASK1_GAP_ANALYSIS_SCRIPT = (
    Path.home() / ".codex" / "skills" / "task1-gap-analysis" / "scripts" / "task1_gap_analysis.py"
)


@dataclass(frozen=True)
class CommandResult:
    """One child script invocation result."""

    child_skill: str
    command: list[str]
    returncode: int
    stdout: str
    stderr: str

    def to_payload(self) -> dict[str, Any]:
        return {
            "child_skill": self.child_skill,
            "command": self.command,
            "returncode": self.returncode,
            "stdout": _json_or_text(self.stdout),
            "stderr": self.stderr,
        }


class Task0Task2Orchestrator:
    """Execute the external OLED duty Task0-Task2 scripts in order."""

    workflow = [
        "task0-report-download",
        "task1-gap-analysis",
        "task2-extract-anomalies",
    ]

    def __init__(self, request: DailyReportRequest, context: RunContext) -> None:
        self.request = request
        self.context = context
        self.workspace = self._resolve_workspace()
        self.output_dir = Path(request.output_dir or context.output_dir).resolve()
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.output_path = self._resolve_output_path()

    def run(self) -> SkillResult:
        try:
            self._assert_scripts_exist()
            command_results = self._run_workflow()
            verification = verify_daily_report_workbook(self.output_path)
            comparison = (
                compare_workbook_values(self.output_path, self.request.reference_workbook)
                if self.request.reference_workbook
                else None
            )
            success = comparison is None or comparison["match"]
            return SkillResult(
                skill_name=TOOL_NAME,
                success=success,
                summary=(
                    f"Task0-Task2 日报流程完成: {self.output_path}"
                    if success
                    else f"Task0-Task2 日报流程完成，但与目标文件不一致: {self.output_path}"
                ),
                artifacts=[
                    ArtifactRef(
                        kind="excel",
                        path=self.output_path,
                        description="Task0-Task2 generated daily report workbook",
                        metadata={"skill": TOOL_NAME},
                    )
                ],
                data={
                    "workflow": self.workflow,
                    "workspace": str(self.workspace),
                    "output_file": str(self.output_path),
                    "steps": [result.to_payload() for result in command_results],
                    "verification": verification,
                    "comparison": comparison,
                },
                error=None
                if success
                else SkillError(
                    code="daily_report.reference_mismatch",
                    message="Generated workbook differs from the reference workbook.",
                    recoverable=True,
                    details={"comparison": comparison},
                ),
            )
        except Exception as exc:
            return SkillResult(
                skill_name=TOOL_NAME,
                success=False,
                summary=f"Task0-Task2 日报流程失败: {exc}",
                data={
                    "workflow": self.workflow,
                    "workspace": str(self.workspace),
                    "output_file": str(self.output_path),
                },
                error=SkillError(
                    code="daily_report.orchestrator.failed",
                    message=str(exc),
                    recoverable=True,
                ),
            )

    def _resolve_workspace(self) -> Path:
        configured = (
            self.request.orchestrator_workspace
            or self.request.source_files.get("orchestrator_workspace")
            or os.getenv("YIELD_REPORT_DUTY_WORKSPACE")
            or DEFAULT_DUTY_WORKSPACE
        )
        return Path(configured).resolve()

    def _resolve_output_path(self) -> Path:
        if self.request.output_name:
            return (self.output_dir / self.request.output_name).resolve()
        suffix = _suffix_from_request(self.request)
        return (self.output_dir / f"{DEFAULT_OUTPUT_PREFIX}-{suffix}.xlsx").resolve()

    def _assert_scripts_exist(self) -> None:
        missing = [
            str(path)
            for path in (
                self._script("task0_report_download.py"),
                self._task1_script(),
                self._script("task2_extract_anomalies.py"),
            )
            if not path.exists()
        ]
        if missing:
            raise FileNotFoundError(f"Task0-Task2 child script missing: {', '.join(missing)}")

    def _run_workflow(self) -> list[CommandResult]:
        results: list[CommandResult] = []
        results.append(
            self._run_command(
                "task0-report-download",
                self._task0_command(),
            )
        )
        if not self.output_path.exists():
            raise FileNotFoundError(f"Task0 did not create workbook: {self.output_path}")

        results.append(
            self._run_command(
                "task1-gap-analysis:self-test",
                self._task1_command("--self-test"),
            )
        )
        if self.request.run_inspection:
            results.append(
                self._run_command(
                    "task1-gap-analysis:inspect-before",
                    self._task1_command("--inspect"),
                )
            )
        results.append(
            self._run_command(
                "task1-gap-analysis",
                self._task1_command("--write"),
            )
        )
        results.append(
            self._run_command(
                "task2-extract-anomalies",
                self._task2_command(),
            )
        )
        if self.request.run_inspection:
            results.append(
                self._run_command(
                    "task2-extract-anomalies:inspect",
                    self._task2_command("--inspect"),
                )
            )
        return results

    def _task0_command(self) -> list[str]:
        command = [
            self._child_python(),
            str(self._script("task0_report_download.py")),
            "--write",
            "--output",
            str(self.output_path),
        ]
        if self.request.report_date:
            command.extend(["--end-date", self.request.report_date])
        if self.request.download_sources:
            command.append("--download-sources")
        return command

    def _task1_script(self) -> Path:
        configured = self.request.source_files.get("task1_gap_analysis_script") or os.getenv(
            "YIELD_REPORT_TASK1_GAP_ANALYSIS_SCRIPT"
        )
        if configured:
            return Path(configured).expanduser().resolve()
        return TASK1_GAP_ANALYSIS_SCRIPT.resolve()

    def _task1_command(self, action: str) -> list[str]:
        command = [self._child_python(), str(self._task1_script())]
        if action != "--self-test":
            command.append(str(self.output_path))
        command.append(action)
        if self.request.orchestrator_now and action != "--self-test":
            command.extend(["--now", self.request.orchestrator_now])
        return command

    def _task2_command(self, action: str = "--write") -> list[str]:
        command = [
            self._child_python(),
            str(self._script("task2_extract_anomalies.py")),
            "--source",
            str(self.output_path),
            action,
        ]
        if self.request.orchestrator_now:
            command.extend(["--now", self.request.orchestrator_now])
        if self.request.task2_max_anomaly_row is not None:
            command.extend(["--max-anomaly-row", str(self.request.task2_max_anomaly_row)])
        return command

    def _script(self, name: str) -> Path:
        return self.workspace / "scripts" / name

    def _child_python(self) -> str:
        configured = self.request.source_files.get("child_python") or os.getenv(
            "YIELD_REPORT_CHILD_PYTHON"
        )
        return str(configured or DEFAULT_CHILD_PYTHON)

    def _run_command(self, child_skill: str, command: list[str]) -> CommandResult:
        completed = subprocess.run(
            command,
            cwd=self.workspace,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
        result = CommandResult(
            child_skill=child_skill,
            command=command,
            returncode=completed.returncode,
            stdout=completed.stdout.strip(),
            stderr=completed.stderr.strip(),
        )
        if completed.returncode != 0:
            raise RuntimeError(
                f"{child_skill} failed with exit code {completed.returncode}: {result.stderr or result.stdout}"
            )
        return result


def verify_daily_report_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Generated workbook does not exist: {path}")

    rows = _read_sheet_values(path, "Data Packet")
    headers = [str(value or "").strip() for value in (rows[0] if rows else [])]
    col_map = {header: index for index, header in enumerate(headers) if header}
    row_count = max(len(rows) - 1, 0)
    return {
        "sheet": "Data Packet",
        "headers": headers,
        "row_count": row_count,
        "nonblank_counts": {
            "1.1 过货影响": _count_nonblank_values(rows, _find_column(col_map, "1.1")),
            "1.3 当日异常": _count_nonblank_values(rows, _find_column(col_map, "1.3")),
            "1.4 已知异常": _count_nonblank_values(rows, _find_column(col_map, "1.4")),
        },
    }


def _read_sheet_values(path: Path, sheet_name: str) -> list[list[Any]]:
    try:
        workbook = load_workbook(path, data_only=True)
    except BadZipFile:
        return _read_sheet_values_with_excel_com(path, sheet_name)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook missing {sheet_name} sheet: {path}")
        sheet = workbook[sheet_name]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_sheet_values_with_excel_com(path: Path, sheet_name: str) -> list[list[Any]]:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError(
            f"Workbook is not readable by openpyxl and Excel COM is unavailable: {path}"
        ) from exc

    pythoncom.CoInitialize()
    app = win32com.client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    workbook = None
    try:
        workbook = app.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        try:
            sheet = workbook.Worksheets(sheet_name)
        except Exception as exc:
            raise ValueError(f"Workbook missing {sheet_name} sheet: {path}") from exc
        used_range = sheet.UsedRange
        first_row = int(used_range.Row)
        first_col = int(used_range.Column)
        row_count = int(used_range.Rows.Count)
        col_count = int(used_range.Columns.Count)
        rows: list[list[Any]] = []
        for row in range(first_row, first_row + row_count):
            rows.append(
                [sheet.Cells(row, col).Value for col in range(first_col, first_col + col_count)]
            )
        return rows
    finally:
        if workbook is not None:
            workbook.Close(False)
        app.Quit()
        pythoncom.CoUninitialize()


def compare_workbook_values(
    generated_path: Path,
    reference_path: Path,
    *,
    sheet_names: tuple[str, ...] = ("Data Packet",),
    max_differences: int = 50,
) -> dict[str, Any]:
    """Compare generated workbook values against a reference workbook."""
    differences: list[dict[str, Any]] = []
    for sheet_name in sheet_names:
        try:
            generated_values = _read_sheet_values(generated_path, sheet_name)
        except ValueError:
            differences.append({"sheet": sheet_name, "issue": "missing_generated_sheet"})
            continue
        try:
            reference_values = _read_sheet_values(reference_path, sheet_name)
        except ValueError:
            differences.append({"sheet": sheet_name, "issue": "missing_reference_sheet"})
            continue

        max_row = max(len(generated_values), len(reference_values))
        max_col = max(_max_columns(generated_values), _max_columns(reference_values))
        for row in range(max_row):
            for col in range(max_col):
                generated = _normalize_cell(_value_at(generated_values, row, col))
                reference = _normalize_cell(_value_at(reference_values, row, col))
                if generated != reference:
                    differences.append(
                        {
                            "sheet": sheet_name,
                            "cell": f"{get_column_letter(col + 1)}{row + 1}",
                            "generated": generated,
                            "reference": reference,
                        }
                    )
                    if len(differences) >= max_differences:
                        return {
                            "match": False,
                            "generated": str(generated_path),
                            "reference": str(reference_path),
                            "differences": differences,
                            "truncated": True,
                        }
    return {
        "match": not differences,
        "generated": str(generated_path),
        "reference": str(reference_path),
        "differences": differences,
        "truncated": False,
    }


def _max_columns(rows: list[list[Any]]) -> int:
    return max((len(row) for row in rows), default=0)


def _value_at(rows: list[list[Any]], row: int, col: int) -> Any:
    if row >= len(rows) or col >= len(rows[row]):
        return None
    return rows[row][col]


def _find_column(col_map: dict[str, int], needle: str) -> int | None:
    for header, col in col_map.items():
        if needle in header:
            return col
    return None


def _count_nonblank_values(rows: list[list[Any]], col: int | None) -> int:
    if col is None:
        return 0
    count = 0
    for row in rows[1:]:
        if col < len(row) and str(row[col] or "").strip():
            count += 1
    return count


def _suffix_from_request(request: DailyReportRequest) -> str:
    if request.orchestrator_now and request.orchestrator_now[:10]:
        day = request.orchestrator_now[:10].replace("-", "")
        return f"{day}-16：00" if "16" in request.orchestrator_now[11:16] else day
    if request.report_date:
        return request.report_date.replace("-", "")
    return "manual"


def _json_or_text(value: str) -> Any:
    if not value:
        return ""
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r\n", "\n").strip()
    return value
