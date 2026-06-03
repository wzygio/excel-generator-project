"""Implementation for the daily_report skill."""

from __future__ import annotations

import html
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from yield_report.agent.spec_model import ArtifactRef, RunContext, SkillError, SkillResult
from yield_report.infrastructure.file_decryption import decrypt_excel_file
from yield_report.skills.daily_report.models import (
    DailyReportPayload,
    DailyReportRequest,
    ExceptionRecord,
    GapItem,
    ProductDailyReport,
    ShippedProduct,
    TrendResult,
)
from yield_report.skills.data_analysis import tool as data_analysis_tool
from yield_report.skills.data_analysis.models import DataAnalysisRequest

logger = logging.getLogger(__name__)

TOOL_NAME = "daily_report"
XLSX_MAGIC = b"PK\x03\x04"
HEADER_SCAN_ROWS = 10

DEFAULT_SECTIONS = ["gap", "trend", "known_exception", "new_exception"]
DEFAULT_TEMPLATE = Path("docs/project_files/V3良率日报每日异常填报表.xlsx")

SOURCE_PATTERNS: dict[str, list[str]] = {
    "spotfire": ["spotfire*.xlsx"],
    "daily_yield": ["V3良率及不良率By月周天*.xlsx"],
    "target_decomposition": ["*良率目标拆解*.xlsx"],
    "gap_template": ["*Gap分析模板*.xlsx", "日良率Gap分析模板*.xlsx"],
    "ct_exception": ["CT良率异常波动管理表*.xlsx"],
    "code_mapping": ["大数据值班当日新增不良HL模板*.xlsx"],
}

DEFECT_GROUPS = [
    "Array_AD",
    "Array_Line",
    "Array_Mura",
    "Array_Pixel",
    "ARRAY_RS查杀",
    "ARRAY其他",
    "CELL其他",
    "OLED_Mura",
    "OLED_RS查杀",
    "OLED其他",
    "OLED_色偏",
    "TP_RS查杀",
    "TP_Short NG",
    "TP其他",
    "TP 容值NG",
    "外观不良",
]

CONCENTRATION_RULES = {
    "Array_AD": "受SCA R品集中过货影响",
    "Array_Line": "受ACA R品集中过货影响",
    "Array_Mura": "受MVI集中过货影响",
    "Array_Pixel": "受MVI集中过货影响",
    "ARRAY_RS查杀": "受CUT站点新批次刚刚投入影响",
    "ARRAY其他": "受MVI集中过货影响",
    "CELL其他": "受MVI集中过货影响",
    "OLED_Mura": "受MVI集中过货影响",
    "OLED_RS查杀": "受CUT站点新批次刚刚投入影响",
    "OLED其他": "受CUT站点新批次刚刚投入影响",
    "OLED_色偏": "受MVI集中过货影响",
    "TP_RS查杀": "受CUT站点新批次刚刚投入影响",
    "TP_Short NG": "受35001站点集中过货影响",
    "TP其他": "受MVI集中过货影响",
    "TP 容值NG": "受35006站点集中过货影响",
    "外观不良": "受APP1站点新批次刚刚投入影响",
}


class DailyReportGenerationError(Exception):
    """Raised when the daily report cannot be generated."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        recoverable: bool = True,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.recoverable = recoverable
        self.details = details or {}


@dataclass
class SheetData:
    """In-memory worksheet values."""

    path: Path
    sheet_name: str
    rows: list[list[Any]]


class ExcelSheetReader:
    """Read standard or enterprise-encrypted Excel sheets into Python values."""

    def read_sheet(self, path: Path, sheet_name: str) -> SheetData:
        path = Path(path)
        if not path.exists():
            raise DailyReportGenerationError(
                "daily_report.file.missing",
                f"源文件不存在: {path}",
                details={"path": str(path), "sheet_name": sheet_name},
            )

        if _is_standard_xlsx(path):
            return self._read_with_openpyxl(path, sheet_name)
        return self._read_with_com(path, sheet_name)

    def _read_with_openpyxl(self, path: Path, sheet_name: str) -> SheetData:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise DailyReportGenerationError(
                "daily_report.file.read_failed",
                f"无法读取 Excel 文件: {path}",
                details={"path": str(path), "error": str(exc)},
            ) from exc

        try:
            actual_sheet = _resolve_sheet_name(workbook.sheetnames, sheet_name)
            if actual_sheet is None:
                raise DailyReportGenerationError(
                    "daily_report.file.missing_sheet",
                    f"文件中缺少 Sheet: {sheet_name}",
                    details={"path": str(path), "available_sheets": workbook.sheetnames},
                )
            worksheet = workbook[actual_sheet]
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
            ]
            return SheetData(path=path, sheet_name=actual_sheet, rows=rows)
        finally:
            workbook.close()

    def _read_with_com(self, path: Path, sheet_name: str) -> SheetData:
        try:
            import win32com.client  # type: ignore[import-untyped]
        except ImportError as exc:
            raise DailyReportGenerationError(
                "daily_report.file.encrypted_read_unavailable",
                "读取企业加密 Excel 需要 pywin32/Excel COM。",
                details={"path": str(path)},
            ) from exc

        excel = None
        workbook = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            for attr, value in (
                ("DisplayAlerts", False),
                ("AskToUpdateLinks", False),
                ("EnableEvents", False),
            ):
                try:
                    setattr(excel, attr, value)
                except Exception:
                    pass
            workbook = excel.Workbooks.Open(str(path.resolve()), 0, True)
            sheet_names = [
                workbook.Worksheets(index).Name
                for index in range(1, workbook.Worksheets.Count + 1)
            ]
            actual_sheet = _resolve_sheet_name(sheet_names, sheet_name)
            if actual_sheet is None:
                raise DailyReportGenerationError(
                    "daily_report.file.missing_sheet",
                    f"文件中缺少 Sheet: {sheet_name}",
                    details={"path": str(path), "available_sheets": sheet_names},
                )
            worksheet = workbook.Worksheets(actual_sheet)
            used_range = worksheet.UsedRange
            raw_values = used_range.Value
            rows = _com_values_to_rows(raw_values)
            return SheetData(path=path, sheet_name=actual_sheet, rows=rows)
        except DailyReportGenerationError:
            raise
        except Exception as exc:
            raise DailyReportGenerationError(
                "daily_report.file.read_failed",
                f"无法通过 Excel COM 读取文件: {path}",
                details={"path": str(path), "sheet_name": sheet_name, "error": str(exc)},
            ) from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass


class DailyReportGenerator:
    """Generate a daily report from deterministic workbook analyzers."""

    def __init__(self, request: DailyReportRequest, context: RunContext) -> None:
        self.request = request
        self.context = context
        self.workspace = context.workspace.resolve()
        self.output_dir = _resolve_output_dir(request, context)
        self.reader = ExcelSheetReader()
        self.warnings: list[str] = []
        self.spec_data = self._load_spec_data()
        self.source_files = self._resolve_source_files()

    def run(self) -> SkillResult:
        report_date = self._resolve_report_date()
        products = self._load_products(report_date)
        if not products:
            raise DailyReportGenerationError(
                "daily_report.data.no_products",
                "未找到当日过货产品。",
                details={"report_date": report_date},
            )

        logger.info(
            "Daily report generation started: report_date=%s product_count=%d source_files=%s",
            report_date,
            len(products),
            {key: str(path) for key, path in self.source_files.items()},
            extra={
                "event": "start",
                "purpose": "business",
                "run_id": self.context.run_id,
                "task_id": TOOL_NAME,
            },
        )
        analysis_result = self._run_data_analysis(report_date, products)
        if not analysis_result.success:
            facts = analysis_result.data.get("daily_report_facts") or {}
            warnings = list(self.warnings) + list(analysis_result.warnings)
            blocked_sections = facts.get("blocked_sections", [])
            logger.error(
                "Daily report generation blocked: report_date=%s product_count=%d "
                "blocked_sections=%d warnings=%s source_files=%s",
                report_date,
                len(products),
                len(blocked_sections),
                warnings,
                {key: str(path) for key, path in self.source_files.items()},
                extra={
                    "event": "failure",
                    "purpose": "business",
                    "run_id": self.context.run_id,
                    "task_id": TOOL_NAME,
                    "error_code": "daily_report.analysis.blocked",
                },
            )
            return SkillResult(
                skill_name=TOOL_NAME,
                success=False,
                summary=f"日报生成阻断: {analysis_result.summary}",
                data={
                    "report_date": report_date,
                    "products": [product.model_dump(mode="json") for product in products],
                    "source_files": {key: str(path) for key, path in self.source_files.items()},
                    "daily_report_facts": facts,
                    "downstream_results": facts.get("downstream_results", []),
                    "blocked_sections": facts.get("blocked_sections", []),
                },
                warnings=warnings,
                error=SkillError(
                    code="daily_report.analysis.blocked",
                    message=analysis_result.error.message if analysis_result.error else analysis_result.summary,
                    recoverable=True,
                    details=analysis_result.error.details if analysis_result.error else {},
                ),
            )

        facts = analysis_result.data.get("daily_report_facts") or {}
        product_reports = [
            _product_report_from_analysis(item)
            for item in facts.get("products", [])
        ]

        payload = DailyReportPayload(
            report_date=report_date,
            products=product_reports,
            source_files=facts.get("source_files") or {key: str(path) for key, path in self.source_files.items()},
            downstream_results=facts.get("downstream_results", []),
            blocked_sections=facts.get("blocked_sections", []),
            warnings=list(self.warnings) + list(analysis_result.warnings),
        )
        artifacts = self._write_outputs(payload)
        payload.output_file = str(artifacts[0].path) if artifacts else ""

        logger.info(
            "Daily report generation completed: report_date=%s product_count=%d output=%s warnings=%s",
            report_date,
            len(product_reports),
            payload.output_file,
            payload.warnings,
            extra={
                "event": "success",
                "purpose": "business",
                "run_id": self.context.run_id,
                "task_id": TOOL_NAME,
                "output_path": payload.output_file,
            },
        )
        return SkillResult(
            skill_name=TOOL_NAME,
            success=True,
            summary=f"日报生成完成: {len(product_reports)} 个产品",
            artifacts=artifacts,
            data=payload.model_dump(mode="json"),
            warnings=list(self.warnings),
        )

    def _load_spec_data(self) -> dict[str, Any]:
        if self.request.spec_path is None:
            return {}
        spec_path = _resolve_path(self.request.spec_path, self.workspace)
        if not spec_path.exists():
            raise DailyReportGenerationError(
                "daily_report.spec.missing",
                f"Spec 文件不存在: {spec_path}",
                details={"spec_path": str(spec_path)},
            )
        raw = yaml.safe_load(spec_path.read_text(encoding="utf-8")) or {}
        if not isinstance(raw, dict):
            raise DailyReportGenerationError(
                "daily_report.spec.invalid",
                "Spec 文件根节点必须是 YAML mapping。",
                details={"spec_path": str(spec_path)},
            )
        return raw

    def _resolve_source_files(self) -> dict[str, Path]:
        source_files: dict[str, Path] = {}
        source_files.update(self._source_files_from_spec())
        source_files.update(
            {key: _resolve_path(path, self.workspace) for key, path in self.request.source_files.items()}
        )

        search_dirs = [self.workspace / "resources", self.workspace / "docs" / "project_files"]
        for alias, patterns in SOURCE_PATTERNS.items():
            if alias in source_files and source_files[alias].exists():
                continue
            found = None
            for directory in search_dirs:
                found = _find_first_file(directory, patterns)
                if found is not None:
                    break
            if found is not None:
                source_files[alias] = found

        missing_required = [
            alias
            for alias in ["spotfire"]
            if alias not in source_files or not source_files[alias].exists()
        ]
        if missing_required:
            raise DailyReportGenerationError(
                "daily_report.file.missing_required",
                f"缺少日报生成必需源文件: {', '.join(missing_required)}",
                details={"missing": missing_required},
            )

        return source_files

    def _source_files_from_spec(self) -> dict[str, Path]:
        inputs = self.spec_data.get("inputs", {})
        if not isinstance(inputs, dict):
            return {}

        files: dict[str, Path] = {}
        for item in inputs.get("local_files", []) or []:
            if not isinstance(item, dict):
                continue
            alias = item.get("alias") or item.get("name")
            raw_path = item.get("path") or item.get("file_path")
            if alias and raw_path:
                files[str(alias)] = _resolve_path(Path(raw_path), self.workspace)
        return files

    def _resolve_report_date(self) -> str:
        if self.request.report_date:
            return _normalize_date(self.request.report_date)

        inputs = self.spec_data.get("inputs", {})
        if isinstance(inputs, dict) and inputs.get("report_date"):
            return _normalize_date(str(inputs["report_date"]))

        spotfire = self.reader.read_sheet(self.source_files["spotfire"], "Sheet1")
        products = extract_shipped_products(spotfire.rows)
        if products:
            return products[0].report_date
        return date.today().isoformat()

    def _load_products(self, report_date: str) -> list[ShippedProduct]:
        spotfire = self.reader.read_sheet(self.source_files["spotfire"], "Sheet1")
        products = extract_shipped_products(spotfire.rows)
        product_filter = self.request.product_models
        if product_filter is None:
            inputs = self.spec_data.get("inputs", {})
            if isinstance(inputs, dict):
                raw_models = inputs.get("product_models")
                if isinstance(raw_models, list) and raw_models:
                    product_filter = [str(item) for item in raw_models]
        if product_filter:
            products = [
                item
                for item in products
                if _matches_product_filter(item, product_filter)
            ]

        return [
            item.model_copy(update={"report_date": report_date})
            for item in products
        ]

    def _run_data_analysis(
        self,
        report_date: str,
        products: list[ShippedProduct],
    ) -> SkillResult:
        return data_analysis_tool.run(
            DataAnalysisRequest(
                analysis_kind="daily_report",
                report_date=report_date,
                sections=self.request.sections or DEFAULT_SECTIONS,
                source_files=self.source_files,
                daily_report_products=[product.model_dump(mode="json") for product in products],
                product_models=[product.product_type for product in products],
                analysis_intent="生成日报 2.1-2.4 结构化分析事实",
            ),
            self.context,
        )

    def _load_target_rates(self, report_date: str) -> dict[tuple[str, str], float]:
        candidates: list[tuple[str, str]] = []
        if "target_decomposition" in self.source_files:
            candidates.append(("target_decomposition", "各产品Group目标"))
        if "gap_template" in self.source_files:
            candidates.append(("gap_template", "良率目标"))

        for alias, sheet_name in candidates:
            try:
                sheet = self.reader.read_sheet(self.source_files[alias], sheet_name)
                rates = extract_target_rates(sheet.rows, _month_label(report_date))
                if rates:
                    return rates
            except DailyReportGenerationError as exc:
                self.warnings.append(f"读取目标表失败({alias}): {exc.message}")
        self.warnings.append("未读取到良率目标，Gap 将按目标 0 计算。")
        return {}

    def _load_code_mapping(self) -> dict[str, dict[str, str]]:
        path = self.source_files.get("code_mapping")
        if path is None:
            return {}
        try:
            sheet = self.reader.read_sheet(path, "Code归属表")
        except DailyReportGenerationError as exc:
            self.warnings.append(f"读取 Code 归属表失败: {exc.message}")
            return {}
        return extract_code_mapping(sheet.rows)

    def _fill_gap_section(
        self,
        report: ProductDailyReport,
        target_rates: dict[tuple[str, str], float],
        defect_code_mapping: dict[str, dict[str, str]],
    ) -> None:
        product = report.product
        if "daily_yield" not in self.source_files:
            report.warnings.append("2.1 skipped: 缺少 daily_yield，使用 spotfire TOP3 信息。")
            return
        if product.is_qualified is True or (product.daily_gap is not None and product.daily_gap >= 0):
            report.warnings.append("2.1 skipped: 产品达标")
            return

        daily_yield = self.reader.read_sheet(self.source_files["daily_yield"], "CT")
        rows = select_product_rows(daily_yield.rows, product.product_type)
        if not rows:
            report.warnings.append("未在 CT 源表中找到该产品。")
            return

        last_col = latest_daily_column(rows)
        if last_col is None:
            report.warnings.append("未找到 CT 源表的日度数据列。")
            return

        group_items: list[GapItem] = []
        for row in rows:
            group = _row_label(row)
            if group not in DEFECT_GROUPS:
                continue
            actual = _to_float(_get_cell(row, last_col))
            if actual is None:
                continue
            target = target_rates.get((product.product_type, group))
            if target is None:
                target = 0.0
            gap = actual - target
            if gap > 0:
                group_items.append(
                    GapItem(
                        defect_group=group,
                        actual_loss_rate=actual,
                        target_loss_rate=target,
                        gap=gap,
                        concentration_reason=CONCENTRATION_RULES.get(group, ""),
                    )
                )

        group_items.sort(key=lambda item: item.gap, reverse=True)
        report.gap_top_items = group_items[:3]
        report.known_exceptions = []

        defect_codes = top_defect_codes(rows, [item.defect_group for item in report.gap_top_items], last_col)
        if not defect_codes and defect_code_mapping:
            defect_codes = [
                code
                for code, metadata in defect_code_mapping.items()
                if metadata.get("group") in {item.defect_group for item in report.gap_top_items}
            ][:3]
        report.top_defect_codes = defect_codes

    def _fill_trend_section(self, report: ProductDailyReport) -> None:
        if "daily_yield" not in self.source_files:
            report.trend = TrendResult(checked=True, skipped_reason="缺少 daily_yield")
            return
        daily_yield = self.reader.read_sheet(self.source_files["daily_yield"], "CT")
        rows = select_product_rows(daily_yield.rows, report.product.product_type)
        if not rows:
            report.trend = TrendResult(checked=True, skipped_reason="未找到产品 CT 数据")
            return

        date_cols = latest_daily_columns(rows, count=3)
        if len(date_cols) < 3:
            report.trend = TrendResult(checked=True, skipped_reason="日度数据列不足三天")
            return

        ct_row = find_metric_row(rows, "CT良率")
        mvi_row = find_metric_row(rows, "CT产出数_MVI产出占比")
        if ct_row is None:
            report.trend = TrendResult(checked=True, skipped_reason="未找到 CT良率 行")
            return

        ct_values = [_to_float(_get_cell(ct_row, col)) for col in date_cols]
        mvi_values = [
            _to_float(_get_cell(mvi_row, col)) if mvi_row is not None else None
            for col in date_cols
        ]
        if any(value is None for value in ct_values):
            report.trend = TrendResult(checked=True, skipped_reason="CT良率三日数据不完整")
            return

        ct_numbers = [float(value) for value in ct_values if value is not None]
        mvi_numbers = [float(value) for value in mvi_values if value is not None]
        report.trend = TrendResult(
            checked=True,
            is_declining=_strictly_descending(ct_numbers),
            mvi_share_increasing=len(mvi_numbers) == 3 and _strictly_ascending(mvi_numbers),
            date_labels=[str(_get_cell(rows[0], col)) for col in date_cols],
            ct_yield_values=ct_numbers,
            mvi_share_values=mvi_numbers,
        )

    def _fill_exception_sections(self, report: ProductDailyReport) -> None:
        if "ct_exception" not in self.source_files:
            report.warnings.append("2.3/2.4 skipped: 缺少 ct_exception，使用 spotfire 异常文本。")
            return
        ct_exception = self.reader.read_sheet(self.source_files["ct_exception"], "CT异常波动调查")
        report_date = _parse_date(report.product.report_date)
        defect_codes = report.top_defect_codes
        report.known_exceptions = extract_exception_records(
            rows=ct_exception.rows,
            product_model=report.product.product_type,
            report_date=report_date,
            defect_codes=defect_codes,
            mode="known",
        )
        report.new_exceptions = extract_exception_records(
            rows=ct_exception.rows,
            product_model=report.product.product_type,
            report_date=report_date,
            defect_codes=None,
            mode="new",
        )

    def _compose_report_text(self, report: ProductDailyReport) -> str:
        deterministic = compose_deterministic_text(report)
        if not self.request.use_llm_polishing:
            return deterministic
        try:
            return polish_text_with_llm(report, deterministic)
        except Exception as exc:
            report.warnings.append(f"LLM 润色失败，已使用确定性文本: {exc}")
            return deterministic

    def _write_outputs(self, payload: DailyReportPayload) -> list[ArtifactRef]:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        output_name = self.request.output_name or f"良率日报_{payload.report_date}.xlsx"
        output_path = self.output_dir / _sanitize_filename(output_name)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        template = _resolve_template_path(self.request, self.workspace)
        ExcelReportWriter().write(payload, template, output_path, self.output_dir)

        artifacts = [
            ArtifactRef(
                kind="excel",
                path=output_path,
                description="良率日报 Excel 文件",
                metadata={"report_date": payload.report_date, "products": len(payload.products)},
            )
        ]

        if self.request.emit_intermediate_artifacts:
            json_path = output_path.with_suffix(".json")
            json_path.write_text(
                payload.model_dump_json(indent=2),
                encoding="utf-8",
            )
            markdown_path = output_path.with_suffix(".md")
            markdown_path.write_text(render_markdown(payload), encoding="utf-8")
            artifacts.extend(
                [
                    ArtifactRef(kind="json", path=json_path, description="日报结构化分析事实"),
                    ArtifactRef(kind="markdown", path=markdown_path, description="日报文本预览"),
                ]
            )
        return artifacts


class ExcelReportWriter:
    """Write the final report workbook."""

    def write(
        self,
        payload: DailyReportPayload,
        template_path: Path,
        output_path: Path,
        work_dir: Path,
    ) -> None:
        standard_template = ensure_standard_workbook(template_path, work_dir / "_normalized")
        if standard_template is None:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "sheet1"
            worksheet["C3"] = "产品类型"
            worksheet["D3"] = "产品"
            worksheet["E3"] = "日期"
            worksheet["F3"] = "当日异常_HTML"
            worksheet["G3"] = "当日异常"
        else:
            try:
                workbook = load_workbook(standard_template)
            except Exception:
                logger.info("Template could not be opened after normalization: %s", standard_template, exc_info=True)
                workbook = self._create_minimal_workbook()
            actual_sheet = _resolve_sheet_name(workbook.sheetnames, "sheet1")
            worksheet = workbook[actual_sheet or workbook.sheetnames[0]]

        self._write_payload_rows(worksheet, payload)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        workbook.close()
        if not _can_open_with_openpyxl(output_path):
            logger.info("Falling back to a minimal workbook because output was not readable: %s", output_path)
            fallback = self._create_minimal_workbook()
            self._write_payload_rows(fallback[fallback.sheetnames[0]], payload)
            fallback.save(output_path)
            fallback.close()

    @staticmethod
    def _create_minimal_workbook() -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "sheet1"
        worksheet["C3"] = "产品类型"
        worksheet["D3"] = "产品"
        worksheet["E3"] = "日期"
        worksheet["F3"] = "当日异常_HTML"
        worksheet["G3"] = "当日异常"
        return workbook

    def _write_payload_rows(self, worksheet: Any, payload: DailyReportPayload) -> None:
        self._clear_existing_data(worksheet, start_row=4)
        for index, product_report in enumerate(payload.products, start=4):
            worksheet.cell(index, 3, product_report.product.product_type)
            worksheet.cell(index, 4, product_report.product.product)
            worksheet.cell(index, 5, product_report.product.report_date)
            worksheet.cell(index, 6, _to_html(product_report.report_text))
            worksheet.cell(index, 7, product_report.report_text)
            worksheet.cell(index, 6).alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.cell(index, 7).alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _clear_existing_data(worksheet: Any, start_row: int) -> None:
        max_row = max(worksheet.max_row, start_row)
        for row in range(start_row, max_row + 1):
            for col in range(3, 8):
                worksheet.cell(row, col).value = None


def _product_report_from_analysis(item: dict[str, Any]) -> ProductDailyReport:
    product = ShippedProduct(**(item.get("product") or {}))
    sections = item.get("sections") or {}
    gap_facts = sections.get("gap", {}).get("facts", {})
    trend_facts = sections.get("trend", {}).get("facts", {}).get("trend", {})
    known_records = sections.get("known_exception", {}).get("facts", {}).get("records", [])
    new_records = sections.get("new_exception", {}).get("facts", {}).get("records", [])
    return ProductDailyReport(
        product=product,
        sections=sections,
        gap_top_items=[GapItem(**record) for record in gap_facts.get("gap_top_items", [])],
        top_defect_codes=list(gap_facts.get("top_defect_codes", [])),
        trend=TrendResult(**trend_facts) if trend_facts else TrendResult(),
        known_exceptions=[ExceptionRecord(**record) for record in known_records],
        new_exceptions=[ExceptionRecord(**record) for record in new_records],
        report_text=item.get("report_text") or "无",
        warnings=list(item.get("warnings") or []),
    )


def execute_daily_report(
    request: DailyReportRequest,
    context: RunContext | None = None,
) -> SkillResult:
    """Generate the daily report and return a structured SkillResult."""
    context = context or RunContext(run_id="manual-run", workspace=Path.cwd())
    try:
        return DailyReportGenerator(request, context).run()
    except DailyReportGenerationError as exc:
        logger.error(
            "Daily report generation failed: %s",
            exc.message,
            extra={
                "event": "failure",
                "purpose": "business",
                "run_id": context.run_id,
                "task_id": TOOL_NAME,
                "error_code": exc.code,
            },
        )
        return SkillResult(
            skill_name=TOOL_NAME,
            success=False,
            summary=f"日报生成失败: {exc.message}",
            data={
                "report_date": request.report_date,
                "sections": request.sections,
                "output_name": request.output_name,
            },
            error=SkillError(
                code=exc.code,
                message=exc.message,
                recoverable=exc.recoverable,
                details=exc.details,
            ),
        )
    except Exception as exc:
        logger.exception(
            "Daily report generation failed",
            extra={
                "event": "failure",
                "purpose": "business",
                "run_id": context.run_id,
                "task_id": TOOL_NAME,
                "error_code": "daily_report.execution.failed",
            },
        )
        return SkillResult(
            skill_name=TOOL_NAME,
            success=False,
            summary=f"日报生成失败: {exc}",
            error=SkillError(
                code="daily_report.execution.failed",
                message=str(exc),
                recoverable=True,
            ),
        )


def extract_shipped_products(rows: list[list[Any]]) -> list[ShippedProduct]:
    header_index = _find_header_row(rows, ["产品类型", "产品", "日期", "当日异常"])
    if header_index is None:
        raise DailyReportGenerationError(
            "daily_report.spotfire.missing_header",
            "spotfire Sheet1 缺少产品类型/产品/日期/当日GAP 表头。",
        )
    headers = [_norm_header(value) for value in rows[header_index]]
    col_product_type = _find_col(headers, ["产品类型", "producttype"])
    col_product = _find_col(headers, ["产品", "product"])
    col_date = _find_col(headers, ["日期", "date"])
    col_target = _find_col(headers, ["目标", "target"])
    col_yield = _find_col(headers, ["良率", "yield"])
    col_qualified = _find_col(headers, ["是否达标"])
    col_gap = _find_col(headers, ["当日gap", "当日gap"])
    col_daily_exception = _find_col(headers, ["当日异常"])
    col_known_exception = _find_col(headers, ["已hl异常", "已hl"])

    products: list[ShippedProduct] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        product_type = _string_cell(_get_cell(row, col_product_type))
        product = _string_cell(_get_cell(row, col_product))
        if not product_type and not product:
            continue
        products.append(
            ShippedProduct(
                product_type=product_type or product,
                product=product,
                report_date=_normalize_date(_string_cell(_get_cell(row, col_date))),
                target_yield=_to_float(_get_cell(row, col_target)),
                actual_yield=_to_float(_get_cell(row, col_yield)),
                is_qualified=_parse_qualified(_get_cell(row, col_qualified)),
                daily_gap=_to_float(_get_cell(row, col_gap)),
                existing_daily_exception=_string_cell(_get_cell(row, col_daily_exception)),
                existing_known_exception=_string_cell(_get_cell(row, col_known_exception)),
                source_row=row_number,
            )
        )
    return products


def extract_target_rates(rows: list[list[Any]], month_label: str) -> dict[tuple[str, str], float]:
    header_index = _find_header_row(rows, ["项目代号", "Group"])
    if header_index is None:
        return {}
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["项目代号"])
    group_col = _find_col(headers, ["group"])
    month_col = _find_col(headers, [_norm_header(month_label)])
    if product_col is None or group_col is None or month_col is None:
        return {}

    rates: dict[tuple[str, str], float] = {}
    for row in rows[header_index + 1 :]:
        product = _string_cell(_get_cell(row, product_col))
        group = _string_cell(_get_cell(row, group_col))
        value = _to_float(_get_cell(row, month_col))
        if product and group and value is not None:
            rates[(product, group)] = value
    return rates


def extract_code_mapping(rows: list[list[Any]]) -> dict[str, dict[str, str]]:
    header_index = _find_header_row(rows, ["DefectDesc", "GroupType"])
    if header_index is None:
        return {}
    headers = [_norm_header(value) for value in rows[header_index]]
    desc_col = _find_col(headers, ["defectdesc", "代码描述"])
    group_col = _find_col(headers, ["grouptype", "组别"])
    factory_col = _find_col(headers, ["factory", "归属工厂"])
    mapping: dict[str, dict[str, str]] = {}
    for row in rows[header_index + 1 :]:
        desc = _string_cell(_get_cell(row, desc_col))
        if not desc:
            continue
        mapping[desc] = {
            "group": _string_cell(_get_cell(row, group_col)),
            "factory": _string_cell(_get_cell(row, factory_col)),
        }
    return mapping


def select_product_rows(rows: list[list[Any]], product_model: str) -> list[list[Any]]:
    header_index = _find_header_row(rows, ["ProductCode", "Operation"])
    if header_index is None:
        return []
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["productcode", "产品型号"])
    if product_col is None:
        return []

    selected: list[list[Any]] = [rows[header_index]]
    current_product = ""
    for row in rows[header_index + 1 :]:
        product_value = _string_cell(_get_cell(row, product_col))
        if product_value:
            current_product = product_value
        if current_product == product_model:
            selected.append(row)
    return selected


def latest_daily_column(rows: list[list[Any]]) -> int | None:
    columns = latest_daily_columns(rows, count=1)
    return columns[-1] if columns else None


def latest_daily_columns(rows: list[list[Any]], count: int) -> list[int]:
    if not rows:
        return []
    header = rows[0]
    candidates = [
        index
        for index, value in enumerate(header)
        if _is_daily_label(_string_cell(value))
    ]
    return candidates[-count:]


def find_metric_row(rows: list[list[Any]], metric_name: str) -> list[Any] | None:
    wanted = _norm_header(metric_name)
    for row in rows[1:]:
        if _norm_header(_row_label(row)) == wanted:
            return row
    return None


def top_defect_codes(
    rows: list[list[Any]],
    defect_groups: list[str],
    last_col: int,
    limit: int = 3,
) -> list[str]:
    if not rows:
        return []
    headers = [_norm_header(value) for value in rows[0]]
    group_col = _find_col(headers, ["defectgroup", "不良分组"])
    code_col = (group_col + 1) if group_col is not None else None
    group_set = set(defect_groups)
    scored: list[tuple[float, str]] = []

    current_group = ""
    for row in rows[1:]:
        explicit_group = _string_cell(_get_cell(row, group_col))
        row_label = _row_label(row)
        if explicit_group:
            current_group = explicit_group
        elif row_label in group_set:
            current_group = row_label
        if current_group not in group_set:
            continue
        code = _string_cell(_get_cell(row, code_col))
        if not code or code in DEFECT_GROUPS or code == current_group:
            continue
        value = _to_float(_get_cell(row, last_col))
        if value is None:
            continue
        scored.append((value, code))

    scored.sort(key=lambda item: item[0], reverse=True)
    return [code for _, code in scored[:limit]]


def extract_exception_records(
    *,
    rows: list[list[Any]],
    product_model: str,
    report_date: date,
    defect_codes: list[str] | None,
    mode: str,
) -> list[ExceptionRecord]:
    if mode == "known" and defect_codes is not None and not defect_codes:
        return []

    header_index = _find_header_row(rows, ["产品", "不良", "通报日期"])
    if header_index is None:
        return []
    headers = [_norm_header(value) for value in rows[header_index]]
    product_col = _find_col(headers, ["产品"])
    defect_col = _find_col(headers, ["不良"])
    date_col = _find_col(headers, ["通报日期"])
    reply_col = _find_col(headers, ["工艺整合&工艺回复的改善及挽救进展", "工艺整合"])
    daily_col = _find_col(headers, ["日良损"])
    month_col = _find_col(headers, ["当月"])
    week_col = _find_col(headers, ["当周"])
    batch_col = _find_col(headers, ["本批次"])

    if product_col is None or defect_col is None or date_col is None:
        return []

    defect_set = set(defect_codes or [])
    start_date = report_date - timedelta(days=30)
    records: list[ExceptionRecord] = []
    for row in rows[header_index + 1 :]:
        product = _string_cell(_get_cell(row, product_col))
        if product != product_model:
            continue
        defect = _string_cell(_get_cell(row, defect_col))
        if defect_set and defect not in defect_set:
            continue
        record_date = _parse_date_value(_get_cell(row, date_col))
        if record_date is None:
            continue
        if mode == "known" and not (start_date <= record_date <= report_date):
            continue
        if mode == "new" and record_date != report_date:
            continue

        reply = _string_cell(_get_cell(row, reply_col))
        fields = _extract_reply_fields(reply)
        records.append(
            ExceptionRecord(
                product_model=product,
                defect_code=defect,
                report_datetime=_string_cell(_get_cell(row, date_col)),
                daily_loss=_format_loss(_get_cell(row, daily_col)),
                monthly_loss=_format_loss(_get_cell(row, month_col)),
                weekly_loss=_format_loss(_get_cell(row, week_col)),
                batch_loss=_format_loss(_get_cell(row, batch_col)),
                raw_reply=reply,
                **fields,
            )
        )
    return records


def compose_deterministic_text(report: ProductDailyReport) -> str:
    lines: list[str] = []
    if report.gap_top_items:
        total_actual = sum(item.actual_loss_rate for item in report.gap_top_items)
        total_gap = sum(item.gap for item in report.gap_top_items)
        reasons = "、".join(
            sorted({item.concentration_reason for item in report.gap_top_items if item.concentration_reason})
        )
        groups = "、".join(item.defect_group for item in report.gap_top_items)
        line = (
            f"1.1 {groups} total不良率{_percent(total_actual)}，"
            f"GAP {_percent(total_gap)}"
        )
        if reasons:
            line += f"；当天Gap主要{reasons}"
        lines.append(line)
    elif report.product.existing_known_exception:
        lines.append(report.product.existing_known_exception)

    trend = report.trend
    if trend.is_declining:
        label_range = "~".join([trend.date_labels[0], trend.date_labels[-1]]) if trend.date_labels else "近三天"
        if trend.mvi_share_increasing:
            line = (
                f"1.2 {label_range} CT良率持续下降原因："
                f"MVI产出占比提升({_percent_chain(trend.mvi_share_values)})"
            )
        else:
            line = f"1.2 {label_range} CT良率持续下降，未识别到MVI产出占比连续升高证据"
        lines.append(line)

    for record in report.known_exceptions:
        lines.append(_format_exception_record("1.3", record))

    for record in report.new_exceptions:
        lines.append(_format_exception_record("【异常】", record))

    if report.product.existing_daily_exception:
        lines.append(report.product.existing_daily_exception)

    if not lines:
        lines.append("无")
    return "\n\n".join(lines)


def polish_text_with_llm(report: ProductDailyReport, fallback_text: str) -> str:
    from shared_kernel.infrastructure.llm_handler import llm_manager

    prompt = (
        "请将以下结构化日报事实润色为良率日报段落，保持编号和【字段】格式，"
        "不要编造未提供的事实。\n\n"
        f"产品: {report.product.product_type}\n"
        f"事实JSON:\n{report.model_dump_json(indent=2)}\n\n"
        f"当前文本:\n{fallback_text}"
    )
    text = llm_manager.chat(
        provider="deepseek",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
        max_tokens=2048,
    )
    return text.strip() or fallback_text


def render_markdown(payload: DailyReportPayload) -> str:
    lines = [f"# 良率日报 {payload.report_date}", ""]
    for product_report in payload.products:
        lines.extend(
            [
                f"## {product_report.product.product_type}",
                "",
                product_report.report_text,
                "",
            ]
        )
    if payload.warnings:
        lines.extend(["## Warnings", ""])
        lines.extend([f"- {warning}" for warning in payload.warnings])
        lines.append("")
    return "\n".join(lines)


def ensure_standard_workbook(path: Path, output_dir: Path) -> Path | None:
    path = Path(path)
    if not path.exists():
        return None
    if _can_open_with_openpyxl(path):
        return path

    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        decrypted = decrypt_excel_file(path, output_dir)
        if _can_open_with_openpyxl(decrypted):
            return decrypted
    except Exception:
        logger.info("Project decryption wrapper failed for template: %s", path, exc_info=True)

    return _save_as_xlsx_with_com(path, output_dir)


def _save_as_xlsx_with_com(path: Path, output_dir: Path) -> Path | None:
    try:
        import win32com.client  # type: ignore[import-untyped]
    except ImportError:
        return None

    output_path = output_dir / f"{path.stem}.xlsx"
    excel = None
    workbook = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for attr, value in (
            ("DisplayAlerts", False),
            ("AskToUpdateLinks", False),
            ("EnableEvents", False),
        ):
            try:
                setattr(excel, attr, value)
            except Exception:
                pass
        workbook = excel.Workbooks.Open(str(path.resolve()), 0, True)
        if output_path.exists():
            output_path.unlink()
        workbook.SaveAs(str(output_path.resolve()), FileFormat=51)
        workbook.Close(False)
        workbook = None
        return output_path if _can_open_with_openpyxl(output_path) else None
    except Exception:
        logger.info("Excel COM SaveAs failed for template: %s", path, exc_info=True)
        return None
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _resolve_output_dir(request: DailyReportRequest, context: RunContext) -> Path:
    output_dir = request.output_dir or context.output_dir
    output_dir = Path(output_dir)
    if not output_dir.is_absolute():
        output_dir = context.workspace / output_dir
    return output_dir


def _resolve_template_path(request: DailyReportRequest, workspace: Path) -> Path:
    if request.template_ref is not None:
        return _resolve_path(request.template_ref, workspace)
    return workspace / DEFAULT_TEMPLATE


def _resolve_path(path: Path, workspace: Path) -> Path:
    path = Path(path)
    return path if path.is_absolute() else workspace / path


def _find_first_file(directory: Path, patterns: list[str]) -> Path | None:
    if not directory.exists():
        return None
    for pattern in patterns:
        for path in sorted(directory.glob(pattern)):
            if path.is_file() and not path.name.startswith("~$"):
                return path
    return None


def _find_header_row(rows: list[list[Any]], required_headers: list[str]) -> int | None:
    required = [_norm_header(item) for item in required_headers]
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        normalized = [_norm_header(value) for value in row]
        if all(any(req in cell for cell in normalized) for req in required):
            return index
    return None


def _find_col(headers: list[str], candidates: list[str]) -> int | None:
    normalized_candidates = [_norm_header(item) for item in candidates]
    for candidate in normalized_candidates:
        for index, header in enumerate(headers):
            if candidate and candidate == header:
                return index
    for candidate in normalized_candidates:
        for index, header in enumerate(headers):
            if candidate and candidate in header:
                return index
    return None


def _resolve_sheet_name(sheet_names: list[str], requested: str) -> str | None:
    if requested in sheet_names:
        return requested
    requested_norm = requested.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() == requested_norm:
            return sheet_name
    return None


def _com_values_to_rows(raw_values: Any) -> list[list[Any]]:
    if raw_values is None:
        return []
    if not isinstance(raw_values, tuple):
        return [[raw_values]]
    if raw_values and not isinstance(raw_values[0], tuple):
        return [list(raw_values)]
    return [list(row) for row in raw_values]


def _get_cell(row: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _row_label(row: list[Any]) -> str:
    for index in [3, 4]:
        value = _string_cell(_get_cell(row, index))
        if value:
            return value
    return ""


def _string_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if "%" in str(value) else number


def _parse_qualified(value: Any) -> bool | None:
    text = _string_cell(value)
    if not text:
        return None
    if "不达标" in text:
        return False
    if "达标" in text:
        return True
    return None


def _normalize_date(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        return date.today().isoformat()
    parsed = _parse_date_value(value)
    if parsed is not None:
        return parsed.isoformat()
    compact = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(?:\.0)?", value)
    if compact:
        return f"{compact.group(1)}-{compact.group(2)}-{compact.group(3)}"
    return value


def _parse_date(value: str) -> date:
    parsed = _parse_date_value(value)
    if parsed is None:
        raise DailyReportGenerationError(
            "daily_report.input.invalid_date",
            f"无法解析日期: {value}",
        )
    return parsed


def _parse_date_value(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    compact = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(?:\.0)?", text)
    if compact:
        return date(int(compact.group(1)), int(compact.group(2)), int(compact.group(3)))
    text = text.replace("/", "-")
    if text.endswith("+00:00"):
        text = text[:-6]
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _month_label(report_date: str) -> str:
    parsed = _parse_date(report_date)
    return f"{parsed.month}月"


def _is_daily_label(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}", value.strip()))


def _matches_product_filter(product: ShippedProduct, product_filter: list[str]) -> bool:
    haystack = f"{product.product_type}\n{product.product}"
    return any(item in haystack for item in product_filter)


def _strictly_descending(values: list[float]) -> bool:
    return len(values) >= 2 and all(left > right for left, right in zip(values, values[1:]))


def _strictly_ascending(values: list[float]) -> bool:
    return len(values) >= 2 and all(left < right for left, right in zip(values, values[1:]))


def _extract_reply_fields(text: str) -> dict[str, str]:
    return {
        "exception_reason": _extract_tag(text, "异常原因"),
        "inline_monitoring": _extract_tag(text, "Inline监控"),
        "is_stopped": _extract_tag(text, "是否止血"),
        "impact_scope": _extract_tag(text, "影响范围"),
        "improvement_measures": _extract_tag(text, "改善措施"),
    }


def _extract_tag(text: str, tag: str) -> str:
    pattern = rf"【{re.escape(tag)}】([\s\S]*?)(?=【[^】]+】|$)"
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _format_exception_record(prefix: str, record: ExceptionRecord) -> str:
    title = f"{prefix} {record.defect_code}"
    if record.daily_loss:
        title += f" {record.daily_loss}"
    if record.monthly_loss:
        title += f"(当月 {record.monthly_loss})"
    fields = [
        title,
        f"【异常原因】{record.exception_reason or '/'}",
        f"【Inline监控】{record.inline_monitoring or '/'}",
        f"【是否止血】{record.is_stopped or '/'}",
        f"【影响范围】{record.impact_scope or '/'}",
        f"【改善措施】{record.improvement_measures or '/'}",
    ]
    return "\n".join(fields)


def _format_loss(value: Any) -> str:
    number = _to_float(value)
    if number is None:
        return _string_cell(value)
    return _percent(number)


def _percent(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value * 100:.2f}%"


def _percent_chain(values: list[float]) -> str:
    return "→".join(_percent(value) for value in values)


def _to_html(text: str) -> str:
    return html.escape(text).replace("\n", "<br/>")


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", _string_cell(value).lower())


def _is_standard_xlsx(path: Path) -> bool:
    try:
        with Path(path).open("rb") as file:
            return file.read(4) == XLSX_MAGIC
    except OSError:
        return False


def _can_open_with_openpyxl(path: Path) -> bool:
    if not _is_standard_xlsx(path):
        return False
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        return bool(workbook.sheetnames)
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.close()


def _sanitize_filename(filename: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", filename)
